"""Export authored data from a live workbook into a typed snapshot.

The live workbook is only ever opened for reading; every write happens later,
in-process, during a fresh build. Values are read through openpyxl with
``data_only=True``: authored cells are literals, so no cached-formula gaps can
occur, and dates come back as datetimes that must be pure dates.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import UTC, date, datetime
from enum import Enum
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.utils.cell import range_boundaries

from .schema import (
    DATA_TABLES,
    EXAMPLE_MARKER,
    SETTINGS_TYPES,
    key_problems,
    schema_fingerprint,
)
from .snapshot import Snapshot

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from .schema import ColumnSchema, TableSchema


class _ExportProblem(Enum):
    MISSING_SHEET = "workbook {} has no {} sheet"
    MISSING_TABLE = "sheet {} has no table {}"
    BLANK_KEY = "{} row {} holds data but no {}; assign one in the workbook and re-run"
    DUPLICATE_KEY = "{} holds duplicate {} values: {}"
    TIME_OF_DAY = "{}[{}] row {} holds a time of day ({}); authored dates must be pure dates"
    CELL_TYPE = "{}[{}] row {} holds {!r}; expected {}"
    SETTINGS_BAND = "Config sheet has no Setting/Value band"
    SETTING_VALUE = "Config setting {} holds {!r}; expected {}"


class ExportError(RuntimeError):
    """Report unreadable or invalid authored data during export."""

    def __init__(self, problem: _ExportProblem, *details: object) -> None:
        """Create an error from a stable diagnostic template."""
        super().__init__(problem.value.format(*details))


@dataclass(frozen=True, kw_only=True, slots=True)
class ExportResult:
    """One exported snapshot with its per-table reconciliation notes."""

    snapshot: Snapshot
    skipped_examples: dict[str, int]
    added_columns: dict[str, tuple[str, ...]]
    unknown_columns: dict[str, tuple[str, ...]]
    notes: tuple[str, ...]


def read_table(worksheet: Worksheet, table_name: str) -> tuple[list[str], list[list[object]]]:
    """Read one workbook table's headers and body values.

    Returns:
        The header names and every body row's cell values.

    Raises:
        ExportError: If the worksheet has no such table.

    """
    if table_name not in worksheet.tables:
        raise ExportError(_ExportProblem.MISSING_TABLE, worksheet.title, table_name)
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [str(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
    rows = [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row + 1, max_row + 1)
    ]
    return headers, rows


def _coerce_date(table: str, column: str, row_index: int, value: object) -> date:
    if isinstance(value, datetime):
        if value.time() != datetime.min.time():
            raise ExportError(_ExportProblem.TIME_OF_DAY, table, column, row_index, value.time())
        return value.date()
    if isinstance(value, date):
        return value
    raise ExportError(_ExportProblem.CELL_TYPE, table, column, row_index, value, "a date")


def _coerce_int(table: str, column: str, row_index: int, value: object) -> int:
    if isinstance(value, bool):
        raise ExportError(_ExportProblem.CELL_TYPE, table, column, row_index, value, "an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    raise ExportError(_ExportProblem.CELL_TYPE, table, column, row_index, value, "an integer")


def _coerce_text(
    table: str,
    column: str,
    row_index: int,
    value: object,
    notes: list[str],
) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, (bool, date, datetime)):
        raise ExportError(_ExportProblem.CELL_TYPE, table, column, row_index, value, "text")
    if isinstance(value, (int, float)):
        text = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
        notes.append(f"{table}[{column}] row {row_index}: numeric {value!r} exported as text")
        return text
    raise ExportError(_ExportProblem.CELL_TYPE, table, column, row_index, value, "text")


def _coerce_cell(
    table: str,
    column: ColumnSchema,
    row_index: int,
    value: object,
    notes: list[str],
) -> object:
    if value is None or (isinstance(value, str) and not value):
        return None
    if column.value_type == "date":
        return _coerce_date(table, column.name, row_index, value)
    if column.value_type == "bool":
        if isinstance(value, bool):
            return value
        raise ExportError(
            _ExportProblem.CELL_TYPE,
            table,
            column.name,
            row_index,
            value,
            "a boolean",
        )
    if column.value_type == "int":
        return _coerce_int(table, column.name, row_index, value)
    return _coerce_text(table, column.name, row_index, value, notes)


def _split_headers(
    table_schema: TableSchema,
    headers: list[str],
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    known = set(table_schema.workbook_columns)
    unknown = tuple(header for header in headers if header not in known)
    present = set(headers)
    added = tuple(name for name in table_schema.column_names if name not in present)
    return added, unknown


def _row_is_example(row: dict[str, object]) -> bool:
    return any(isinstance(value, str) and EXAMPLE_MARKER in value for value in row.values())


def _require_keys(table_schema: TableSchema, rows: list[dict[str, object]]) -> None:
    blank_row, duplicates = key_problems(table_schema.key, tuple(rows))
    if blank_row is not None:
        raise ExportError(
            _ExportProblem.BLANK_KEY,
            table_schema.table,
            blank_row,
            table_schema.key,
        )
    if duplicates:
        raise ExportError(
            _ExportProblem.DUPLICATE_KEY,
            table_schema.table,
            table_schema.key,
            ", ".join(duplicates),
        )


class _FormulaProbe:
    """Classify unknown workbook columns as formula or authored on demand."""

    def __init__(self, path: Path) -> None:
        self._path = path
        self._workbook: Workbook | None = None
        self._tables: dict[tuple[str, str], tuple[list[str], list[list[object]]]] = {}

    def _table_values(
        self,
        table_schema: TableSchema,
    ) -> tuple[list[str], list[list[object]]]:
        key = (table_schema.sheet, table_schema.table)
        if key not in self._tables:
            if self._workbook is None:
                self._workbook = openpyxl.load_workbook(self._path, data_only=False)
            if table_schema.sheet not in self._workbook.sheetnames:
                raise ExportError(
                    _ExportProblem.MISSING_SHEET,
                    self._path.name,
                    table_schema.sheet,
                )
            self._tables[key] = read_table(
                self._workbook[table_schema.sheet],
                table_schema.table,
            )
        return self._tables[key]

    def is_formula_column(self, table_schema: TableSchema, header: str) -> bool:
        """Return whether one unknown column stores formulas.

        Returns:
            True when any of the column's body cells holds a formula.

        """
        headers, rows = self._table_values(table_schema)
        column = headers.index(header)
        return any(
            isinstance(row[column], str) and str(row[column]).startswith("=") for row in rows
        )


def _export_table(
    worksheet: Worksheet,
    table_schema: TableSchema,
    probe: _FormulaProbe,
    result_notes: list[str],
) -> tuple[tuple[dict[str, object], ...], int, tuple[str, ...], tuple[str, ...]]:
    headers, body = read_table(worksheet, table_schema.table)
    added, unknown = _split_headers(table_schema, headers)

    authored_unknown = tuple(
        header for header in unknown if not probe.is_formula_column(table_schema, header)
    )
    for header in unknown:
        if header in authored_unknown:
            result_notes.append(
                f"{table_schema.table}[{header}] is not in the current schema; "
                "its values are exported and must be resolved before injection"
            )
        else:
            result_notes.append(
                f"{table_schema.table}[{header}] is an unknown formula column; not exported"
            )

    column_map = {column.name: column for column in table_schema.columns}
    exported_names = [
        header for header in headers if header in column_map or header in authored_unknown
    ]
    rows: list[dict[str, object]] = []
    skipped = 0
    for index, body_row in enumerate(body, start=1):
        values = dict(zip(headers, body_row, strict=True))
        row: dict[str, object] = {}
        for name in exported_names:
            column_schema = column_map.get(name)
            value = (
                _coerce_cell(table_schema.table, column_schema, index, values[name], result_notes)
                if column_schema is not None
                else values[name]
            )
            if value is not None:
                row[name] = value
        if not row:
            continue
        if _row_is_example(row):
            skipped += 1
            continue
        if any(isinstance(value, str) and "EXAMPLE" in value.upper() for value in row.values()):
            label = row.get(table_schema.key or "", f"row {index}")
            result_notes.append(
                f"{table_schema.table} {label}: contains example-styled text and was retained"
            )
        rows.append(row)

    _require_keys(table_schema, rows)
    return tuple(rows), skipped, added, authored_unknown


def _find_settings_band(worksheet: Worksheet) -> int:
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == "Setting" and worksheet.cell(row, 2).value == "Value":
            return row
    raise ExportError(_ExportProblem.SETTINGS_BAND)


def _coerce_setting(name: str, value: object) -> object:
    value_type = SETTINGS_TYPES.get(name)
    if value_type is None:
        return value
    if value_type == "int":
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ExportError(_ExportProblem.SETTING_VALUE, name, value, "an integer")
        if isinstance(value, float):
            if not value.is_integer():
                raise ExportError(_ExportProblem.SETTING_VALUE, name, value, "an integer")
            return int(value)
        return value
    if not isinstance(value, str) or not value:
        raise ExportError(_ExportProblem.SETTING_VALUE, name, value, "text")
    return value


def export_settings(worksheet: Worksheet) -> dict[str, object]:
    """Read the Config settings band into name-value pairs.

    Returns:
        Every setting label mapped to its typed value.

    """
    header_row = _find_settings_band(worksheet)
    settings: dict[str, object] = {}
    for row in range(header_row + 1, worksheet.max_row + 2):
        label = worksheet.cell(row, 1).value
        if label in {None, ""}:
            break
        name = f"cfg{label}"
        settings[name] = _coerce_setting(name, worksheet.cell(row, 2).value)
    return settings


def export_workbook(path: Path) -> ExportResult:
    """Export every authored table row and Config setting from one workbook.

    Returns:
        The typed snapshot and its reconciliation notes.

    Raises:
        ExportError: If the workbook structure or authored values are invalid.

    """
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = openpyxl.load_workbook(path, data_only=True)
    probe = _FormulaProbe(path)

    notes: list[str] = []
    tables: dict[str, tuple[dict[str, object], ...]] = {}
    skipped_examples: dict[str, int] = {}
    added_columns: dict[str, tuple[str, ...]] = {}
    unknown_columns: dict[str, tuple[str, ...]] = {}
    for table_schema in DATA_TABLES:
        if table_schema.sheet not in workbook.sheetnames:
            raise ExportError(_ExportProblem.MISSING_SHEET, path.name, table_schema.sheet)
        rows, skipped, added, unknown = _export_table(
            workbook[table_schema.sheet],
            table_schema,
            probe,
            notes,
        )
        tables[table_schema.table] = rows
        if skipped:
            skipped_examples[table_schema.table] = skipped
        if added:
            added_columns[table_schema.table] = added
        if unknown:
            unknown_columns[table_schema.table] = unknown

    if "Config" not in workbook.sheetnames:
        raise ExportError(_ExportProblem.MISSING_SHEET, path.name, "Config")
    settings = export_settings(workbook["Config"])

    snapshot = Snapshot(
        schema_fingerprint=schema_fingerprint(),
        workbook_digest=digest,
        exported_at=datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
        settings=settings,
        tables=tables,
    )
    return ExportResult(
        snapshot=snapshot,
        skipped_examples=skipped_examples,
        added_columns=added_columns,
        unknown_columns=unknown_columns,
        notes=tuple(notes),
    )
