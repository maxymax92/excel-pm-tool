"""Exercise formula guards with malformed hierarchy and dependency data."""

import logging
from datetime import date

import openpyxl

from ..pipeline import build_one
from ..spec.config import TYPES
from .common import ERROR_VALUE_RE, temporary_examples, temporary_workbook, workbook_error_cells
from .excel import recalculate

LOGGER = logging.getLogger(__name__)
MIN_HIERARCHY_LEVEL = 0
MAX_HIERARCHY_LEVEL = max(level for _item_type, level in TYPES)


def main() -> int:
    """Run malformed-data QA and report each observable failure.

    Returns:
        A process exit status: zero for success and one for failure.

    """
    fails: list[str] = []
    with temporary_examples() as examples, temporary_workbook("PM_abuse") as out:
        examples.ITEMS_EXAMPLES = [
            {
                "ID": "I-1001",
                "Title": "Circular A",
                "Type": "Epic",
                "Status": "In Progress",
                "Parent": "I-1002",
                "Owner": "Max",
            },
            {
                "ID": "I-1002",
                "Title": "Circular B",
                "Type": "Epic",
                "Status": "In Progress",
                "Parent": "I-1001",
                "Owner": "Max",
            },
            {
                "ID": "I-1003",
                "Title": "Orphan (bad parent)",
                "Type": "Story",
                "Status": "Ready",
                "Parent": "I-9999",
                "Owner": "Max",
            },
            {
                "ID": "I-1004",
                "Title": "Self parent",
                "Type": "Task",
                "Status": "Backlog",
                "Parent": "I-1004",
                "Owner": "Max",
            },
            {
                "ID": "I-1005",
                "Title": "He said \"hi\", <b> & 'x' — 日本語 " + "y" * 200,
                "Type": "Story",
                "Status": "Review",
                "Owner": "Max",
                "Due": date(2020, 1, 1),
            },
            {
                "ID": "I-1006",
                "Title": "Dup one",
                "Type": "Story",
                "Status": "Done",
                "Owner": "Max",
                "DoneDate": date(2026, 6, 1),
            },
            {
                "ID": "I-1006",
                "Title": "Dup two",
                "Type": "Story",
                "Status": "In Progress",
                "Owner": "Max",
            },
            {
                "ID": "I-1007",
                "Title": "Tangled deps",
                "Type": "Story",
                "Status": "In Progress",
                "Parent": "I-1001",
                "BlockedBy": "I-1007, I-9999, I-1001",
                "Owner": "Max",
                "Start": date(2024, 1, 1),
                "Due": date(2024, 6, 1),
            },
            {
                "ID": "I-1008",
                "Title": "Weird numbers",
                "Type": "Story",
                "Status": "In Progress",
                "Owner": "Nobody",
                "Delivery Health": "Blocked",
            },
        ]
        examples.PEOPLE_EXAMPLES = [
            {"Person": "Max", "Role": "PM", "Team": "Core"},
        ]
        examples.RAID_EXAMPLES = [
            {
                "RaidID": "R-001",
                "Type": "Risk",
                "Title": 'Risk "q" & <html>',
                "RelatedID": "I-9999",
                "Owner": "Ghost",
                "Status": "Open",
                "Prob": 5,
                "Impact": 5,
            },
        ]
        build_one(out, with_vba=False)
        recalculate(out)

        workbook = openpyxl.load_workbook(out, data_only=True)
        try:
            fails.extend(f"error cell {value}" for value in workbook_error_cells(workbook))
            items = workbook["Items"]
            columns = {cell.value: index + 1 for index, cell in enumerate(items[2])}
            for row_index in range(3, 3 + len(examples.ITEMS_EXAMPLES)):
                level = items.cell(row=row_index, column=columns["Level"]).value
                if not (
                    isinstance(level, int) and MIN_HIERARCHY_LEVEL <= level <= MAX_HIERARCHY_LEVEL
                ):
                    fails.append(f"Items row {row_index} Level out of bounds: {level!r}")
                for column in ("Scope", "WbsKey", "EffStart", "EffDue", "WaitingOn"):
                    value = items.cell(row=row_index, column=columns[column]).value
                    if isinstance(value, str) and ERROR_VALUE_RE.match(value):
                        fails.append(f"Items row {row_index} {column} errored: {value!r}")
        finally:
            workbook.close()

    if fails:
        LOGGER.error("ABUSE QA: %s FAILURE(S)", len(fails))
        for failure in fails[:40]:
            LOGGER.error("  FAIL %s", failure)
        return 1
    LOGGER.info("ABUSE QA: ALL PASS")
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    raise SystemExit(main())
