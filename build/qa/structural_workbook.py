"""Workbook schema, surface, validation, protection, and layout checks."""

from __future__ import annotations

import re
from typing import TYPE_CHECKING

from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .structural_shared import (
    EXPECTED_SHEETS,
    EXPECTED_TABLES,
    ITEMS_CORE,
    ITEMS_DATA_ROW,
    ITEMS_HDR_ROW,
    PEOPLE_HEADERS,
    PROTECTED_SHEETS,
    RAID_CORE,
    RAID_SYSTEM,
    column_dimension_for,
    grouped_column_failures,
    normalise_number_format,
    table_headers,
    validation_for,
)

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook

MIN_ITEMS_VALIDATED_COLUMNS = 10
MIN_ITEMS_CF_RANGES = 6
EXPECTED_ITEM_OUTLINES = (0, 3, 4, 1)
DATE_NUMBER_FORMAT = "dd mmm yyyy"
DATE_TARGETS = {
    "Items": (
        "Start",
        "Due",
        "Created",
        "Updated",
        "InProgressSince",
        "DoneDate",
        "BlockedSince",
        "LatestUpdateOn",
    ),
    "RAID": ("NextReview", "Raised", "Closed", "Updated"),
}
DATE_TABLES = {"Items": "tblItems", "RAID": "tblRAID"}


def _inventory_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    if workbook.sheetnames != EXPECTED_SHEETS:
        failures.append(f"sheet order mismatch: {workbook.sheetnames}")
    for sheet, expected in EXPECTED_TABLES.items():
        if sheet not in workbook.sheetnames:
            continue
        actual = set(workbook[sheet].tables)
        wanted = set(expected)
        if actual != wanted:
            failures.append(f"tables on {sheet}: got {sorted(actual)}, want {sorted(wanted)}")
    return failures


def _required_objects_present(workbook: Workbook) -> bool:
    if not set(EXPECTED_SHEETS).issubset(workbook.sheetnames):
        return False
    return all(
        set(expected).issubset(workbook[sheet].tables)
        for sheet, expected in EXPECTED_TABLES.items()
    )


def _overview_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Overview"]
    failures: list[str] = []
    if workbook.active.title != "Overview":
        failures.append(f"active sheet is {workbook.active.title!r}, want 'Overview'")
    if worksheet["A1"].value != "Executive Status Summary":
        failures.append(f"Overview does not start at the panels: {worksheet['A1'].value!r}")

    titles = {
        "A1": "Executive Status Summary",
        "F1": "Top RAID",
        "M1": "Coming up",
        "Q1": "Recent progress",
    }
    failures.extend(
        f"Overview title {reference}: got {worksheet[reference].value!r}, want {wanted!r}"
        for reference, wanted in titles.items()
        if worksheet[reference].value != wanted
    )

    header_sets = {
        "A2": ["Item", "Delivery Health", "Owner", "Due"],
        "F2": ["Type", "Description", "Severity", "Owner", "Next review", "Latest Status"],
        "M2": ["Milestones / Decisions / Deadlines", "Date", "Scope"],
        "Q2": ["Completed work", "Type", "Owner", "Completed", "Scope"],
    }
    for start, wanted in header_sets.items():
        start_cell = worksheet[start]
        actual = [
            worksheet.cell(row=start_cell.row, column=start_cell.column + offset).value
            for offset in range(len(wanted))
        ]
        if actual != wanted:
            failures.append(f"Overview headers at {start}: got {actual}, want {wanted}")

    if worksheet.freeze_panes != "F3":
        failures.append(f"Overview freeze pane is {worksheet.freeze_panes!r}, want 'F3'")
    if "$A$1:$U$7" not in str(worksheet.print_area):
        failures.append(f"Overview print area is {worksheet.print_area!r}, want A1:U7")
    cf_ranges = {str(conditional.sqref) for conditional in worksheet.conditional_formatting}
    failures.extend(
        f"Overview conditional formatting missing on {reference}"
        for reference in ("B3:B7", "D3:D7", "H3:H7", "J3:J7", "N3:N7")
        if reference not in cf_ranges
    )
    if worksheet.data_validations.dataValidation:
        failures.append("Overview must have no inputs (found data validation)")
    return failures


def _data_surface_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    items = workbook["Items"]
    item_headers = table_headers(items, "tblItems")
    if item_headers[: len(ITEMS_CORE)] != ITEMS_CORE:
        failures.append(
            f"Items core headers: got {item_headers[: len(ITEMS_CORE)]}, want {ITEMS_CORE}"
        )
    failures.extend(grouped_column_failures(items, len(ITEMS_CORE), len(item_headers), "Items"))
    if items.freeze_panes != "D3":
        failures.append(f"Items freeze pane is {items.freeze_panes!r}, want 'D3'")

    people_headers = table_headers(workbook["Config"], "tblPeople")
    if people_headers != PEOPLE_HEADERS:
        failures.append(f"People headers: got {people_headers}, want {PEOPLE_HEADERS}")

    raid = workbook["RAID"]
    raid_headers = table_headers(raid, "tblRAID")
    expected_raid_headers = RAID_CORE + RAID_SYSTEM
    if raid_headers != expected_raid_headers:
        failures.append(f"RAID headers: got {raid_headers}, want {expected_raid_headers}")
    failures.extend(grouped_column_failures(raid, len(RAID_CORE), len(raid_headers), "RAID"))
    return failures


def _example_failures(workbook: Workbook) -> list[str]:
    items = workbook["Items"]
    columns = {
        str(cell.value): cell.column for cell in items[ITEMS_HDR_ROW] if cell.value is not None
    }
    actual = [
        (
            items.cell(row, columns["ID"]).value,
            items.cell(row, columns["Type"]).value,
            items.cell(row, columns["Status"]).value,
        )
        for row in range(ITEMS_DATA_ROW, ITEMS_DATA_ROW + len(EXPECTED_ITEM_OUTLINES))
    ]
    wanted = [
        ("I-1001", "Project", "In Progress"),
        ("I-1002", "Epic", "In Progress"),
        ("I-1003", "Task", "Done"),
        ("I-1004", "Release", "Ready"),
    ]
    failures: list[str] = []
    if actual != wanted:
        failures.append(f"Items examples differ: {actual}")
    for offset, wanted_level in enumerate(EXPECTED_ITEM_OUTLINES):
        row = ITEMS_DATA_ROW + offset
        actual_level = items.row_dimensions[row].outline_level
        if actual_level != wanted_level:
            failures.append(
                f"Items example row {row} outline level {actual_level}, want {wanted_level}"
            )
    if [workbook["RAID"]["B3"].value, workbook["RAID"]["B4"].value] != [
        "Risk",
        "Decision",
    ]:
        failures.append("RAID examples must demonstrate one Risk and one Decision")
    return failures


def _config_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Config"]
    expected_headers = {
        "tblStatuses": ["Status", "IsActive", "IsDone", "IsCancelled", "IsDeleted"],
        "tblRaidStatuses": ["RaidStatus", "IsClosed", "IsDeleted"],
        "tblDeliveryHealth": ["Delivery Health"],
        "tblTypes": ["Type", "Level"],
        "tblSeverity": ["Severity", "MinScore"],
    }
    failures = [
        f"Config {table_name} headers differ: {table_headers(worksheet, table_name)}"
        for table_name, wanted in expected_headers.items()
        if table_headers(worksheet, table_name) != wanted
    ]
    expected_refs = {
        "tblStatuses": "E3:I10",
        "tblTypes": "K3:L17",
        "tblPriorities": "N3:N8",
        "tblTeams": "P3:P4",
        "tblRaidTypes": "R3:T8",
        "tblRaidStatuses": "V3:X7",
        "tblSeverity": "Z3:AA7",
        "tblDeliveryHealth": "AC3:AC7",
        "tblPeople": "AE3:AG4",
    }
    failures.extend(
        f"Config {table_name} is {worksheet.tables[table_name].ref}, "
        f"want side-by-side band {wanted}"
        for table_name, wanted in expected_refs.items()
        if worksheet.tables[table_name].ref != wanted
    )
    config_text = "\n".join(
        str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None
    ).lower()
    failures.extend(
        f"Config guidance missing {fragment!r}"
        for fragment in ("project/product", "latest status", "entering dates", "key dates")
        if fragment not in config_text
    )
    return failures


def _plan_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Plan"]
    failures: list[str] = []
    if worksheet["A1"].value != "Plan - schedule & key dates":
        failures.append("Plan title changed unexpectedly")
    if worksheet["A2"].value != "Scope" or worksheet["A3"].value != "Depth":
        failures.append("Plan filter labels are not Scope (r2) / Depth (r3)")
    if worksheet["D2"].value != "From" or worksheet["D3"].value != "To":
        failures.append("Plan window labels are not From (D2) / To (D3)")

    scope_validation = validation_for(worksheet, "B2")
    if (
        scope_validation is None
        or scope_validation.type != "list"
        or "dvScopeLabels" not in str(scope_validation.formula1)
    ):
        failures.append("Plan Scope selector is not fed by dvScopeLabels")
    depth_validation = validation_for(worksheet, "B3")
    if (
        depth_validation is None
        or depth_validation.type != "list"
        or not all(value in str(depth_validation.formula1) for value in ("1", "6"))
    ):
        failures.append("Plan Depth selector is not the 1-6 level list")
    failures.extend(
        f"Plan helper column {column} is visible"
        for column in ("BG", "BH", "BI")
        if not column_dimension_for(worksheet, column_index_from_string(column)).hidden
    )
    failures.extend(
        f"Plan window input {reference} lacks date validation"
        for reference in ("E2", "E3")
        if (validation_for(worksheet, reference) is None)
        or validation_for(worksheet, reference).type != "date"
    )
    failures.extend(
        "Plan "
        f"{reference} format is "
        f"{normalise_number_format(worksheet[reference].number_format)!r}, "
        f"want {DATE_NUMBER_FORMAT}"
        for reference in ("D6", "E6")
        if normalise_number_format(worksheet[reference].number_format) != DATE_NUMBER_FORMAT
    )
    return failures


def _date_validation_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    for sheet, columns in DATE_TARGETS.items():
        worksheet = workbook[sheet]
        table_name = DATE_TABLES[sheet]
        table = worksheet.tables[table_name]
        min_col, min_row, _max_col, _max_row = range_boundaries(table.ref)
        headers = table_headers(worksheet, table_name)
        for header in columns:
            coordinate = f"{get_column_letter(min_col + headers.index(header))}{min_row + 1}"
            validation = validation_for(worksheet, coordinate)
            if validation is None or validation.type != "date":
                failures.append(f"{sheet} {header} has no date validation")
                continue
            if (
                validation.promptTitle != "Enter a date"
                or "Ctrl+;" not in str(validation.prompt)
                or not validation.showInputMessage
                or validation.errorStyle not in {None, "stop"}
            ):
                failures.append(f"{sheet} {header} date validation lacks the standard prompt")
            actual_format = normalise_number_format(worksheet[coordinate].number_format)
            if actual_format != DATE_NUMBER_FORMAT:
                failures.append(
                    f"{sheet} {header} format is {actual_format!r}, want {DATE_NUMBER_FORMAT}"
                )
    return failures


def _protection_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    for worksheet in workbook.worksheets:
        if worksheet.title in PROTECTED_SHEETS and not worksheet.protection.sheet:
            failures.append(f"{worksheet.title} should ship protected (view/system sheet)")
        if worksheet.title not in PROTECTED_SHEETS and worksheet.protection.sheet:
            failures.append(f"{worksheet.title} must stay unprotected (growable tables)")
    return failures


def _items_attention_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Items"]
    validated_columns: set[int] = set()
    for validation in worksheet.data_validations.dataValidation:
        for cell_range in str(validation.sqref).split():
            match = re.match(r"\$?([A-Z]+)\$?\d+(?::\$?([A-Z]+)\$?\d+)?", cell_range)
            if match is None:
                continue
            first = column_index_from_string(match.group(1))
            last = column_index_from_string(match.group(2)) if match.group(2) else first
            validated_columns.update(range(first, last + 1))
    failures: list[str] = []
    if len(validated_columns) < MIN_ITEMS_VALIDATED_COLUMNS:
        failures.append(
            f"Items DV covers only {len(validated_columns)} column(s), "
            f"want >= {MIN_ITEMS_VALIDATED_COLUMNS}"
        )
    if len(list(worksheet.conditional_formatting)) < MIN_ITEMS_CF_RANGES:
        failures.append(f"Items conditional formatting rules < {MIN_ITEMS_CF_RANGES}")
    return failures


def workbook_failures(workbook: Workbook) -> list[str]:
    """Check workbook inventory, surfaces, controls, dates, and protection.

    Returns:
        Every workbook-layer structural violation.

    """
    failures = _inventory_failures(workbook)
    if not _required_objects_present(workbook):
        return failures
    return [
        *failures,
        *_overview_failures(workbook),
        *_data_surface_failures(workbook),
        *_example_failures(workbook),
        *_config_failures(workbook),
        *_plan_failures(workbook),
        *_date_validation_failures(workbook),
        *_protection_failures(workbook),
        *_items_attention_failures(workbook),
    ]
