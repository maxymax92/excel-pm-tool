"""Verify that an empty workbook presents explicit, error-free view states."""

from __future__ import annotations

import logging
import zipfile

import openpyxl

from ..core.layout import REGISTRY
from ..pipeline import build_one
from .common import temporary_examples, temporary_workbook, workbook_error_cells
from .excel import recalculate

LOGGER = logging.getLogger(__name__)


def _overview_failures(workbook: openpyxl.Workbook) -> list[str]:
    failures: list[str] = []
    if "Overview" not in workbook.sheetnames:
        return ["Overview sheet missing"]

    overview = workbook["Overview"]
    checks = (
        ("scopes empty", overview["A3"].value, "No open items at configured levels."),
        ("RAID empty type", overview["F3"].value, "—"),
        ("RAID empty description", overview["G3"].value, "No high-severity open RAID."),
        (
            "coming up empty",
            overview["M3"].value,
            "No upcoming milestones, decisions or deadlines.",
        ),
        ("recent empty", overview["Q3"].value, "No dated completions in this period."),
    )
    failures.extend(
        f"Overview {label}: got {got!r}, want {want!r}"
        for label, got, want in checks
        if got != want
    )
    rendered = " ".join(
        str(cell.value) for row in overview.iter_rows() for cell in row if cell.value is not None
    )
    if "1900" in rendered:
        failures.append("Overview exposes a zero-date (1900)")
    return failures


def _plan_failures(workbook: openpyxl.Workbook) -> list[str]:
    failures: list[str] = []
    plan = workbook["Plan"]
    if plan["A6"].value != "— none —":
        failures.append(f"Plan empty guard: got {plan['A6'].value!r}")
    if plan["F2"].value not in {None, ""}:
        failures.append(f"Plan missing-dates counter is not empty: {plan['F2'].value!r}")
    return failures


def _spill_zone_failures() -> list[str]:
    expected_zones = {
        "overview-projects",
        "overview-risks",
        "overview-milestones",
        "overview-completed",
    }
    actual_zones = {zone.label for zone in REGISTRY.zones.get("Overview", [])}
    if actual_zones == expected_zones:
        return []
    return [f"Overview spill zones: got {sorted(actual_zones)}, want {sorted(expected_zones)}"]


def main() -> int:
    """Build and verify the empty workbook state.

    Returns:
        A process exit status: zero for success and one for failure.

    """
    fails: list[str] = []
    error_count = 0
    with temporary_examples() as examples, temporary_workbook("PM_empty") as out:
        examples.ITEMS_EXAMPLES = []
        examples.PEOPLE_EXAMPLES = []
        examples.RAID_EXAMPLES = []
        build_one(out, with_vba=False)
        recalculate(out)

        workbook = openpyxl.load_workbook(out, data_only=True)
        try:
            errors = workbook_error_cells(workbook)
            error_count = len(errors)
            fails.extend(f"error {value}" for value in errors[:20])
            fails.extend(_overview_failures(workbook))
            fails.extend(_plan_failures(workbook))
            fails.extend(_spill_zone_failures())
        finally:
            workbook.close()

        with zipfile.ZipFile(out) as package:
            bad_part = package.testzip()
            if bad_part is not None:
                fails.append(f"zip corrupt: {bad_part}")

    if fails:
        error_suffix = f" ({error_count} error cells)" if error_count else ""
        LOGGER.error(
            "EMPTY-STATE QA: %s issue(s)%s",
            len(fails),
            error_suffix,
        )
        for failure in fails:
            LOGGER.error("  FAIL %s", failure)
        return 1
    LOGGER.info("EMPTY-STATE QA: ALL PASS")
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    raise SystemExit(main())
