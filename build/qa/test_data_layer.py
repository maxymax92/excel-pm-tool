"""Regression contracts for the authored-data export, snapshot and injection layer."""

from __future__ import annotations

import json
import tempfile
import unittest
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import TableColumn, TableFormula

from .. import pipeline
from ..data.export import (
    ExportError,
    ExportResult,
    _coerce_date,
    _coerce_text,
    export_workbook,
)
from ..data.inject import InjectError, injected_source, validate_snapshot
from ..data.migrate import MigrationError, _require_workbook
from ..data.schema import (
    EXAMPLE_MARKER,
    SETTINGS_DEFAULTS,
    TABLES_BY_NAME,
    schema_fingerprint,
)
from ..data.snapshot import (
    RING_LIMIT,
    Snapshot,
    SnapshotError,
    read_snapshot,
    write_snapshot,
)
from ..spec import config, examples
from ..spec.capacity import CONFIG_ROWS, DATA_ROWS


def _is_marked(row: dict[str, object]) -> bool:
    return any(isinstance(value, str) and EXAMPLE_MARKER in value for value in row.values())


FIXTURE_SETTINGS = {
    "cfgDueSoonDays": 9,
    "cfgItemIDPrefix": "I-",
    "cfgNextItemID": 3004,
    "cfgRaidIDPrefix": "R-",
    "cfgNextRaidID": 102,
}

ITEM_ROWS = (
    {
        "ID": "I-3001",
        "Title": "Rollout programme",
        "Type": "Project",
        "Status": "In Progress",
        "Delivery Health": "On track",
        "Priority": "P1",
        "Owner": "Ana Ruiz",
        "Start": date(2026, 6, 1),
        "Due": date(2026, 9, 30),
        "Latest Status": "Tracking to plan.",
        "Source": "api:delivery",
        "Source ID": "project-3001",
        "Created": date(2026, 5, 20),
        "Updated": date(2026, 7, 10),
        "LatestUpdateOn": date(2026, 7, 10),
    },
    {
        "ID": "I-3002",
        "Title": "Vendor workstream",
        "Type": "Workstream",
        "Parent": "I-3001",
        "Status": "In Progress",
        "Priority": "P2",
        "Owner": "Ana Ruiz",
        "Start": date(2026, 6, 5),
        "Due": date(2026, 8, 14),
        "Latest Status": "Contracting is underway.",
        "Created": date(2026, 5, 25),
        "Updated": date(2026, 7, 8),
        "InProgressSince": date(2026, 6, 5),
        "LatestUpdateOn": date(2026, 7, 8),
    },
    {
        "ID": "I-3003",
        "Title": "Contract signed",
        "Type": "Task",
        "Parent": "I-3002",
        "Status": "Done",
        "Priority": "P2",
        "Owner": "Ben Cole",
        "Start": date(2026, 6, 5),
        "Due": date(2026, 6, 20),
        "Latest Status": "Signed and archived.",
        "Created": date(2026, 5, 25),
        "Updated": date(2026, 6, 20),
        "InProgressSince": date(2026, 6, 5),
        "DoneDate": date(2026, 6, 20),
        "LatestUpdateOn": date(2026, 6, 20),
    },
)

RAID_ROWS = (
    {
        "RaidID": "R-101",
        "Type": "Risk",
        "Title": "Vendor capacity may slip",
        "Detail": "Contracting depends on one supplier team.",
        "RelatedID": "I-3002",
        "Owner": "Ana Ruiz",
        "Status": "Open",
        "Prob": 3,
        "Impact": 4,
        "Response": "Secure a named backup supplier.",
        "NextReview": date(2026, 7, 20),
        "Raised": date(2026, 7, 1),
        "Updated": date(2026, 7, 10),
    },
)


def _config_tables() -> dict[str, tuple[dict[str, object], ...]]:
    return {
        "tblStatuses": tuple(
            {
                "Status": status,
                "IsActive": active,
                "IsDone": done,
                "IsCancelled": cancelled,
                "IsDeleted": deleted,
            }
            for status, active, done, cancelled, deleted in config.STATUSES
        ),
        "tblTypes": (
            *({"Type": name, "Level": level} for name, level in config.TYPES),
            {"Type": "Workstream", "Level": 3},
        ),
        "tblPriorities": tuple({"Priority": priority} for priority in config.PRIORITIES),
        "tblTeams": ({"Team": "Alpha"}, {"Team": "Beta"}),
        "tblRaidTypes": tuple(
            {"RaidType": name, "IsAlert": alert, "IsDecision": decision}
            for name, alert, decision in config.RAID_TYPES
        ),
        "tblRaidStatuses": tuple(
            {"RaidStatus": name, "IsClosed": closed, "IsDeleted": deleted}
            for name, closed, deleted in config.RAID_STATUSES
        ),
        "tblSeverity": tuple(
            {"Severity": severity, "MinScore": score} for severity, score in config.SEVERITY
        ),
        "tblDeliveryHealth": tuple({"Delivery Health": value} for value in config.DELIVERY_HEALTH),
        "tblPeople": (
            {"Person": "Ana Ruiz", "Role": "Delivery Lead", "Team": "Alpha"},
            {"Person": "Ben Cole", "Role": "Engineer", "Team": "Beta"},
        ),
    }


def _snapshot(
    *,
    settings: dict[str, object] | None = None,
    tables: dict[str, tuple[dict[str, object], ...]] | None = None,
) -> Snapshot:
    merged_tables = {
        **_config_tables(),
        "tblItems": tuple(dict(row) for row in ITEM_ROWS),
        "tblRAID": tuple(dict(row) for row in RAID_ROWS),
    }
    if tables:
        merged_tables.update(tables)
    return Snapshot(
        schema_fingerprint=schema_fingerprint(),
        workbook_digest="0" * 64,
        exported_at="2026-07-15T00:00:00Z",
        settings=dict(FIXTURE_SETTINGS if settings is None else settings),
        tables=merged_tables,
    )


def _build_exported(directory: str, *, injected: bool) -> ExportResult:
    workbook = Path(directory) / "fixture.xlsx"
    if injected:
        snapshot = _snapshot()
        with injected_source(snapshot, validate_snapshot(snapshot)):
            pipeline.build_one(workbook, with_vba=False)
    else:
        pipeline.build_one(workbook, with_vba=False)
    return export_workbook(workbook)


def _add_legacy_item_formula_columns(workbook: Path) -> None:
    """Add the removed schedule-envelope columns to a current workbook fixture."""
    document = openpyxl.load_workbook(workbook, data_only=False)
    worksheet = document["Items"]
    table = worksheet.tables["tblItems"]
    min_col, header_row, max_col, max_row = range_boundaries(table.ref)
    headers = {
        str(worksheet.cell(header_row, column).value) for column in range(min_col, max_col + 1)
    }

    for name, direct_column in (("EffStart", "Start"), ("EffDue", "Due")):
        if name in headers:
            continue
        max_col += 1
        worksheet.cell(header_row, max_col, name)
        formula = f'=IF([@ID]="","",[@{direct_column}])'
        for row in range(header_row + 1, max_row + 1):
            worksheet.cell(row, max_col, formula)
        table.tableColumns.append(
            TableColumn(
                id=max(column.id for column in table.tableColumns) + 1,
                name=name,
                calculatedColumnFormula=TableFormula(
                    attr_text=(f'IF([[#This Row],ID]="","",[[#This Row],{direct_column}])')
                ),
            )
        )
        headers.add(name)

    table.ref = f"{get_column_letter(min_col)}{header_row}:{get_column_letter(max_col)}{max_row}"
    document.save(workbook)


class LegacyFormulaColumnCompatibilityTests(unittest.TestCase):
    """Keep old calculated columns outside the authored-data contract."""

    def test_obsolete_schedule_formulas_do_not_compromise_round_trip(self) -> None:
        """Export and reinject authored values around legacy formula-only columns."""
        source = _snapshot()
        source_reconciliation = validate_snapshot(source)
        with tempfile.TemporaryDirectory() as directory:
            legacy_workbook = Path(directory) / "legacy.xlsx"
            rebuilt_workbook = Path(directory) / "rebuilt.xlsx"
            with injected_source(source, source_reconciliation):
                pipeline.build_one(legacy_workbook, with_vba=False)
            _add_legacy_item_formula_columns(legacy_workbook)

            exported = export_workbook(legacy_workbook)
            formula_notes = tuple(note for note in exported.notes if "Eff" in note)
            self.assertEqual(
                formula_notes,
                (
                    "tblItems[EffStart] is an unknown formula column; not exported",
                    "tblItems[EffDue] is an unknown formula column; not exported",
                ),
            )
            self.assertNotIn("tblItems", exported.unknown_columns)
            self.assertEqual(exported.snapshot.tables, source.tables)
            self.assertEqual(exported.snapshot.settings, source_reconciliation.settings)

            exported_reconciliation = validate_snapshot(exported.snapshot)
            with injected_source(exported.snapshot, exported_reconciliation):
                pipeline.build_one(rebuilt_workbook, with_vba=False)
            rebuilt = export_workbook(rebuilt_workbook)

        self.assertEqual(rebuilt.snapshot.tables, exported.snapshot.tables)
        self.assertEqual(rebuilt.snapshot.settings, exported.snapshot.settings)


class DataRoundTripTests(unittest.TestCase):
    """Prove injected rows and settings export back exactly."""

    snapshot: Snapshot
    result: ExportResult

    @classmethod
    def setUpClass(cls) -> None:
        """Build one injected workbook and export it."""
        super().setUpClass()
        directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(directory.cleanup)
        cls.snapshot = _snapshot()
        cls.result = _build_exported(directory.name, injected=True)

    def test_every_table_round_trips_exactly(self) -> None:
        """Preserve every authored row, value and order through build and export."""
        for table, rows in self.snapshot.tables.items():
            with self.subTest(table=table):
                self.assertEqual(list(self.result.snapshot.tables[table]), list(rows))

    def test_settings_round_trip_as_the_merged_band(self) -> None:
        """Export the full settings band exactly as injected."""
        reconciliation = validate_snapshot(self.snapshot)
        self.assertEqual(self.result.snapshot.settings, reconciliation.settings)

    def test_reconciliation_is_quiet_for_consistent_data(self) -> None:
        """Report no bumps, warnings or empty tables for coherent fixtures."""
        reconciliation = validate_snapshot(self.snapshot)
        self.assertEqual(reconciliation.counter_bumps, {})
        self.assertEqual(reconciliation.warnings, ())
        self.assertEqual(reconciliation.empty_tables, ())
        self.assertTrue(reconciliation.fingerprint_matches)

    def test_export_records_the_current_schema_fingerprint(self) -> None:
        """Stamp snapshots with the schema they were exported under."""
        self.assertEqual(self.result.snapshot.schema_fingerprint, schema_fingerprint())


class ShippedExampleExportTests(unittest.TestCase):
    """Pin the example-row skip rule against the shipped fixtures."""

    result: ExportResult

    @classmethod
    def setUpClass(cls) -> None:
        """Build the default example workbook and export it."""
        super().setUpClass()
        directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(directory.cleanup)
        cls.result = _build_exported(directory.name, injected=False)

    def test_marked_example_rows_are_skipped(self) -> None:
        """Skip only rows carrying the exact delete-me marker."""
        marked_items = [row for row in examples.ITEMS_EXAMPLES if _is_marked(row)]
        marked_raid = [row for row in examples.RAID_EXAMPLES if _is_marked(row)]
        self.assertTrue(marked_items)
        self.assertEqual(self.result.skipped_examples.get("tblItems"), len(marked_items))
        self.assertEqual(self.result.skipped_examples.get("tblRAID", 0), len(marked_raid))

    def test_unmarked_example_like_rows_are_retained_and_noted(self) -> None:
        """Never drop a row that merely mentions the word example."""
        retained_rows = [row for row in examples.RAID_EXAMPLES if not _is_marked(row)]
        raid_ids = [row.get("RaidID") for row in self.result.snapshot.tables["tblRAID"]]
        self.assertEqual(raid_ids, [str(row["RaidID"]) for row in retained_rows])
        for row in retained_rows:
            if "EXAMPLE" in str(row).upper():
                raid_id = str(row["RaidID"])
                self.assertTrue(any(raid_id in note for note in self.result.notes))

    def test_default_settings_export_exactly(self) -> None:
        """Read the shipped settings band back verbatim."""
        self.assertEqual(self.result.snapshot.settings, SETTINGS_DEFAULTS)


class InjectionGateTests(unittest.TestCase):
    """Halt on data loss or unbuildable snapshots; warn on red-flag values."""

    def test_duplicate_ids_halt(self) -> None:
        """Refuse duplicate keys before any build."""
        rows = (dict(ITEM_ROWS[0]), dict(ITEM_ROWS[0]))
        with self.assertRaisesRegex(InjectError, "duplicate ID values: I-3001"):
            validate_snapshot(_snapshot(tables={"tblItems": rows}))

    def test_case_variant_duplicate_ids_halt_like_excel(self) -> None:
        """Treat workbook identifiers as case-insensitive at the rebuild gate."""
        first = dict(ITEM_ROWS[0])
        second = dict(ITEM_ROWS[1])
        second["ID"] = "i-3001"
        with self.assertRaisesRegex(InjectError, "duplicate ID values: I-3001"):
            validate_snapshot(_snapshot(tables={"tblItems": (first, second)}))

    def test_orphan_columns_halt(self) -> None:
        """Refuse to drop values from columns outside the schema."""
        row = dict(ITEM_ROWS[0])
        row["Estimate"] = "13d"
        with self.assertRaisesRegex(InjectError, "Estimate"):
            validate_snapshot(_snapshot(tables={"tblItems": (row,)}))

    def test_orphan_settings_halt(self) -> None:
        """Refuse to drop settings outside the schema."""
        settings = dict(FIXTURE_SETTINGS)
        settings["cfgRetired"] = 4
        with self.assertRaisesRegex(InjectError, "cfgRetired"):
            validate_snapshot(_snapshot(settings=settings))

    def test_unknown_item_type_halts(self) -> None:
        """Refuse types the taxonomy cannot level."""
        row = dict(ITEM_ROWS[0])
        row["Type"] = "Ghost"
        with self.assertRaisesRegex(InjectError, "Ghost"):
            validate_snapshot(_snapshot(tables={"tblItems": (row,)}))

    def test_config_and_identifier_choices_compare_case_insensitively(self) -> None:
        """Match Excel list and lookup behavior during rebuild validation."""
        item = dict(ITEM_ROWS[1])
        item.update({
            "Type": "workstream",
            "Status": "in progress",
            "Owner": "ana ruiz",
            "Parent": "i-3001",
        })
        raid = dict(RAID_ROWS[0])
        raid.update({
            "Type": "risk",
            "Status": "open",
            "Owner": "ana ruiz",
            "RelatedID": "i-3001",
        })

        reconciliation = validate_snapshot(
            _snapshot(tables={"tblItems": (ITEM_ROWS[0], item), "tblRAID": (raid,)})
        )

        self.assertFalse(
            any("outside the configured list" in warning for warning in reconciliation.warnings),
            reconciliation.warnings,
        )

    def test_capacity_breach_halts(self) -> None:
        """Refuse more rows than every workbook layer supports."""
        rows = tuple(
            {"ID": f"I-{4000 + index}", "Title": "Bulk", "Type": "Task", "Status": "Backlog"}
            for index in range(DATA_ROWS + 1)
        )
        with self.assertRaisesRegex(InjectError, str(DATA_ROWS)):
            validate_snapshot(_snapshot(tables={"tblItems": rows}))

    def test_stale_counters_are_bumped_and_reported(self) -> None:
        """Keep the VBA ID counters above every existing identifier."""
        settings = dict(FIXTURE_SETTINGS)
        settings["cfgNextItemID"] = 1
        reconciliation = validate_snapshot(_snapshot(settings=settings))
        self.assertEqual(reconciliation.counter_bumps, {"cfgNextItemID": (1, 3004)})

    def test_red_flag_values_warn_without_halting(self) -> None:
        """Report list-breaking values the workbook itself flags in red."""
        row = dict(ITEM_ROWS[0])
        row["Status"] = "Bogus"
        reconciliation = validate_snapshot(_snapshot(tables={"tblItems": (row,)}))
        self.assertTrue(any("Bogus" in warning for warning in reconciliation.warnings))

    def test_missing_type_halts_before_the_build(self) -> None:
        """Refuse rows the outline writer would crash on."""
        row = dict(ITEM_ROWS[0])
        del row["Type"]
        with self.assertRaisesRegex(InjectError, "have no Type: I-3001"):
            validate_snapshot(_snapshot(tables={"tblItems": (row,)}))

    def test_missing_title_halts_before_the_build(self) -> None:
        """Refuse rows the outline writer would crash on."""
        row = dict(ITEM_ROWS[0])
        del row["Title"]
        with self.assertRaisesRegex(InjectError, "have no Title: I-3001"):
            validate_snapshot(_snapshot(tables={"tblItems": (row,)}))


class AuthoredOwnershipTests(unittest.TestCase):
    """Retain the workbook column owner at the mutation boundary."""

    def test_item_and_raid_schema_retains_column_kinds(self) -> None:
        """Distinguish inputs, VBA stamps and system identity metadata."""
        item_columns = {column.name: column.kind for column in TABLES_BY_NAME["tblItems"].columns}
        raid_columns = {column.name: column.kind for column in TABLES_BY_NAME["tblRAID"].columns}
        self.assertEqual(item_columns["Title"], "I")
        self.assertEqual(item_columns["ID"], "V")
        self.assertEqual(item_columns["Source"], "S")
        self.assertEqual(item_columns["Source ID"], "S")
        self.assertEqual(raid_columns["RaidID"], "V")
        self.assertEqual(raid_columns["Source"], "S")
        self.assertEqual(raid_columns["Source ID"], "S")

    def test_source_identity_round_trips_with_authored_rows(self) -> None:
        """Keep source pairs as durable snapshot values."""
        snapshot = _snapshot()
        row = snapshot.tables["tblItems"][0]
        self.assertEqual(row["Source"], "api:delivery")
        self.assertEqual(row["Source ID"], "project-3001")


class DeletedRoleMigrationTests(unittest.TestCase):
    """Normalize legacy status tables without losing other Config values."""

    @staticmethod
    def _legacy_statuses() -> tuple[dict[str, object], ...]:
        return tuple(
            {
                "Status": status,
                "IsActive": active,
                "IsDone": done,
                "IsCancelled": cancelled,
            }
            for status, active, done, cancelled, deleted in config.STATUSES
            if not deleted
        )

    @staticmethod
    def _legacy_raid_statuses() -> tuple[dict[str, object], ...]:
        return tuple(
            {"RaidStatus": status, "IsClosed": closed}
            for status, closed, deleted in config.RAID_STATUSES
            if not deleted
        )

    def test_legacy_tables_gain_exact_deleted_roles(self) -> None:
        """Append one normal Deleted row to each legacy status table."""
        legacy_items = self._legacy_statuses()
        legacy_raid = self._legacy_raid_statuses()
        reconciliation = validate_snapshot(
            _snapshot(tables={"tblStatuses": legacy_items, "tblRaidStatuses": legacy_raid})
        )
        normalized = reconciliation.normalized_snapshot.tables
        item_deleted = [row for row in normalized["tblStatuses"] if row["IsDeleted"]]
        raid_deleted = [row for row in normalized["tblRaidStatuses"] if row["IsDeleted"]]
        self.assertEqual(
            item_deleted,
            [
                {
                    "Status": "Deleted",
                    "IsActive": False,
                    "IsDone": True,
                    "IsCancelled": True,
                    "IsDeleted": True,
                }
            ],
        )
        self.assertEqual(
            raid_deleted,
            [{"RaidStatus": "Deleted", "IsClosed": True, "IsDeleted": True}],
        )
        self.assertTrue(any("appended Deleted" in value for value in reconciliation.adjustments))
        for original, migrated in zip(
            legacy_items,
            normalized["tblStatuses"][: len(legacy_items)],
            strict=True,
        ):
            self.assertEqual({**original, "IsDeleted": False}, migrated)

    def test_existing_case_insensitive_label_is_normalized_in_place(self) -> None:
        """Assign exact roles to one existing Deleted label without appending."""
        existing = (
            *self._legacy_statuses(),
            {
                "Status": "dElEtEd",
                "IsActive": True,
                "IsDone": False,
                "IsCancelled": False,
            },
        )
        reconciliation = validate_snapshot(_snapshot(tables={"tblStatuses": existing}))
        rows = reconciliation.normalized_snapshot.tables["tblStatuses"]
        matching = [row for row in rows if row["IsDeleted"]]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0]["Status"], "dElEtEd")
        self.assertEqual(
            {name: matching[0][name] for name in ("IsActive", "IsDone", "IsCancelled")},
            {"IsActive": False, "IsDone": True, "IsCancelled": True},
        )
        self.assertTrue(any("normalized" in value for value in reconciliation.adjustments))

    def test_ambiguous_deleted_labels_halt(self) -> None:
        """Refuse to guess between multiple case-insensitive Deleted labels."""
        rows = (
            *self._legacy_statuses(),
            {"Status": "Deleted", "IsActive": False, "IsDone": True, "IsCancelled": True},
            {"Status": "DELETED", "IsActive": False, "IsDone": True, "IsCancelled": True},
        )
        with self.assertRaisesRegex(InjectError, "multiple deletion-role candidates"):
            validate_snapshot(_snapshot(tables={"tblStatuses": rows}))

    def test_multiple_explicit_deleted_roles_halt(self) -> None:
        """Require exactly one effective deletion role in current Config."""
        rows = (
            {
                "Status": "Deleted",
                "IsActive": False,
                "IsDone": True,
                "IsCancelled": True,
                "IsDeleted": True,
            },
            {
                "Status": "Archived",
                "IsActive": False,
                "IsDone": True,
                "IsCancelled": True,
                "IsDeleted": True,
            },
        )
        with self.assertRaisesRegex(InjectError, "multiple deletion-role candidates"):
            validate_snapshot(_snapshot(tables={"tblStatuses": rows}))

    def test_deleted_append_respects_config_capacity(self) -> None:
        """Halt rather than overflow Config when no Deleted row fits."""
        rows = tuple(
            {
                "Status": f"State {index}",
                "IsActive": False,
                "IsDone": False,
                "IsCancelled": False,
            }
            for index in range(CONFIG_ROWS)
        )
        with self.assertRaisesRegex(InjectError, "cannot append Deleted"):
            validate_snapshot(_snapshot(tables={"tblStatuses": rows}))


class CellCoercionTests(unittest.TestCase):
    """Pin the export-side cell typing rules."""

    def test_time_of_day_in_a_date_cell_halts(self) -> None:
        """Refuse datetimes that are not pure dates."""
        cell_value = datetime.combine(date(2026, 7, 1), time(10, 30))
        with self.assertRaisesRegex(ExportError, "time of day"):
            _coerce_date("tblItems", "Due", 3, cell_value)

    def test_numeric_text_is_converted_with_a_note(self) -> None:
        """Convert numbers typed into text cells and say so."""
        notes: list[str] = []
        self.assertEqual(_coerce_text("tblItems", "Title", 2, 42, notes), "42")
        self.assertTrue(any("exported as text" in note for note in notes))

    def test_booleans_in_text_cells_halt(self) -> None:
        """Refuse values that cannot be honest text."""
        cell_value: object = True
        with self.assertRaisesRegex(ExportError, "expected text"):
            _coerce_text("tblItems", "Title", 2, cell_value, [])


class SnapshotDocumentTests(unittest.TestCase):
    """Protect the persisted snapshot format and its retention ring."""

    def test_write_and_read_round_trip(self) -> None:
        """Persist typed rows, dates and settings losslessly."""
        snapshot = _snapshot()
        with tempfile.TemporaryDirectory() as directory:
            path = write_snapshot(snapshot, directory=directory)
            loaded = read_snapshot(path)
        self.assertEqual(loaded, snapshot)

    def test_ring_is_pruned(self) -> None:
        """Keep only the newest snapshots."""
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory)
            for index in range(RING_LIMIT + 3):
                stale = target / f"pm-data-20260101T{index:06}-000000000000.json"
                stale.write_text("{}", encoding="utf-8")
            write_snapshot(_snapshot(), directory=directory)
            remaining = list(target.glob("pm-data-*.json"))
        self.assertEqual(len(remaining), RING_LIMIT)

    def test_unsupported_format_is_rejected(self) -> None:
        """Refuse documents from an unknown snapshot format."""
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "pm-data-bad.json"
            path.write_text(json.dumps({"format": 99}), encoding="utf-8")
            with self.assertRaisesRegex(SnapshotError, "unsupported format"):
                read_snapshot(path)

    def test_bad_dates_are_rejected(self) -> None:
        """Refuse rows whose date columns cannot parse."""
        snapshot = _snapshot()
        with tempfile.TemporaryDirectory() as directory:
            path = write_snapshot(snapshot, directory=directory)
            document = json.loads(path.read_text(encoding="utf-8"))
            document["tables"]["tblItems"][0]["Start"] = "not-a-date"
            path.write_text(json.dumps(document), encoding="utf-8")
            with self.assertRaisesRegex(SnapshotError, "ISO date"):
                read_snapshot(path)


class MigrateGuardTests(unittest.TestCase):
    """Refuse migration targets that are missing, foreign or open in Excel."""

    def test_missing_workbook_is_rejected(self) -> None:
        """Fail fast when the workbook path does not exist."""
        with (
            tempfile.TemporaryDirectory() as directory,
            self.assertRaisesRegex(MigrationError, "does not exist"),
        ):
            _require_workbook(Path(directory) / "absent.xlsm")

    def test_non_macro_workbook_is_rejected(self) -> None:
        """Only the macro-enabled artifact is a migration target."""
        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "plain.xlsx"
            workbook.touch()
            with self.assertRaisesRegex(MigrationError, "not an .xlsm"):
                _require_workbook(workbook)

    def test_excel_lock_file_is_rejected(self) -> None:
        """Refuse to run while Excel holds the workbook open."""
        with tempfile.TemporaryDirectory() as directory:
            workbook = Path(directory) / "PM_Workbook.xlsm"
            workbook.touch()
            (Path(directory) / "~$PM_Workbook.xlsm").touch()
            with self.assertRaisesRegex(MigrationError, "open in Excel"):
                _require_workbook(workbook)


if __name__ == "__main__":
    unittest.main()
