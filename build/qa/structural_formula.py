"""Formula, name, spill, and conditional-expression structural checks."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.utils import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

from ..spec.capacity import DATA_ROWS
from ..spec.items import DIRECT_BLOCKED_HEALTH_FORMULA
from .structural_shared import EXPECTED_NAMES, formula_parentheses_balanced

if TYPE_CHECKING:
    from openpyxl.cell.cell import Cell
    from openpyxl.workbook.workbook import Workbook

MAX_CALC_FORMULAS = 20
BULK_HELPER_MIN_COLUMN = 14
BULK_HELPER_MAX_COLUMN = 35
WBS_BLANK_SENTINEL = 'REPT("Z",50)'
OBSOLETE_SCHEDULE_FIELDS = ("EffStart", "EffDue")


def _is_formula(cell: Cell) -> bool:
    return cell.data_type == "f" or isinstance(cell.value, ArrayFormula)


def _table_formula_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    for sheet, table_name in (("Items", "tblItems"), ("RAID", "tblRAID")):
        worksheet = workbook[sheet]
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        failures.extend(
            f"{sheet}!{cell.coordinate} is a dynamic-array formula inside a calculated table column"
            for row in worksheet.iter_rows(
                min_row=min_row + 1,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
            for cell in row
            if isinstance(cell.value, ArrayFormula)
        )
    items_table = workbook["Items"].tables["tblItems"]
    item_headers = {column.name for column in items_table.tableColumns}
    failures.extend(
        f"Items table still contains obsolete calculated column {column}"
        for column in OBSOLETE_SCHEDULE_FIELDS
        if column in item_headers
    )
    wbs_column = next(
        (column for column in items_table.tableColumns if column.name == "WbsKey"),
        None,
    )
    wbs_formula_value = None if wbs_column is None else wbs_column.calculatedColumnFormula
    wbs_formula = str(getattr(wbs_formula_value, "text", "") or "")
    if WBS_BLANK_SENTINEL not in wbs_formula:
        failures.append(
            "Items WbsKey blank rows lack the high text sentinel required by Mac Excel sort"
        )
    return failures


def _defined_name_failures(workbook: Workbook) -> list[str]:
    names = {
        name: str(workbook.defined_names[name].value)
        for name in workbook.defined_names
        if not name.casefold().startswith("_xlpm.")
    }
    failures = [f"defined name missing: {name}" for name in sorted(EXPECTED_NAMES - set(names))]
    failures.extend(
        f"unexpected defined name: {name}" for name in sorted(set(names) - EXPECTED_NAMES)
    )
    scope_labels = names.get("dvScopeLabels", "")
    if scope_labels and f"$M$2:$M${DATA_ROWS + 2}" not in scope_labels:
        failures.append("dvScopeLabels is not aligned to the selector helper")
    failures.extend(
        f"LAMBDA name {name} mis-encoded: {value[:80]}"
        for name, value in names.items()
        if name.startswith("fn") and ("_xlfn.LAMBDA(" not in value or "_xlfn._xlpm" in value)
    )
    sort_date = names.get("fnSortDate", "")
    failures.extend(
        f"fnSortDate does not use the item's direct schedule field {token}"
        for token in ("tblItems[Start]", "tblItems[Due]")
        if token not in sort_date
    )
    return failures


def _calc_formula_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Calc"]
    formulas = [cell for row in worksheet.iter_rows() for cell in row if _is_formula(cell)]
    failures: list[str] = []
    if len(formulas) > MAX_CALC_FORMULAS:
        failures.append(
            f"Calc contains {len(formulas)} formulas; "
            f"the bounded helper layer allows {MAX_CALC_FORMULAS}"
        )
    bulk_formula = next(
        (
            cell
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=DATA_ROWS + 1,
                min_col=BULK_HELPER_MIN_COLUMN,
                max_col=BULK_HELPER_MAX_COLUMN,
            )
            for cell in row
            if _is_formula(cell)
        ),
        None,
    )
    if bulk_formula is not None:
        failures.append(f"Calc contains a bulk helper formula at {bulk_formula.coordinate}")
    blocked_count = _formula_text(worksheet["AK2"])
    if DIRECT_BLOCKED_HEALTH_FORMULA not in blocked_count:
        failures.append("Calc blocked count does not use the shared final-nonblank identity")
    return failures


def _formula_text(cell: Cell) -> str:
    value = cell.value
    formula = str(getattr(value, "text", value) or "")
    for prefix in ("_xlfn.", "_xlws.", "_xlpm."):
        formula = formula.replace(prefix, "")
    return formula


def _overview_formula_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Overview"]
    failures: list[str] = []
    for reference in ("A3", "F3", "M3", "Q3"):
        cell = worksheet[reference]
        formula = _formula_text(cell)
        if cell.data_type != "f":
            failures.append(f"Overview spill anchor {reference} is not a formula")
        if not any(source in formula for source in ("tblItems", "tblRAID", "Calc!")):
            failures.append(f"Overview spill anchor {reference} lacks a live data source")

    recent_formula = _formula_text(worksheet["Q3"])
    failures.extend(
        f"Recent progress lacks {token}"
        for token in ("tblItems[DoneDate]", "cfgReportDays")
        if token not in recent_formula
    )

    summary_formula = _formula_text(worksheet["A3"])
    failures.extend(
        f"Executive Status Summary is not Config-level driven ({token} missing)"
        for token in (
            "cfgExecutiveStatusMaxLevel",
            "tblItems[Level]",
            "lstDoneStatus",
            "lstCancelledStatus",
            "tblItems[Delivery Health]",
            "dvDeliveryHealth",
            DIRECT_BLOCKED_HEALTH_FORMULA,
            'own,FILTER(tblItems[Owner],pred,"")',
            "due,FILTER(tblItems[Due],pred,0)",
        )
        if token not in summary_formula
    )
    failures.extend(
        f"Executive Status Summary still rolls health through descendants ({token} present)"
        for token in ("MAP(ids", "tblItems[A2]=itemid", "tblItems[A5]=itemid")
        if token in summary_formula
    )
    if "Calc!$N" in summary_formula:
        failures.append("Executive Status Summary depends on the forbidden Calc bulk grid")
    if "XLOOKUP(tblItems[Scope],tblItems[ID],tblItems[Owner]" in summary_formula:
        failures.append("Executive Status Summary still substitutes the scope Owner")
    return failures


def _plan_formula_failures(workbook: Workbook) -> list[str]:
    worksheet = workbook["Plan"]
    helper = str(worksheet["BG2"].value or "")
    failures: list[str] = []
    if 'OR(B2="",B2="All")' not in helper:
        failures.append("Plan blank Scope does not resolve to All")
    if worksheet["BG2"].data_type != "f" or "B2" not in helper:
        failures.append("Plan scope-ID helper BG2 is missing or not formula-driven")
    if worksheet["F2"].data_type != "f" or _formula_text(worksheet["F2"]) != "=BG3":
        failures.append("Plan status rail F2 is not an explicit formula linked to BG3")
    metadata_formula = _formula_text(worksheet["D1"])
    if worksheet["D1"].data_type != "f" or "TODAY()" not in metadata_formula:
        failures.append("Plan reporting-date metadata is not an explicit live formula")

    row_formula = _formula_text(worksheet["A6"])
    failures.extend(
        f"Plan row inclusion is missing direct structural predicate {token}"
        for token in ("tblItems[ID]", "tblItems[Scope]", "tblItems[Level]")
        if token not in row_formula
    )
    failures.extend(
        f"Plan row inclusion still depends on schedule state {token}"
        for token in ("tblItems[Start]", "tblItems[Due]", "tblItems[IsPoint]")
        if token in row_formula
    )
    failures.extend(
        f"Plan {coordinate} does not display the item's direct {column}"
        for coordinate, column in (("D6", "Start"), ("E6", "Due"))
        if any(
            token not in _formula_text(worksheet[coordinate])
            for token in ("fnItemLookup", f"tblItems[{column}]")
        )
    )

    category_formula = _formula_text(worksheet["BI6"])
    failures.extend(
        f"Plan category helper does not use direct {token}"
        for token in ("tblItems[Status]", "tblItems[Due]")
        if token not in category_formula
    )

    axis_formula = _formula_text(worksheet["F5"])
    failures.extend(
        f"Plan automatic date window does not use direct {token}"
        for token in ("tblItems[Start]", "tblItems[Due]")
        if token not in axis_formula
    )

    grid_formula = _formula_text(worksheet["F6"])
    failures.extend(
        f"Plan timeline does not use direct {token}"
        for token in ("tblItems[Start]", "tblItems[Due]", "tblItems[IsPoint]")
        if token not in grid_formula
    )
    return failures


def _obsolete_schedule_reference_failures(workbook: Workbook) -> list[str]:
    """Reject removed schedule-envelope fields from every formula-bearing layer.

    Returns:
        Every formula, name, conditional-format or validation reference found.

    """
    formula_sources: list[tuple[str, str]] = []
    for worksheet in workbook.worksheets:
        formula_sources.extend(
            (f"{worksheet.title}!{cell.coordinate}", _formula_text(cell))
            for row in worksheet.iter_rows()
            for cell in row
            if _is_formula(cell)
        )
        for table in worksheet.tables.values():
            formula_sources.extend(
                (
                    f"{worksheet.title}/{table.name}[{column.name}]",
                    str(getattr(column.calculatedColumnFormula, "text", "") or ""),
                )
                for column in table.tableColumns
                if column.calculatedColumnFormula is not None
            )
        formula_sources.extend(
            (f"{worksheet.title} conditional format", str(formula))
            for rules in worksheet.conditional_formatting._cf_rules.values()
            for rule in rules
            for formula in rule.formula or ()
        )
        formula_sources.extend(
            (f"{worksheet.title} validation", str(formula))
            for validation in worksheet.data_validations.dataValidation
            for formula in (validation.formula1, validation.formula2)
            if formula is not None
        )
    formula_sources.extend(
        (f"defined name {name}", str(workbook.defined_names[name].value))
        for name in workbook.defined_names
    )
    return [
        f"{source} still references obsolete schedule field {token}"
        for source, formula in formula_sources
        for token in OBSOLETE_SCHEDULE_FIELDS
        if token in formula
    ]


def _conditional_formula_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    for worksheet in workbook.worksheets:
        for rules in worksheet.conditional_formatting._cf_rules.values():
            failures.extend(
                f"{worksheet.title} conditional-format formula is malformed: {formula}"
                for rule in rules
                for formula in rule.formula or ()
                if not formula_parentheses_balanced(formula)
            )
    return failures


def formula_failures(workbook: Workbook) -> list[str]:
    """Check table formulas, names, bounded helpers, views, and CF expressions.

    Returns:
        Every formula-layer structural violation.

    """
    return [
        *_table_formula_failures(workbook),
        *_defined_name_failures(workbook),
        *_calc_formula_failures(workbook),
        *_overview_formula_failures(workbook),
        *_plan_formula_failures(workbook),
        *_obsolete_schedule_reference_failures(workbook),
        *_conditional_formula_failures(workbook),
    ]
