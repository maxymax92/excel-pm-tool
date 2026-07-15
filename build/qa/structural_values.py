"""Cached-value checks for the Excel-calculated release artifacts."""

from __future__ import annotations

from datetime import UTC, date, datetime, time
from typing import TYPE_CHECKING

from .common import ERROR_VALUE_RE
from .structural_shared import (
    ITEMS_HDR_ROW,
    RAID_DATA_ROW,
    RAID_HDR_ROW,
    normalise_cached_date,
)

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

MAX_REPORTED_ERROR_VALUES = 25
FLOAT_TOLERANCE = 1e-9
REPORT_DAYS = 14
ValueCheck = tuple[str, object, object]


def _dt(year: int, month: int, day: int) -> datetime:
    return datetime.combine(date(year, month, day), time.min)


def _item_checks(worksheet: Worksheet) -> list[ValueCheck]:
    columns = {
        str(cell.value): index + 1
        for index, cell in enumerate(worksheet[ITEMS_HDR_ROW])
        if cell.value is not None
    }

    def value(row: int, column: str) -> object:
        return worksheet.cell(row=row + ITEMS_HDR_ROW, column=columns[column]).value

    return [
        ("I-1001 Children", value(1, "Children"), 2),
        ("I-1002 Children", value(2, "Children"), 1),
        ("I-1003 Children", value(3, "Children"), 0),
        ("I-1004 Children", value(4, "Children"), 0),
        ("I-1001 Level", value(1, "Level"), 1),
        ("I-1002 Level", value(2, "Level"), 4),
        ("I-1003 Level", value(3, "Level"), 5),
        ("I-1004 Level", value(4, "Level"), 2),
        ("I-1003 Scope", value(3, "Scope"), "I-1001"),
        ("I-1004 Scope", value(4, "Scope"), "I-1001"),
        ("Items schema has no EffStart", "EffStart" not in columns, True),
        ("Items schema has no EffDue", "EffDue" not in columns, True),
        ("I-1001 direct Start", value(1, "Start"), _dt(2026, 6, 15)),
        ("I-1001 direct Due", value(1, "Due"), _dt(2026, 10, 30)),
        ("I-1002 direct Due", value(2, "Due"), _dt(2026, 8, 14)),
        ("I-1004 direct Start blank", value(4, "Start") in {None, ""}, True),
        ("I-1004 direct Due", value(4, "Due"), _dt(2026, 7, 20)),
        ("I-1004 IsPoint", value(4, "IsPoint"), True),
        ("I-1001 IsPoint", value(1, "IsPoint"), False),
        ("I-1001 WbsKey one segment", len(str(value(1, "WbsKey"))), 13),
        ("I-1003 WbsKey three segments", len(str(value(3, "WbsKey"))), 39),
        ("I-1003 IsBlocked", value(3, "IsBlocked"), False),
        ("I-1002 ParentTitle", value(2, "ParentTitle"), "Example Project"),
        ("I-1002 ParentLevel", value(2, "ParentLevel"), 1),
        ("I-1003 BlockedRefsValid", value(3, "BlockedRefsValid"), True),
        ("I-1001 delivery health", value(1, "Delivery Health"), "At risk"),
        ("I-1004 LatestUpdateOn stamped", value(4, "LatestUpdateOn"), _dt(2026, 7, 12)),
    ]


def _raid_checks(worksheet: Worksheet) -> list[ValueCheck]:
    columns = {
        str(cell.value): index + 1
        for index, cell in enumerate(worksheet[RAID_HDR_ROW])
        if cell.value is not None
    }
    return [
        (
            "R-001 Score",
            worksheet.cell(row=RAID_DATA_ROW, column=columns["Score"]).value,
            12,
        ),
        (
            "R-001 Severity",
            worksheet.cell(row=RAID_DATA_ROW, column=columns["Severity"]).value,
            "High",
        ),
        (
            "R-001 Scope",
            worksheet.cell(row=RAID_DATA_ROW, column=columns["Scope"]).value,
            "I-1001",
        ),
    ]


def _overview_checks(worksheet: Worksheet) -> list[ValueCheck]:
    today = datetime.now(tz=UTC).date()
    key_date_upcoming = date(2026, 7, 20) >= today
    decision_upcoming = date(2026, 7, 15) >= today
    done_in_window = (today - date(2026, 7, 3)).days <= REPORT_DAYS
    coming_label = "No upcoming milestones, decisions or deadlines."
    if decision_upcoming:
        coming_label = "Decision · Approve example launch scope"
    elif key_date_upcoming:
        coming_label = "Release · Example release"
    coming_date = None
    if decision_upcoming:
        coming_date = _dt(2026, 7, 15)
    elif key_date_upcoming:
        coming_date = _dt(2026, 7, 20)

    checks: list[ValueCheck] = [
        (
            "Overview scope row",
            str(worksheet["A3"].value).startswith("I-1001 · Project · Example Project"),
            True,
        ),
        ("Overview scope delivery health", worksheet["B3"].value, "At risk"),
        ("Overview scope due is TEXT (never ###)", worksheet["D3"].value, "30 Oct 2026"),
        ("Overview scope numeric date helper", worksheet["E3"].value, _dt(2026, 10, 30)),
        ("Overview RAID type first", worksheet["F3"].value, "Risk"),
        (
            "Overview RAID description",
            str(worksheet["G3"].value).startswith("Example risk"),
            True,
        ),
        ("Overview RAID severity", worksheet["H3"].value, "High"),
        (
            "Overview RAID latest status",
            worksheet["K3"].value,
            "Mitigate: weekly checkpoint with supplier",
        ),
        ("Overview RAID numeric date helper", worksheet["L3"].value, _dt(2026, 7, 20)),
        ("Overview coming-up item surfaced", worksheet["M3"].value, coming_label),
        (
            "Overview completed surfaced",
            worksheet["Q3"].value,
            "Example completed deliverable"
            if done_in_window
            else "No dated completions in this period.",
        ),
        ("Overview Coming Up numeric date helper", worksheet["P3"].value, coming_date),
    ]
    if decision_upcoming and key_date_upcoming:
        checks.append((
            "Overview release follows the decision",
            worksheet["M4"].value,
            "Release · Example release",
        ))
    return checks


def _plan_checks(worksheet: Worksheet) -> list[ValueCheck]:
    key_date_glyphs = [worksheet.cell(row=9, column=column).value for column in range(6, 58)]
    interval_glyphs = {"\u2713", "\u25cf", "!", "\u00d7", "\u2014"}
    return [
        ("Plan row 1 is the Project", worksheet["A6"].value, "I-1001"),
        ("Plan row 2 is the Epic", worksheet["A7"].value, "I-1002"),
        ("Plan row 3 is the Task", worksheet["A8"].value, "I-1003"),
        ("Plan row 4 is the Release", worksheet["A9"].value, "I-1004"),
        ("Plan L1 title unindented", worksheet["B6"].value, "Example Project"),
        ("Plan L4 title indented", worksheet["B7"].value, "         Example epic"),
        (
            "Plan L5 title indented",
            worksheet["B8"].value,
            "            Example completed deliverable",
        ),
        ("Plan point Start blank", worksheet["D9"].value in {None, ""}, True),
        ("Plan point direct Due", worksheet["E9"].value, _dt(2026, 7, 20)),
        ("Plan point has exactly one diamond", key_date_glyphs.count("\u25c6"), 1),
        (
            "Plan point has no inherited interval glyph",
            any(glyph in interval_glyphs for glyph in key_date_glyphs),
            False,
        ),
        ("Plan level helper", worksheet["BH6"].value, 1),
        ("Plan category helper: done task", worksheet["BI8"].value, "D"),
        ("Plan axis starts 8 Jun", worksheet["F5"].value, _dt(2026, 6, 8)),
    ]


def expected_value_checks(workbook: Workbook) -> list[ValueCheck]:
    """Return the representative calculated-value assertions.

    Returns:
        Label, observed value, and expected value triples.

    """
    return [
        *_item_checks(workbook["Items"]),
        *_raid_checks(workbook["RAID"]),
        ("Calc itemIDs first", workbook["Calc"]["A2"].value, "I-1001"),
        ("Calc people first", workbook["Calc"]["B2"].value, "Max"),
        ("Calc scope selectors[0]", workbook["Calc"]["M2"].value, "All"),
        *_overview_checks(workbook["Overview"]),
        *_plan_checks(workbook["Plan"]),
    ]


def _error_value_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    error_count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ERROR_VALUE_RE.match(cell.value):
                    error_count += 1
                    if error_count <= MAX_REPORTED_ERROR_VALUES:
                        failures.append(
                            f"error value {worksheet.title}!{cell.coordinate} = {cell.value}"
                        )
    if error_count > MAX_REPORTED_ERROR_VALUES:
        remaining = error_count - MAX_REPORTED_ERROR_VALUES
        failures.append(f"... and {remaining} more error values")
    return failures


def _expected_value_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    for name, got, want in expected_value_checks(workbook):
        normalized = normalise_cached_date(got, want, workbook.epoch)
        matches = normalized == want
        if isinstance(want, float) and isinstance(normalized, (int, float)):
            matches = abs(normalized - want) < FLOAT_TOLERANCE
        if not matches:
            failures.append(f"VALUE {name}: got {normalized!r} want {want!r}")
    return failures


def cached_value_failures(workbook: Workbook, *, check_expected: bool) -> list[str]:
    """Check formula error caches and optionally representative values.

    Returns:
        Every cached-value failure.

    """
    failures = _error_value_failures(workbook)
    if not check_expected:
        return failures
    return [*failures, *_expected_value_failures(workbook)]
