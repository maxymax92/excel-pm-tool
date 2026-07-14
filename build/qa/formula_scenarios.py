"""Value-check health and dependency formulas with a focused scenario."""

import sys
from datetime import date

import openpyxl

from ..pipeline import build_one
from ..spec import config
from .common import temporary_examples, temporary_workbook
from .excel import recalculate

FLOAT_TOLERANCE = 1e-9


def main() -> None:
    """Build and verify the focused health and dependency scenario."""
    fails: list[str] = []
    configured_health = list(config.DELIVERY_HEALTH)
    try:
        # The internal blank proves that the direct-blocked identity comes from
        # the final nonblank option rather than a position derived from COUNTA.
        config.DELIVERY_HEALTH = ["On track", "", "At risk", "Off track", "Blocked"]
        with temporary_examples() as examples, temporary_workbook("PM_scenario") as out:
            examples.ITEMS_EXAMPLES = [
                {
                    "ID": "I-1001",
                    "Title": "Root epic",
                    "Type": "Epic",
                    "Status": "In Progress",
                    "Owner": "Alice",
                    "Created": date(2024, 1, 1),
                },
                {
                    "ID": "I-1002",
                    "Title": "Predecessor (blocked)",
                    "Type": "Story",
                    "Parent": "I-1001",
                    "Status": "Backlog",
                    "Owner": "Alice",
                    "Due": date(2030, 6, 1),
                    "Delivery Health": "Blocked",
                    "Created": date(2024, 1, 1),
                },
                {
                    "ID": "I-1003",
                    "Title": "Blocked and overdue",
                    "Type": "Story",
                    "Parent": "I-1001",
                    "Status": "In Progress",
                    "Owner": "Alice",
                    "Start": date(2024, 1, 1),
                    "Due": date(2024, 6, 1),
                    "BlockedBy": "I-1002",
                    "Created": date(2024, 1, 1),
                    "InProgressSince": date(2024, 1, 1),
                },
                {
                    "ID": "I-1004",
                    "Title": "Overdue",
                    "Type": "Story",
                    "Parent": "I-1001",
                    "Status": "In Progress",
                    "Owner": "Alice",
                    "Due": date(2020, 1, 1),
                    "Created": date(2024, 1, 1),
                    "InProgressSince": date(2020, 1, 1),
                },
            ]
            examples.PEOPLE_EXAMPLES = [
                {"Person": "Alice", "Role": "PM", "Team": "Core"},
            ]
            examples.RAID_EXAMPLES = []
            build_one(out, with_vba=False)
            recalculate(out)

            workbook = openpyxl.load_workbook(out, data_only=True)
            try:
                items = workbook["Items"]
                columns = {cell.value: index + 1 for index, cell in enumerate(items[2])}

                def value(item_index: int, column: str) -> object:
                    return items.cell(row=item_index + 2, column=columns[column]).value

                checks = [
                    ("I-1001 Children", value(1, "Children"), 3),
                    ("I-1003 WaitingOn", value(3, "WaitingOn"), "I-1002"),
                    ("I-1003 IsBlocked", value(3, "IsBlocked"), True),
                    ("I-1003 Health", value(3, "Health"), "R"),
                    ("I-1004 Health", value(4, "Health"), "R"),
                    ("I-1002 IsBlocked across Config gap", value(2, "IsBlocked"), True),
                    ("I-1002 Health", value(2, "Health"), "A"),
                    ("I-1002 WaitingOn", value(2, "WaitingOn"), None),
                    ("I-1004 DueIn", (value(4, "DueIn") or 0) < 0, True),
                    (
                        "Config Delivery Health gap fixture",
                        [workbook["Config"].cell(row, 27).value for row in range(4, 9)],
                        [value or None for value in config.DELIVERY_HEALTH],
                    ),
                ]
                len(checks)
                for name, got, want in checks:
                    matches = (
                        abs(got - want) < FLOAT_TOLERANCE
                        if isinstance(want, float) and isinstance(got, (int, float))
                        else got == want
                    )
                    if not matches:
                        fails.append(f"{name}: got {got!r}, want {want!r}")
            finally:
                workbook.close()
    finally:
        config.DELIVERY_HEALTH = configured_health

    if fails:
        sys.stdout.write("FORMULA SCENARIO QA FAIL\n")
        sys.stdout.write("".join(f" - {failure}\n" for failure in fails))
        sys.exit(1)
    sys.stdout.write("FORMULA SCENARIO QA PASS\n")


if __name__ == "__main__":
    main()
