"""Regression contracts for deterministic plans and approved workbook apply."""

from __future__ import annotations

import asyncio
import json
import re
import sys
import tempfile
import unittest
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from unittest.mock import patch
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

import openpyxl
from defusedxml import ElementTree as DefusedET
from openpyxl.utils.cell import get_column_letter, range_boundaries

from .. import pipeline
from ..data import __main__ as data_cli
from ..data.apply import apply_workbook, evaluate_apply
from ..data.bridge import describe_workbook, plan_workbook
from ..data.contract import parse_changeset
from ..data.diagnostics import Diagnostic
from ..data.inject import injected_source, validate_snapshot
from ..paths import ROOT
from .test_agent_contract import _change_set
from .test_data_layer import ITEM_ROWS, RAID_ROWS, _snapshot

LONDON = ZoneInfo("Europe/London")


@dataclass(frozen=True, slots=True)
class _CliResult:
    returncode: int
    stdout: bytes
    stderr: bytes


async def _run_cli(arguments: tuple[str, ...], stdin: bytes | None) -> _CliResult:
    process = await asyncio.create_subprocess_exec(
        sys.executable,
        "-m",
        "build.data",
        *arguments,
        cwd=ROOT,
        stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await process.communicate(stdin)
    return _CliResult(process.returncode or 0, stdout, stderr)


def _payload(document: dict[str, object], *, pretty: bool = False) -> bytes:
    """Return one strict UTF-8 change-set payload.

    Returns:
        Compact or pretty JSON bytes with the same parsed meaning.

    """
    indent = 2 if pretty else None
    return json.dumps(
        document,
        ensure_ascii=False,
        allow_nan=False,
        indent=indent,
        sort_keys=pretty,
    ).encode()


def _upsert(*, title: str = "Rollout programme v2") -> dict[str, object]:
    return {
        "operation_id": "update-project",
        "op": "upsert",
        "entity": "item",
        "identity": {"workbook_id": "I-3001"},
        "set": {"Title": title},
    }


def _trim_table_columns(
    workbook: openpyxl.Workbook,
    sheet: str,
    table_name: str,
    count: int,
) -> None:
    worksheet = workbook[sheet]
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col - count)}{max_row}"
    table.ref = new_ref
    table.tableColumns = table.tableColumns[:-count]
    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref


class PlanContractTests(unittest.TestCase):
    """Bind exact read-only plans to state, warnings and local time."""

    directory: tempfile.TemporaryDirectory[str]
    workbook: Path
    morning = datetime(2026, 7, 15, 9, 30, tzinfo=LONDON)

    @classmethod
    def setUpClass(cls) -> None:
        """Build one current workbook for deterministic planning tests."""
        super().setUpClass()
        cls.directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(cls.directory.cleanup)
        cls.workbook = Path(cls.directory.name) / "planning.xlsx"
        snapshot = _snapshot()
        with injected_source(snapshot, validate_snapshot(snapshot)):
            pipeline.build_one(cls.workbook, with_vba=False)

    def _document(self, *operations: dict[str, object]) -> dict[str, object]:
        document = _change_set()
        document["target"] = describe_workbook(self.workbook, now=self.morning)["target"]
        document["operations"] = list(operations)
        return document

    def test_plan_returns_exact_diffs_time_boundary_and_token(self) -> None:
        """Expose reviewable field changes and the complete midnight boundary."""
        result = plan_workbook(
            _payload(self._document(_upsert())),
            self.workbook,
            now=self.morning,
        )

        self.assertTrue(result["valid"], result["errors"])
        self.assertFalse(result["conflict"])
        self.assertTrue(result["has_changes"])
        self.assertEqual(
            result["effective_time"],
            {
                "date": "2026-07-15",
                "timezone": "Europe/London",
                "utc_offset": "+01:00",
                "expires_at": "2026-07-16T00:00:00+01:00",
            },
        )
        self.assertEqual(
            result["summary"],
            {"create": 0, "update": 1, "mark_deleted": 0, "noop": 0, "field_changes": 2},
        )
        operation = result["operations"][0]
        self.assertEqual(operation["action"], "update")
        self.assertEqual(operation["workbook_id"], "I-3001")
        self.assertEqual(
            {diff["field"]: (diff["before"], diff["after"]) for diff in operation["diffs"]},
            {
                "Title": ("Rollout programme", "Rollout programme v2"),
                "Updated": ("2026-07-10", "2026-07-15"),
            },
        )
        self.assertEqual(result["errors"], [])
        self.assertTrue(result["warnings"])
        self.assertRegex(result["plan_token"], re.compile(r"^[0-9a-f]{64}$"))

    def test_token_is_stable_across_json_formatting_and_same_day_time(self) -> None:
        """Exclude whitespace, object order, export time and clock time within a day."""
        document = self._document(_upsert())
        first = plan_workbook(
            _payload(document),
            self.workbook,
            now=datetime(2026, 7, 15, 0, 1, tzinfo=LONDON),
        )
        second = plan_workbook(
            _payload(deepcopy(document), pretty=True),
            self.workbook,
            now=datetime(2026, 7, 15, 23, 59, tzinfo=LONDON),
        )
        self.assertEqual(first["plan_token"], second["plan_token"])

    def test_token_changes_with_effective_date_or_intended_state(self) -> None:
        """Bind the lifecycle date and complete intended authored state."""
        first = plan_workbook(
            _payload(self._document(_upsert())),
            self.workbook,
            now=self.morning,
        )
        next_day = plan_workbook(
            _payload(self._document(_upsert())),
            self.workbook,
            now=datetime(2026, 7, 16, 9, 30, tzinfo=LONDON),
        )
        changed = plan_workbook(
            _payload(self._document(_upsert(title="A different intended title"))),
            self.workbook,
            now=self.morning,
        )
        self.assertNotEqual(first["plan_token"], next_day["plan_token"])
        self.assertNotEqual(first["plan_token"], changed["plan_token"])

    def test_token_binds_warning_set(self) -> None:
        """Make warning acknowledgement part of the approved token."""
        payload = _payload(self._document(_upsert()))
        baseline = plan_workbook(payload, self.workbook, now=self.morning)
        added_warning = Diagnostic(
            code="test.warning",
            severity="warning",
            phase="plan",
            pointer="",
            operation_id=None,
            message="A deterministic warning for token binding.",
            hint="Review it.",
        )
        with patch("build.data.bridge._planning_warnings", return_value=(added_warning,)):
            changed = plan_workbook(payload, self.workbook, now=self.morning)
        self.assertNotEqual(baseline["plan_token"], changed["plan_token"])
        self.assertEqual(changed["warnings"][0]["code"], "test.warning")

    def test_target_mismatch_is_a_conflict_without_token(self) -> None:
        """Refuse planning against any stale describe target value."""
        for field in (
            "workbook_sha256",
            "workbook_schema_fingerprint",
            "build_schema_fingerprint",
        ):
            with self.subTest(field=field):
                document = self._document(_upsert())
                document["target"][field] = "f" * 64
                result = plan_workbook(_payload(document), self.workbook, now=self.morning)

                self.assertFalse(result["valid"])
                self.assertTrue(result["conflict"])
                self.assertNotIn("plan_token", result)
                self.assertEqual(result["errors"][0]["code"], f"target.{field}")
                self.assertEqual(result["operations"], [])

    def test_replayed_values_are_a_no_change_plan(self) -> None:
        """Return an approvable no-op without lifecycle or counter changes."""
        operation = _upsert(title="Rollout programme")
        operation["set"]["Due"] = "2026-09-30"
        result = plan_workbook(
            _payload(self._document(operation)),
            self.workbook,
            now=self.morning,
        )

        self.assertTrue(result["valid"], result["errors"])
        self.assertFalse(result["has_changes"])
        self.assertEqual(result["operations"][0]["action"], "noop")
        self.assertEqual(result["operations"][0]["diffs"], [])
        self.assertEqual(result["summary"]["noop"], 1)
        self.assertIn("plan_token", result)

    def test_new_validation_error_keeps_proposed_diffs_but_blocks_token(self) -> None:
        """Show the exact rejected change while keeping the batch atomic."""
        operation = _upsert()
        operation["set"] = {"Due": "2026-05-01"}
        result = plan_workbook(
            _payload(self._document(operation)),
            self.workbook,
            now=self.morning,
        )

        self.assertFalse(result["valid"])
        self.assertFalse(result["conflict"])
        self.assertNotIn("plan_token", result)
        self.assertEqual(result["operations"][0]["action"], "update")
        self.assertIn("Due", {diff["field"] for diff in result["operations"][0]["diffs"]})
        self.assertIn("date.order", {error["code"] for error in result["errors"]})

    def test_malformed_change_set_returns_structured_errors(self) -> None:
        """Keep strict parser diagnostics in the plan JSON boundary."""
        result = plan_workbook(b'{"contract":', self.workbook, now=self.morning)
        self.assertFalse(result["valid"])
        self.assertFalse(result["conflict"])
        self.assertIsNone(result["target"])
        self.assertEqual(result["errors"][0]["code"], "json.syntax")
        self.assertNotIn("plan_token", result)


class LegacyPlanTests(unittest.TestCase):
    """Report source-column and deletion-role migration before approval."""

    def test_legacy_shape_migration_is_reviewed_and_token_bound(self) -> None:
        """Plan current columns and roles without mutating the legacy workbook."""
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "legacy.xlsx"
            item_rows = tuple(
                {key: value for key, value in row.items() if key not in {"Source", "Source ID"}}
                for row in ITEM_ROWS
            )
            raid_rows = tuple(
                {key: value for key, value in row.items() if key not in {"Source", "Source ID"}}
                for row in RAID_ROWS
            )
            snapshot = _snapshot(tables={"tblItems": item_rows, "tblRAID": raid_rows})
            with injected_source(snapshot, validate_snapshot(snapshot)):
                pipeline.build_one(workbook_path, with_vba=False)

            workbook = openpyxl.load_workbook(workbook_path)
            _trim_table_columns(workbook, "Items", "tblItems", 2)
            _trim_table_columns(workbook, "RAID", "tblRAID", 2)
            _trim_table_columns(workbook, "Config", "tblStatuses", 1)
            _trim_table_columns(workbook, "Config", "tblRaidStatuses", 1)
            workbook.save(workbook_path)

            now = datetime(2026, 7, 15, 10, 0, tzinfo=LONDON)
            document = _change_set()
            document["target"] = describe_workbook(workbook_path, now=now)["target"]
            document["operations"] = []
            before = workbook_path.read_bytes()
            result = plan_workbook(_payload(document), workbook_path, now=now)

            self.assertTrue(result["valid"], result["errors"])
            self.assertTrue(result["has_changes"])
            self.assertEqual(result["operations"], [])
            adjustments = "\n".join(result["migration_adjustments"])
            self.assertIn("'Source'", adjustments)
            self.assertIn("'Source ID'", adjustments)
            self.assertIn("'IsDeleted'", adjustments)
            self.assertIn("added IsDeleted=False", adjustments)
            self.assertIn("plan_token", result)
            self.assertEqual(workbook_path.read_bytes(), before)


class ApplyContractTests(unittest.TestCase):
    """Replan approvals and keep every rejected apply free of publication."""

    directory: tempfile.TemporaryDirectory[str]
    workbook: Path
    morning = datetime(2026, 7, 15, 9, 30, tzinfo=LONDON)

    @classmethod
    def setUpClass(cls) -> None:
        """Build one macro-enabled workbook for apply and CLI guards."""
        super().setUpClass()
        cls.directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(cls.directory.cleanup)
        cls.workbook = Path(cls.directory.name) / "apply.xlsm"
        snapshot = _snapshot()
        with injected_source(snapshot, validate_snapshot(snapshot)):
            pipeline.build_one(cls.workbook, with_vba=True)

    def _document(self, *operations: dict[str, object]) -> dict[str, object]:
        document = _change_set()
        document["target"] = describe_workbook(self.workbook, now=self.morning)["target"]
        document["operations"] = list(operations)
        return document

    def test_approved_apply_replans_and_uses_existing_publication_pipeline(self) -> None:
        """Persist the pre-change snapshot then rebuild the intended state once."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        snapshot_path = Path(self.directory.name) / "pre.json"
        backup_path = Path(self.directory.name) / "backup.xlsm"

        with (
            patch("build.data.apply.require_current_vba") as require_vba,
            patch("build.data.apply.write_snapshot", return_value=snapshot_path) as write,
            patch(
                "build.data.apply.rebuild_and_publish",
                return_value=backup_path,
            ) as publish,
        ):
            result = apply_workbook(
                payload,
                self.workbook,
                approve=approved,
                now=self.morning,
            )

        self.assertTrue(result["valid"], result["errors"])
        self.assertTrue(result["applied"])
        self.assertEqual(result["publication"]["status"], "published")
        self.assertEqual(result["publication"]["snapshot"], str(snapshot_path))
        self.assertEqual(result["publication"]["backup"], str(backup_path))
        require_vba.assert_called_once_with()
        write.assert_called_once()
        publish.assert_called_once()
        intended = publish.call_args.args[1]
        self.assertEqual(intended.tables["tblItems"][0]["Title"], "Rollout programme v2")
        self.assertEqual(intended.tables["tblItems"][0]["Updated"].isoformat(), "2026-07-15")

    def test_noop_apply_creates_no_snapshot_backup_or_build(self) -> None:
        """Short-circuit an approved replay before every mutation side effect."""
        operation = _upsert(title="Rollout programme")
        payload = _payload(self._document(operation))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]

        with (
            patch("build.data.apply.require_current_vba") as require_vba,
            patch("build.data.apply.write_snapshot") as write,
            patch("build.data.apply.rebuild_and_publish") as publish,
        ):
            result = apply_workbook(
                payload,
                self.workbook,
                approve=approved,
                now=self.morning,
            )

        self.assertTrue(result["valid"])
        self.assertFalse(result["applied"])
        self.assertEqual(result["publication"]["status"], "no_change")
        require_vba.assert_not_called()
        write.assert_not_called()
        publish.assert_not_called()

    def test_wrong_or_expired_token_is_a_conflict_before_side_effects(self) -> None:
        """Reject both arbitrary and next-day approvals before build setup."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        scenarios = (
            ("0" * 64, self.morning),
            (approved, datetime(2026, 7, 16, 9, 30, tzinfo=LONDON)),
        )
        for token, now in scenarios:
            with (
                self.subTest(now=now),
                patch("build.data.apply.write_snapshot") as write,
                patch("build.data.apply.rebuild_and_publish") as publish,
            ):
                evaluation = evaluate_apply(
                    payload,
                    self.workbook,
                    approve=token,
                    now=now,
                )
                self.assertEqual(evaluation.exit_code, 3)
                self.assertTrue(evaluation.result["conflict"])
                self.assertFalse(evaluation.result["applied"])
                self.assertIn(
                    "approval.token_mismatch",
                    {error["code"] for error in evaluation.result["errors"]},
                )
                write.assert_not_called()
                publish.assert_not_called()

    def test_changed_workbook_is_a_conflict_before_side_effects(self) -> None:
        """Recheck the change-set target during apply planning."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        original = self.workbook.read_bytes()
        try:
            self.workbook.write_bytes(original + b"\n")
            with (
                patch("build.data.apply.write_snapshot") as write,
                patch("build.data.apply.rebuild_and_publish") as publish,
            ):
                evaluation = evaluate_apply(
                    payload,
                    self.workbook,
                    approve=approved,
                    now=self.morning,
                )
                self.assertEqual(evaluation.exit_code, 3)
                self.assertTrue(evaluation.result["conflict"])
                write.assert_not_called()
                publish.assert_not_called()
        finally:
            self.workbook.write_bytes(original)

    def test_mid_publication_excel_lock_reports_close_and_retry(self) -> None:
        """Distinguish a newly opened workbook from a changed approved digest."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        snapshot_path = Path(self.directory.name) / "pre.json"
        lock = self.workbook.parent / f"~${self.workbook.name}"

        def open_during_publication(*_args: object) -> None:
            lock.touch()
            raise pipeline.PublicationPreconditionError(
                pipeline._PipelineProblem.PRECONDITION_PRESENT,
                lock,
            )

        try:
            with (
                patch("build.data.apply.require_current_vba"),
                patch("build.data.apply.write_snapshot", return_value=snapshot_path),
                patch(
                    "build.data.apply.rebuild_and_publish",
                    side_effect=open_during_publication,
                ),
            ):
                evaluation = evaluate_apply(
                    payload,
                    self.workbook,
                    approve=approved,
                    now=self.morning,
                )
        finally:
            if lock.exists():
                lock.unlink()

        self.assertEqual(evaluation.exit_code, 2)
        self.assertFalse(evaluation.result["conflict"])
        diagnostic = evaluation.result["errors"][-1]
        self.assertEqual(diagnostic["code"], "apply.workbook")
        self.assertIn("appears open in Excel", diagnostic["message"])
        self.assertIn("close every workbook in Excel", diagnostic["message"])
        self.assertEqual(
            diagnostic["hint"], "Supply the unlocked macro-enabled release workbook and retry."
        )
        self.assertNotIn("new exact digest", diagnostic["hint"])

    def test_publication_failure_is_exit_four_with_original_error(self) -> None:
        """Return a machine failure without replacing the original exception text."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        with (
            patch("build.data.apply.require_current_vba"),
            patch(
                "build.data.apply.write_snapshot",
                return_value=Path(self.directory.name) / "pre.json",
            ),
            patch(
                "build.data.apply.rebuild_and_publish",
                side_effect=OSError("disk full"),
            ),
        ):
            evaluation = evaluate_apply(
                payload,
                self.workbook,
                approve=approved,
                now=self.morning,
            )
        self.assertEqual(evaluation.exit_code, 4)
        self.assertFalse(evaluation.result["valid"])
        self.assertEqual(evaluation.result["errors"][-1]["code"], "publication.failed")
        self.assertIn("OSError: disk full", evaluation.result["errors"][-1]["message"])

    def test_publication_failure_reports_original_and_rollback_error_chain(self) -> None:
        """Keep the failed operation visible beside secondary recovery diagnostics."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        original = OSError("workbook replacement failed")
        recovery = RuntimeError("rollback failures: PermissionError: restore denied")
        recovery.__cause__ = original
        with (
            patch("build.data.apply.require_current_vba"),
            patch(
                "build.data.apply.write_snapshot",
                return_value=Path(self.directory.name) / "pre.json",
            ),
            patch("build.data.apply.rebuild_and_publish", side_effect=recovery),
        ):
            evaluation = evaluate_apply(
                payload,
                self.workbook,
                approve=approved,
                now=self.morning,
            )

        message = evaluation.result["errors"][-1]["message"]
        self.assertEqual(evaluation.exit_code, 4)
        self.assertIn("RuntimeError: rollback failures: PermissionError: restore denied", message)
        self.assertIn("OSError: workbook replacement failed", message)

    def test_corrupt_calculated_package_is_exit_four(self) -> None:
        """Report package verification failures through the machine apply result."""
        payload = _payload(self._document(_upsert()))
        approved = plan_workbook(payload, self.workbook, now=self.morning)["plan_token"]
        with (
            patch("build.data.apply.require_current_vba"),
            patch(
                "build.data.apply.write_snapshot",
                return_value=Path(self.directory.name) / "pre.json",
            ),
            patch(
                "build.data.apply.rebuild_and_publish",
                side_effect=BadZipFile("calculated workbook is corrupt"),
            ),
        ):
            evaluation = evaluate_apply(
                payload,
                self.workbook,
                approve=approved,
                now=self.morning,
            )

        self.assertEqual(evaluation.exit_code, 4)
        self.assertIn(
            "BadZipFile: calculated workbook is corrupt",
            evaluation.result["errors"][-1]["message"],
        )


class PublicationPreconditionTests(unittest.TestCase):
    """Capture the exact approved source bytes inside publication setup."""

    def test_stale_digest_creates_no_backup_or_publication_temporary(self) -> None:
        """Reject the transaction before applying either destination."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "PM_Workbook.xlsm"
            replacement = root / "calculated.xlsm"
            backup = root / "backup.xlsm"
            target.write_bytes(b"current")
            replacement.write_bytes(b"replacement")

            with self.assertRaises(pipeline.PublicationPreconditionError):
                pipeline.publish_transaction(
                    {backup: target, target: replacement},
                    expected_digests={target: "0" * 64},
                )

            self.assertEqual(target.read_bytes(), b"current")
            self.assertFalse(backup.exists())
            self.assertEqual(list(root.glob(".*.publish.*")), [])

    def test_lock_created_during_preparation_blocks_before_publication(self) -> None:
        """Recheck Excel's lock immediately before replacing the live workbook."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "PM_Workbook.xlsm"
            replacement = root / "calculated.xlsm"
            backup = root / "backup.xlsm"
            lock = root / "~$PM_Workbook.xlsm"
            target.write_bytes(b"approved")
            replacement.write_bytes(b"replacement")
            prepare = pipeline._prepare_publication

            def prepare_then_open(
                plan: pipeline.PublicationPlan,
            ) -> dict[Path, Path]:
                prepared = prepare(plan)
                lock.touch()
                return prepared

            with (
                patch(
                    "build.pipeline._prepare_publication",
                    side_effect=prepare_then_open,
                ),
                self.assertRaisesRegex(
                    pipeline.PublicationPreconditionError,
                    "precondition path exists",
                ),
            ):
                pipeline.publish_transaction(
                    {backup: target, target: replacement},
                    expected_digests={target: sha256(b"approved").hexdigest()},
                    required_absent=(lock,),
                )

            self.assertEqual(target.read_bytes(), b"approved")
            self.assertFalse(backup.exists())
            self.assertEqual(list(root.glob(".*.publish.*")), [])


class LiteralTextPackageTests(unittest.TestCase):
    """Keep agent-authored strings literal in the generated Excel package."""

    def test_formula_prefixes_and_urls_are_text_without_hyperlinks(self) -> None:
        """Disable XlsxWriter inference while preserving explicit workbook formulas."""
        literals = (
            "=1+1",
            "+SUM(A1:A2)",
            "-cmd|' /C calc'!A0",
            "@SUM(1,1)",
            "https://example.com/status?item=I-3001",
        )
        item_rows = [dict(row) for row in ITEM_ROWS]
        item_rows[0]["Title"] = literals[0]
        item_rows[0]["Latest Status"] = literals[1]
        item_rows[1]["Title"] = literals[2]
        item_rows[1]["Latest Status"] = literals[3]
        item_rows[2]["Title"] = literals[4]
        snapshot = _snapshot(tables={"tblItems": tuple(item_rows)})

        with tempfile.TemporaryDirectory() as directory:
            package = Path(directory) / "literal-text.xlsx"
            with injected_source(snapshot, validate_snapshot(snapshot)):
                pipeline.build_one(package, with_vba=False)

            workbook = openpyxl.load_workbook(package, data_only=False)
            items = workbook["Items"]
            literal_cells = [
                cell for row in items.iter_rows() for cell in row if cell.value in literals
            ]
            self.assertEqual({cell.value for cell in literal_cells}, set(literals))
            self.assertTrue(all(cell.data_type == "s" for cell in literal_cells))
            self.assertTrue(all(cell.hyperlink is None for cell in literal_cells))
            self.assertTrue(
                any(
                    cell.data_type == "f"
                    for sheet in (workbook["Overview"], workbook["Calc"])
                    for row in sheet.iter_rows()
                    for cell in row
                )
            )

            with ZipFile(package) as archive:
                shared = DefusedET.fromstring(archive.read("xl/sharedStrings.xml"))
                namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                shared_values = {
                    "".join(node.text or "" for node in item.iter(f"{namespace}t"))
                    for item in shared.findall(f"{namespace}si")
                }
                self.assertTrue(set(literals).issubset(shared_values))
                relationships = DefusedET.fromstring(
                    archive.read("xl/worksheets/_rels/sheet3.xml.rels")
                )
                self.assertFalse(
                    any(
                        relationship.attrib.get("Type", "").endswith("/hyperlink")
                        for relationship in relationships
                    )
                )


class PublicationTransactionTests(unittest.TestCase):
    """Publish the approved workbook and backup as one rollback unit."""

    def test_approved_bytes_become_transactional_backup(self) -> None:
        """Publish replacement and exact rollback backup as one transaction."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "PM_Workbook.xlsm"
            replacement = root / "calculated.xlsm"
            backup = root / "backup.xlsm"
            target.write_bytes(b"approved")
            replacement.write_bytes(b"replacement")

            pipeline.publish_transaction(
                {backup: target, target: replacement},
                expected_digests={target: sha256(b"approved").hexdigest()},
            )

            self.assertEqual(backup.read_bytes(), b"approved")
            self.assertEqual(target.read_bytes(), b"replacement")

    def test_partial_publication_restores_workbook_and_removes_backup(self) -> None:
        """Roll back the backup destination when the workbook replacement fails."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "PM_Workbook.xlsm"
            replacement = root / "calculated.xlsm"
            backup = root / "backup.xlsm"
            target.write_bytes(b"approved")
            replacement.write_bytes(b"replacement")

            def fail_after_backup(
                _plan: pipeline.PublicationPlan,
                prepared: dict[Path, Path],
            ) -> None:
                prepared[backup].replace(backup)
                del prepared[backup]
                message = "workbook replacement failed"
                raise OSError(message)

            with (
                patch("build.pipeline._apply_publication", side_effect=fail_after_backup),
                self.assertRaisesRegex(pipeline.PublicationError, "every destination was restored"),
            ):
                pipeline.publish_transaction(
                    {backup: target, target: replacement},
                    expected_digests={target: sha256(b"approved").hexdigest()},
                )

            self.assertEqual(target.read_bytes(), b"approved")
            self.assertFalse(backup.exists())
            self.assertEqual(list(root.glob(".*.publish.*")), [])


class AgentCliTests(unittest.TestCase):
    """Expose strict JSON-only command results and documented exit codes."""

    directory: tempfile.TemporaryDirectory[str]
    workbook: Path
    now = datetime(2026, 7, 15, 9, 30, tzinfo=LONDON)

    @classmethod
    def setUpClass(cls) -> None:
        """Build one workbook for isolated subprocess commands."""
        super().setUpClass()
        cls.directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(cls.directory.cleanup)
        cls.workbook = Path(cls.directory.name) / "cli.xlsm"
        snapshot = _snapshot()
        with injected_source(snapshot, validate_snapshot(snapshot)):
            pipeline.build_one(cls.workbook, with_vba=True)

    @staticmethod
    def _run(
        *arguments: str,
        stdin: bytes | None = None,
    ) -> _CliResult:
        return asyncio.run(_run_cli(arguments, stdin))

    def _change_file(self) -> Path:
        document = _change_set()
        document["target"] = describe_workbook(self.workbook, now=self.now)["target"]
        document["operations"] = [
            {
                "operation_id": "replay",
                "op": "upsert",
                "entity": "item",
                "identity": {"workbook_id": "I-3001"},
                "set": {"Title": "Rollout programme"},
            }
        ]
        path = Path(self.directory.name) / "changes.json"
        path.write_bytes(_payload(document))
        return path

    def test_describe_and_plan_emit_one_json_document(self) -> None:
        """Keep successful stdout machine-readable and operational text on stderr."""
        described = self._run("describe", str(self.workbook))
        self.assertEqual(described.returncode, 0, described.stderr.decode())
        self.assertEqual(json.loads(described.stdout)["result"], "describe")

        planned = self._run("plan", str(self._change_file()), str(self.workbook))
        self.assertEqual(planned.returncode, 0, planned.stderr.decode())
        self.assertEqual(json.loads(planned.stdout)["result"], "plan")

    def test_stale_target_and_wrong_apply_token_exit_three(self) -> None:
        """Reserve conflict status for target and approval preconditions."""
        change_file = self._change_file()
        document = json.loads(change_file.read_bytes())
        document["target"]["workbook_sha256"] = "e" * 64
        change_file.write_bytes(_payload(document))
        stale = self._run("plan", str(change_file), str(self.workbook))
        self.assertEqual(stale.returncode, 3)
        self.assertTrue(json.loads(stale.stdout)["conflict"])

        change_file = self._change_file()
        wrong = self._run(
            "apply",
            str(change_file),
            str(self.workbook),
            "--approve",
            "0" * 64,
        )
        self.assertEqual(wrong.returncode, 3)
        self.assertTrue(json.loads(wrong.stdout)["conflict"])

    def test_output_cannot_overwrite_workbook_or_change_set(self) -> None:
        """Reject destructive output aliases before opening the destination."""
        change_file = self._change_file()
        original = change_file.read_bytes()
        collided = self._run(
            "plan",
            str(change_file),
            str(self.workbook),
            "--output",
            str(change_file),
        )
        self.assertEqual(collided.returncode, 2)
        self.assertEqual(change_file.read_bytes(), original)
        self.assertEqual(json.loads(collided.stdout)["errors"][0]["code"], "cli.output_collision")

        workbook_collision = self._run(
            "describe",
            str(self.workbook),
            "--output",
            str(self.workbook),
        )
        self.assertEqual(workbook_collision.returncode, 2)

    def test_argument_errors_are_structured_exit_two_results(self) -> None:
        """Avoid mixing argparse usage text into the machine result stream."""
        missing_approval = self._run(
            "apply",
            str(self._change_file()),
            str(self.workbook),
        )
        self.assertEqual(missing_approval.returncode, 2)
        self.assertEqual(missing_approval.stderr, b"")
        result = json.loads(missing_approval.stdout)
        self.assertEqual(result["result"], "cli")
        self.assertEqual(result["errors"][0]["code"], "cli.arguments")

    def test_dash_input_and_atomic_output_file(self) -> None:
        """Read stdin and emit the result only through the requested file."""
        payload = self._change_file().read_bytes()
        output = Path(self.directory.name) / "result.json"
        planned = self._run(
            "plan",
            "-",
            str(self.workbook),
            "--output",
            str(output),
            stdin=payload,
        )
        self.assertEqual(planned.returncode, 0, planned.stderr.decode())
        self.assertEqual(planned.stdout, b"")
        self.assertEqual(planned.stderr.decode(), f"result: {output.resolve()}\n")
        self.assertEqual(json.loads(output.read_bytes())["result"], "plan")

    def test_output_replace_failure_preserves_destination_and_cleans_temporary(self) -> None:
        """Keep an existing JSON result intact when atomic replacement fails."""
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "result.json"
            original = b'{"result":"previous"}\n'
            output.write_bytes(original)

            with (
                patch(
                    "build.data.snapshot.Path.replace",
                    side_effect=OSError("replace blocked"),
                ),
                self.assertRaisesRegex(OSError, "replace blocked"),
            ):
                data_cli._emit({"result": "plan"}, output)

            self.assertEqual(output.read_bytes(), original)
            self.assertEqual(list(root.glob(".result.*.json")), [])

    def test_bridge_command_signatures_match_the_public_contract(self) -> None:
        """Pin optional workbooks/output and apply's required approval token."""
        expected = {
            "describe": "usage: build.data describe [-h] [--output OUTPUT] [workbook]",
            "plan": ("usage: build.data plan [-h] [--output OUTPUT] change_set [workbook]"),
            "apply": (
                "usage: build.data apply [-h] [--output OUTPUT] --approve APPROVE "
                "change_set [workbook]"
            ),
        }
        for command, usage in expected.items():
            with self.subTest(command=command):
                result = self._run(command, "--help")
                normalized = re.sub(r"\s+", " ", result.stdout.decode()).strip()
                self.assertEqual(result.returncode, 0, result.stderr.decode())
                self.assertIn(usage, normalized)

    def test_cli_surface_is_exactly_the_five_provider_neutral_commands(self) -> None:
        """Retain export/migrate beside the three bridge commands and nothing else."""
        result = self._run("--help")
        self.assertEqual(result.returncode, 0, result.stderr.decode())
        match = re.search(rb"\{(?P<commands>[^}]+)\}", result.stdout)
        self.assertIsNotNone(match)
        commands = set(match.group("commands").decode().split(","))
        self.assertEqual(commands, {"export", "migrate", "describe", "plan", "apply"})


class ProviderNeutralRepositoryTests(unittest.TestCase):
    """Keep runtime, release gates and user guidance provider-neutral."""

    def test_documented_change_set_is_complete_and_schema_valid(self) -> None:
        """Parse the guide's full example through the production contract."""
        guide = (ROOT / "docs" / "agent-data-bridge.md").read_text(encoding="utf-8")
        match = re.search(
            r"## Complete change-set example.*?```json\n(?P<json>.*?)\n```",
            guide,
            flags=re.DOTALL,
        )
        self.assertIsNotNone(match)
        document = parse_changeset(match.group("json").encode())
        self.assertEqual(document["version"], "1.0.0")
        self.assertEqual(len(document["operations"]), 3)
        normalized_guide = re.sub(r"\s+", " ", guide)
        for required in (
            "## Describe",
            "## Plan, revise and apply",
            "Correct the change set and rerun `plan` until it is valid.",
            ".venv/bin/python -m build.data apply /tmp/pm-changes.json PM_Workbook.xlsm",
            "## Paste-ready agent prompt",
            "Stop and wait for my explicit approval.",
            "Do not infer approval from this prompt.",
            "After I approve that exact token, run:",
            "python -m build.data apply /tmp/pm-changes.json",
        ):
            with self.subTest(required=required):
                self.assertIn(required, normalized_guide)

    def test_runtime_release_and_guides_have_no_provider_specific_contract(self) -> None:
        """Limit the product surface to describe, plan, apply, export and migrate."""
        forbidden = (
            "monday" + ".com",
            "MONDAY" + "_API_TOKEN",
            "dist/" + "monday",
            "test_" + "monday_import",
        )
        sources = [
            ROOT / "README.md",
            ROOT / "build" / "qa" / "release.py",
            *(ROOT / "docs").rglob("*.md"),
            *(ROOT / "build" / "data").rglob("*.py"),
        ]
        for source in sources:
            content = source.read_text(encoding="utf-8")
            for token in forbidden:
                with self.subTest(source=source.relative_to(ROOT), token=token):
                    self.assertNotIn(token, content)

        provider_module = ROOT / "build" / "data" / ("monday" + ".py")
        provider_tests = ROOT / "build" / "qa" / ("test_" + "monday_import.py")
        self.assertFalse(provider_module.exists())
        self.assertFalse(provider_tests.exists())


if __name__ == "__main__":
    unittest.main()
