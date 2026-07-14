"""Design-system QA for the generated workbook packages.

This suite deliberately tests stable presentation contracts rather than
formula results.  It proves that the central design tokens remain accessible,
the OOXML package carries the current Office theme signature, macro action
shapes keep their branded DrawingML styling and assigned VBA action, and every visible sheet
opens with the intended workbook chrome.

Usage:
    .venv/bin/python -m build.qa.design [workbook ...]

With no arguments the suite requires both release files in ``dist/``.
"""

from __future__ import annotations

import math
import re
import sys
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

if TYPE_CHECKING:
    from collections.abc import Iterator

    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.colors import Color
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.worksheet import Worksheet

from ..core.design import (
    ALIGNMENT,
    COLORS,
    MACRO_ACTIONS,
    OFFICE_THEME_SIGNATURE,
    ROWS,
    TYPOGRAPHY,
    WORKBOOK_WINDOW,
)
from ..paths import DIST
from ..spec.capacity import CONFIG_ROWS, DATA_ROWS
from ..vba.registry import MODULES

AA_NORMAL_TEXT = 4.5
AA_NON_TEXT = 3.0
ARGB_LENGTH = 8
SRGB_LINEAR_THRESHOLD = 0.04045
OPENING_WINDOW_SCALE = 0.75
MAX_HIERARCHY_LEVEL = 6
XML_ATTRIBUTE = re.compile(
    rb"(?P<name>[A-Za-z_][\w:.-]*)\s*=\s*(?P<quote>['\"])(?P<value>.*?)(?P=quote)",
    re.DOTALL,
)


# These are the design system's actual normal-size text/background pairings.
# Checking the aliases (rather than a duplicate list of hex values) means a
# future token edit is evaluated automatically.
CONTRAST_PAIRS = (
    ("text", "surface"),
    ("text_secondary", "surface"),
    ("text_muted", "canvas"),
    ("formula_fg", "formula_bg"),
    ("header_fg", "header_bg"),
    ("brand", "surface"),
    ("info_fg", "info_bg"),
    ("info_soft_fg", "info_soft_bg"),
    ("rag_g_fg", "rag_g_bg"),
    ("rag_a_fg", "rag_a_bg"),
    ("rag_r_fg", "rag_r_bg"),
    ("example_fg", "example_bg"),
    ("bar_done_fg", "bar_done_bg"),
    ("bar_active_fg", "bar_active_bg"),
    ("bar_plan_fg", "bar_plan_bg"),
    ("bar_over_fg", "bar_over_bg"),
    ("bar_cancel_fg", "bar_cancel_bg"),
)

NON_TEXT_CONTRAST_PAIRS = (("border_strong", "input_bg"),)


@dataclass(frozen=True)
class SheetCheck:
    """Workbook-sheet context shared by focused design assertions."""

    path: Path
    sheet: Worksheet
    fails: list[str]


@dataclass(frozen=True)
class PackageCheck:
    """OOXML-package context shared by focused package assertions."""

    path: Path
    package: zipfile.ZipFile
    fails: list[str]


def normalise_hex(value: str) -> str:
    """Return an uppercase six-digit RGB value.

    Returns:
        The normalized six-digit RGB value.

    Raises:
        ValueError: If ``value`` is not a six- or eight-digit RGB value.

    """
    text = str(value).strip().lstrip("#").upper()
    if len(text) == ARGB_LENGTH:  # OOXML/openpyxl ARGB -> RGB
        text = text[2:]
    if not re.fullmatch(r"[0-9A-F]{6}", text):
        msg = f"not a six-digit RGB colour: {value!r}"
        raise ValueError(msg)
    return text


def relative_luminance(value: str) -> float:
    """Return the WCAG relative luminance for an RGB colour.

    Returns:
        Relative luminance in the inclusive range zero to one.

    """
    rgb = [int(normalise_hex(value)[i : i + 2], 16) / 255 for i in (0, 2, 4)]
    linear = [
        channel / 12.92 if channel <= SRGB_LINEAR_THRESHOLD else ((channel + 0.055) / 1.055) ** 2.4
        for channel in rgb
    ]
    return 0.2126 * linear[0] + 0.7152 * linear[1] + 0.0722 * linear[2]


def contrast_ratio(foreground: str, background: str) -> float:
    """Return the WCAG contrast ratio between two RGB colours.

    Returns:
        The contrast ratio, where one represents no contrast.

    """
    light, dark = sorted(
        (relative_luminance(foreground), relative_luminance(background)),
        reverse=True,
    )
    return (light + 0.05) / (dark + 0.05)


def token_checks(fails: list[str]) -> None:
    """Validate every normal-text semantic colour pairing against WCAG AA."""
    _contrast_token_checks(fails)
    _typography_token_checks(fails)
    _macro_token_checks(fails)
    _vba_source_token_checks(fails)


def _contrast_token_checks(fails: list[str]) -> None:
    """Validate text and non-text semantic colour pairs."""
    for foreground, background in CONTRAST_PAIRS:
        missing = [name for name in (foreground, background) if name not in COLORS]
        if missing:
            fails.append(f"tokens: missing colour alias(es): {', '.join(missing)}")
            continue
        try:
            ratio = contrast_ratio(COLORS[foreground], COLORS[background])
        except ValueError as exc:
            fails.append(f"tokens: {foreground}/{background}: {exc}")
            continue
        if ratio + 1e-9 < AA_NORMAL_TEXT:
            fails.append(
                f"tokens: {foreground} on {background} is {ratio:.2f}:1; "
                f"normal text requires at least {AA_NORMAL_TEXT:.1f}:1"
            )

    for foreground, background in NON_TEXT_CONTRAST_PAIRS:
        ratio = contrast_ratio(COLORS[foreground], COLORS[background])
        if ratio + 1e-9 < AA_NON_TEXT:
            fails.append(
                f"tokens: {foreground} against {background} is {ratio:.2f}:1; "
                f"control boundaries require at least {AA_NON_TEXT:.1f}:1"
            )


def _typography_token_checks(fails: list[str]) -> None:
    """Validate the typography and alignment token contracts."""
    if TYPOGRAPHY.get("body_font") != OFFICE_THEME_SIGNATURE["minor_font"]:
        fails.append(
            "tokens: body font and Office minor theme font do not agree "
            f"({TYPOGRAPHY.get('body_font')!r} vs "
            f"{OFFICE_THEME_SIGNATURE['minor_font']!r})"
        )
    if TYPOGRAPHY.get("display_font") != OFFICE_THEME_SIGNATURE["major_font"]:
        fails.append(
            "tokens: display font and Office major theme font do not agree "
            f"({TYPOGRAPHY.get('display_font')!r} vs "
            f"{OFFICE_THEME_SIGNATURE['major_font']!r})"
        )
    required_alignment_roles = {
        "text",
        "narrative",
        "number",
        "date",
        "panel_text",
        "panel_date",
        "control",
        "axis",
        "metadata",
    }
    missing_roles = required_alignment_roles - set(ALIGNMENT)
    if missing_roles:
        fails.append(f"tokens: missing semantic alignment roles {sorted(missing_roles)}")


def _macro_token_checks(fails: list[str]) -> None:
    """Validate the Overview export-action token contract."""
    export_action = MACRO_ACTIONS.get("export", {})
    for key, expected in (
        ("cell", "A32"),
        ("caption", "Export to Markdown"),
    ):
        if export_action.get(key) != expected:
            fails.append(
                f"tokens: Overview export {key} is "
                f"{export_action.get(key)!r}; expected {expected!r}"
            )


def _vba_source_token_checks(fails: list[str]) -> None:
    """Validate design-relevant workbook-event source contracts."""
    missing = [module.filename for module in MODULES if not module.path.exists()]
    if missing:
        fails.extend(f"VBA source: {filename} is missing" for filename in missing)
        return
    sources = {module.name: module.path.read_text(encoding="utf-8") for module in MODULES}
    events = sources["ThisWorkbook"]
    combined = "\n".join(sources.values())
    if "Application.WindowState = xlMaximized" not in events:
        fails.append("VBA source: workbook-open handler does not maximise the Excel window")
    if "ThisWorkbook.ForceFullCalculation = False" not in events:
        fails.append("VBA source: workbook-open handler does not disable forced calculation")
    if "ITEMS_DELIVERY_HEALTH_COLUMN" not in events:
        fails.append("VBA source: the Delivery Health column is not mapped")
    if "PMTool.IsBlockedDeliveryHealth" not in events:
        fails.append("VBA source: Delivery Health does not drive blocked-state stamping")
    obsolete_contracts = (
        "ITEMS_BLOCKED_COLUMN",
        "PMTool.EnsureItemCheckboxes",
        "PMTool.EnsureCheckboxCell",
    )
    fails.extend(
        f"VBA source: obsolete checkbox contract is present: {contract}"
        for contract in obsolete_contracts
        if contract in combined
    )


def _xml_attributes(tag: bytes) -> dict[str, str]:
    """Return decoded attributes from one deterministic OOXML start tag.

    Returns:
        A mapping of qualified attribute names to decoded values.

    """
    return {
        match.group("name").decode("ascii"): unescape(match.group("value").decode("utf-8"))
        for match in XML_ATTRIBUTE.finditer(tag)
    }


def _xml_tag(data: bytes, local_name: str) -> bytes | None:
    """Return the first start tag with the requested OOXML local name.

    Returns:
        The matching start tag, or ``None`` when it is absent.

    """
    name = re.escape(local_name.encode("ascii"))
    match = re.search(rb"<(?:[A-Za-z_][\w.-]*:)?" + name + rb"\b[^>]*>", data)
    return None if match is None else match.group(0)


def _xml_section(data: bytes, local_name: str) -> bytes | None:
    """Return the first complete OOXML element with the requested local name.

    Returns:
        The matching element bytes, or ``None`` when it is absent.

    """
    name = re.escape(local_name.encode("ascii"))
    prefix = rb"(?:[A-Za-z_][\w.-]*:)?"
    match = re.search(
        rb"<" + prefix + name + rb"\b[^>]*>.*?</" + prefix + name + rb"\s*>",
        data,
        re.DOTALL,
    )
    return None if match is None else match.group(0)


def _xml_text(data: bytes, qualified_name: bytes) -> str | None:
    """Return UTF-8 element text for a qualified OOXML element.

    Returns:
        Decoded text, or ``None`` when the element is absent.

    """
    name = re.escape(qualified_name)
    match = re.search(rb"<" + name + rb"\b[^>]*>(.*?)</" + name + rb"\s*>", data, re.DOTALL)
    return None if match is None else unescape(match.group(1).decode("utf-8")).strip()


def _package_part(check: PackageCheck, member: str) -> bytes | None:
    """Read one UTF-8 OOXML part and record a diagnostic when unavailable.

    Returns:
        Raw part bytes, or ``None`` when the part cannot be checked.

    """
    if member not in check.package.namelist():
        check.fails.append(f"{check.path.name}: package has no {member}")
        return None
    data = check.package.read(member)
    try:
        data.decode("utf-8")
    except UnicodeDecodeError as exc:
        check.fails.append(f"{check.path.name}: {member} is not UTF-8: {exc}")
        return None
    if _xml_tag(data, "workbook") is None and not data.lstrip().startswith(b"<?xml"):
        check.fails.append(f"{check.path.name}: {member} is not recognizable OOXML")
        return None
    return data


def package_theme_checks(check: PackageCheck) -> None:
    """Assert the post-processed package has the modern Office signature."""
    label = check.path.name
    theme_name = "xl/theme/theme1.xml"
    data = _package_part(check, theme_name)
    if data is None:
        return
    major = _xml_section(data, "majorFont") or b""
    minor = _xml_section(data, "minorFont") or b""
    accent = _xml_section(data, "accent1") or b""
    dark = _xml_section(data, "dk2") or b""
    actual = {
        "major_font": _xml_attributes(_xml_tag(major, "latin") or b"").get("typeface", ""),
        "minor_font": _xml_attributes(_xml_tag(minor, "latin") or b"").get("typeface", ""),
        "accent1": _xml_attributes(_xml_tag(accent, "srgbClr") or b"").get("val", "").upper(),
        "dk2": _xml_attributes(_xml_tag(dark, "srgbClr") or b"").get("val", "").upper(),
    }
    for key, expected in OFFICE_THEME_SIGNATURE.items():
        if actual[key] != expected:
            check.fails.append(
                f"{label}: Office theme {key} is {actual[key] or '<missing>'!r}; "
                f"expected {expected!r}"
            )


def package_window_checks(check: PackageCheck) -> None:
    """Ensure the first-open workbook window is screen-sized."""
    part = "xl/workbook.xml"
    data = _package_part(check, part)
    if data is None:
        return
    view_tag = _xml_tag(data, "workbookView")
    if view_tag is None:
        check.fails.append(f"{check.path.name}: package has no workbookView")
        return
    view = _xml_attributes(view_tag)
    expected_width = WORKBOOK_WINDOW["width"] * 15
    expected_height = WORKBOOK_WINDOW["height"] * 15
    try:
        width = int(view.get("windowWidth", "0"))
        height = int(view.get("windowHeight", "0"))
    except ValueError:
        check.fails.append(f"{check.path.name}: workbookView has non-numeric geometry")
        return
    if width < int(expected_width * OPENING_WINDOW_SCALE) or height < int(
        expected_height * OPENING_WINDOW_SCALE
    ):
        check.fails.append(
            f"{check.path.name}: opening window is {width}x{height} twips; "
            "expected a screen-sized first-open window"
        )
    if view.get("minimized") in {"1", "true", "True"}:
        check.fails.append(f"{check.path.name}: workbook is configured to open minimized")


def package_calculation_checks(check: PackageCheck) -> None:
    """Require release artifacts to use Excel's normal dependency chain."""
    part = "xl/workbook.xml"
    data = _package_part(check, part)
    if data is None:
        return
    calc_tag = _xml_tag(data, "calcPr")
    if calc_tag is None:
        check.fails.append(f"{check.path.name}: package has no calcPr")
        return
    calc = _xml_attributes(calc_tag)
    if calc.get("fullCalcOnLoad") in {"1", "true", "True"}:
        check.fails.append(f"{check.path.name}: fullCalcOnLoad is enabled")
    if calc.get("forceFullCalc") in {"1", "true", "True"}:
        check.fails.append(f"{check.path.name}: forceFullCalc is enabled")


def package_formula_checks(check: PackageCheck) -> None:
    """Reject array-formula records that occupy any part of a merged range."""
    worksheet_members = sorted(
        name
        for name in check.package.namelist()
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
    )
    for member in worksheet_members:
        data = _package_part(check, member)
        if data is None:
            continue
        merged_ranges = [
            attributes["ref"]
            for tag in re.findall(rb"<(?:[A-Za-z_][\w.-]*:)?mergeCell\b[^>]*>", data)
            if (attributes := _xml_attributes(tag)).get("ref")
        ]
        if not merged_ranges:
            continue
        cells = re.finditer(
            rb"<(?:[A-Za-z_][\w.-]*:)?c\b(?P<attrs>[^>]*)>(?P<body>.*?)"
            rb"</(?:[A-Za-z_][\w.-]*:)?c\s*>",
            data,
            re.DOTALL,
        )
        for cell in cells:
            cell_attrs = _xml_attributes(b"<c " + cell.group("attrs") + b">")
            formula_tag = _xml_tag(cell.group("body"), "f")
            coordinate = cell_attrs.get("r")
            if formula_tag is None or not coordinate:
                continue
            if _xml_attributes(formula_tag).get("t") != "array":
                continue
            row, column = coordinate_to_tuple(coordinate)
            for merged_range in merged_ranges:
                min_column, min_row, max_column, max_row = range_boundaries(merged_range)
                if min_row <= row <= max_row and min_column <= column <= max_column:
                    check.fails.append(
                        f"{check.path.name}/{member}: array formula {coordinate} "
                        f"intersects merged range {merged_range}"
                    )


def _drawing_anchor(
    check: PackageCheck,
    drawing_parts: list[tuple[str, bytes]],
    description: str,
) -> tuple[int, int] | None:
    """Return the zero-based anchor for one accessible action description.

    Returns:
        The zero-based column and row, or ``None`` when no anchor matches.

    """
    for name, data in drawing_parts:
        for anchor in re.findall(
            rb"<xdr:(?:twoCellAnchor|oneCellAnchor)\b[^>]*>.*?"
            rb"</xdr:(?:twoCellAnchor|oneCellAnchor)\s*>",
            data,
            re.DOTALL,
        ):
            identity = _xml_tag(anchor, "cNvPr")
            if identity is None:
                continue
            if _xml_attributes(identity).get("descr") != description:
                continue
            origin = _xml_section(anchor, "from") or b""
            col_text = _xml_text(origin, b"xdr:col")
            row_text = _xml_text(origin, b"xdr:row")
            if col_text is None or row_text is None:
                check.fails.append(f"{check.path.name}: {name} has an incomplete action anchor")
                return None
            try:
                return int(col_text), int(row_text)
            except ValueError:
                check.fails.append(f"{check.path.name}: {name} has a non-numeric action anchor")
                return None
    return None


def _check_action_shape(
    check: PackageCheck,
    text: str,
    drawing_parts: list[tuple[str, bytes]],
    action: dict[str, str],
) -> None:
    """Validate one macro action's assignment, caption, description and anchor."""
    for required, label in (
        (f'macro="[0]!{action["macro"]}"', "macro assignment"),
        (f'descr="{action["description"]}"', "accessible description"),
        (f">{action['caption']}<", "visible caption"),
    ):
        if required not in text:
            check.fails.append(f"{check.path.name}: {action['caption']} action lacks {label}")

    min_col, min_row, _, _ = range_boundaries(action["cell"])
    expected_anchor = (min_col - 1, min_row - 1)
    actual_anchor = _drawing_anchor(check, drawing_parts, action["description"])
    if actual_anchor != expected_anchor:
        check.fails.append(
            f"{check.path.name}: {action['caption']} anchor is {actual_anchor!r}; "
            f"expected zero-based {expected_anchor!r} for {action['cell']}"
        )


def _check_action_palette(check: PackageCheck, drawing: bytes) -> None:
    """Validate macro action colours and typography."""
    lower = drawing.lower()
    for required, label in (
        (COLORS["brand"].lstrip("#").lower().encode(), "brand fill"),
        (COLORS["brand_dark"].lstrip("#").lower().encode(), "brand border"),
        (COLORS["header_fg"].lstrip("#").lower().encode(), "white text"),
        (TYPOGRAPHY["body_font"].lower().encode(), "Aptos text"),
    ):
        if required not in lower:
            check.fails.append(f"{check.path.name}: macro action shapes lack {label}")


def _check_vml_controls(check: PackageCheck) -> None:
    """Reject legacy VML Form Control buttons."""
    vml_names = [
        name
        for name in check.package.namelist()
        if name.startswith("xl/drawings/") and name.lower().endswith(".vml")
    ]
    for name in vml_names:
        raw = check.package.read(name)
        if re.search(rb"ObjectType\s*=\s*[\"']Button[\"']", raw, re.IGNORECASE):
            check.fails.append(f"{check.path.name}: {name} contains a VML Form Control button")


def package_button_checks(check: PackageCheck) -> None:
    """Ensure .xlsm actions are branded DrawingML shapes with live macros."""
    if check.path.suffix.lower() != ".xlsm":
        return
    drawing_names = [
        name
        for name in check.package.namelist()
        if re.match(r"^xl/drawings/drawing\d+\.xml$", name, re.IGNORECASE)
    ]
    drawing_parts = [(name, check.package.read(name)) for name in drawing_names]
    drawing = b"\n".join(data for _, data in drawing_parts)
    try:
        text = drawing.decode("utf-8")
    except UnicodeDecodeError as exc:
        check.fails.append(f"{check.path.name}: drawing XML is not UTF-8: {exc}")
        return
    for action in MACRO_ACTIONS.values():
        _check_action_shape(check, text, drawing_parts, action)
    _check_action_palette(check, drawing)
    _check_vml_controls(check)


def cell_rgb(color: Color | None) -> str | None:
    """Read an explicit cell RGB colour; theme/indexed colours return None.

    Returns:
        The six-digit RGB value, or ``None`` for non-RGB colours.

    """
    if color is None or color.type != "rgb" or color.rgb is None:
        return None
    return normalise_hex(color.rgb)


def expect_height(check: SheetCheck, row: int, expected: float, role: str) -> None:
    """Require one row to use the semantic design-system height."""
    actual = check.sheet.row_dimensions[row].height
    if actual is None or not math.isclose(float(actual), float(expected), abs_tol=0.1):
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}: {role} row {row} height is "
            f"{actual if actual is not None else '<default>'}; expected {expected} pt"
        )


def expect_fill(
    check: SheetCheck,
    coordinate: str,
    expected: str,
    role: str,
) -> None:
    """Require one cell to use the expected solid design-system fill."""
    cell = check.sheet[coordinate]
    actual = cell_rgb(cell.fill.fgColor)
    wanted = normalise_hex(expected)
    if cell.fill.fill_type != "solid" or actual != wanted:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{coordinate}: {role} fill is "
            f"{actual or '<not explicit solid RGB>'}; expected {wanted}"
        )


def validations_covering(ws: Worksheet, coordinate: str) -> list[DataValidation]:
    """Return data-validation rules whose ranges include coordinate.

    Returns:
        The validation rules applied to the requested coordinate.

    """
    collection = ws.data_validations
    return [
        validation
        for validation in collection.dataValidation
        if any(coordinate in cell_range for cell_range in validation.sqref.ranges)
    ]


def expect_stop_validation(check: SheetCheck, coordinate: str, role: str) -> None:
    """Require one input to use visible stop-style validation."""
    rules = validations_covering(check.sheet, coordinate)
    if not rules:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{coordinate}: no {role} validation"
        )
        return
    # OOXML omits errorStyle when it uses Excel's default "stop" behavior;
    # openpyxl exposes that omission as None.
    if not any(
        rule.showErrorMessage is True and rule.errorStyle in {None, "stop"} for rule in rules
    ):
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{coordinate}: {role} validation is not "
            "a visible stop-style error"
        )


def expect_input_prompt(
    check: SheetCheck,
    coordinate: str,
    role: str,
    required_text: tuple[str, ...],
) -> None:
    """Require a visible selection prompt that explains the field contract."""
    rules = validations_covering(check.sheet, coordinate)
    if not rules:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{coordinate}: no {role} validation"
        )
        return
    for rule in rules:
        prompt = f"{rule.promptTitle or ''} {rule.prompt or ''}"
        if rule.showInputMessage is True and all(text in prompt for text in required_text):
            return
    check.fails.append(
        f"{check.path.name}/{check.sheet.title}!{coordinate}: {role} validation does not show "
        f"the required guidance {required_text}"
    )


def expect_alignment(
    check: SheetCheck,
    coordinate: str,
    horizontal: str,
    vertical: str,
    role: str,
) -> None:
    """Require one cell to use the semantic horizontal and vertical alignment."""
    cell = check.sheet[coordinate]
    actual_h = cell.alignment.horizontal
    actual_v = cell.alignment.vertical
    if actual_h != horizontal or actual_v != vertical:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{coordinate}: {role} alignment is "
            f"{actual_h or '<general>'}/{actual_v or '<general>'}; expected "
            f"{horizontal}/{vertical}"
        )


def expect_column_alignment(
    check: SheetCheck,
    column: str,
    horizontal: str,
    vertical: str,
    role: str,
) -> None:
    """Check the column-default format used by unmaterialised spill cells."""
    alignment = check.sheet.column_dimensions[column].alignment
    actual_h = alignment.horizontal
    actual_v = alignment.vertical
    if actual_h != horizontal or actual_v != vertical:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{column}:{column}: {role} alignment is "
            f"{actual_h or '<general>'}/{actual_v or '<general>'}; expected "
            f"{horizontal}/{vertical}"
        )


def style_color(color: Color | None) -> str | None:
    """Normalize an explicit openpyxl colour from a cell or differential style.

    Returns:
        The six-digit RGB value, or ``None`` for non-RGB colours.

    """
    if color is None or color.type != "rgb" or color.rgb is None:
        return None
    return normalise_hex(color.rgb)


def rule_colours(rule: Rule) -> set[str]:
    """Collect text/fill/border RGB values from one conditional-format rule.

    Returns:
        The explicit RGB values used by the rule's differential style.

    """
    result: set[str] = set()
    dxf = rule.dxf
    if dxf is None:
        return result
    if dxf.font is not None:
        value = style_color(dxf.font.color)
        if value:
            result.add(value)
    if dxf.fill is not None:
        for color in (dxf.fill.fgColor, dxf.fill.bgColor):
            value = style_color(color)
            # XlsxWriter/openpyxl represents the unused pattern foreground as
            # Transparent black is package metadata outside the authored palette.
            if value and value != "000000":
                result.add(value)
    if dxf.border is not None:
        for side_name in ("left", "right", "top", "bottom"):
            side = getattr(dxf.border, side_name, None)
            value = style_color(None if side is None else side.color)
            if value:
                result.add(value)
    return result


def matching_rules(ws: Worksheet, *needles: str) -> Iterator[Rule]:
    """Yield conditional-format rules whose formula contains every needle.

    Yields:
        Each conditional-format rule whose formula contains every needle.

    """
    for conditional_format in ws.conditional_formatting:
        rules = ws.conditional_formatting[conditional_format]
        for rule in rules:
            formula = " ".join(rule.formula or ())
            if all(needle in formula for needle in needles):
                yield rule


def expect_border_only_rules(
    check: SheetCheck,
    role: str,
    needles: tuple[str, ...],
) -> None:
    """Ruler DXFs may add borders only; font/fill would corrupt bar/axis ink."""
    rules = list(matching_rules(check.sheet, *needles))
    if not rules:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}: missing {role} border-only rule"
        )
        return
    for rule in rules:
        dxf = rule.dxf
        if dxf is None or dxf.border is None:
            check.fails.append(
                f"{check.path.name}/{check.sheet.title}: {role} rule has no border DXF"
            )
        if dxf is not None and (dxf.font is not None or dxf.fill is not None):
            check.fails.append(
                f"{check.path.name}/{check.sheet.title}: {role} rule changes font/fill; "
                "timeline rulers must be border-only"
            )


def expect_state_rule(
    check: SheetCheck,
    role: str,
    needles: tuple[str, ...],
    colour_tokens: tuple[str, ...],
    *,
    require_stop: bool = False,
) -> None:
    """Require one semantic conditional-format rule and its design colours."""
    rules = list(matching_rules(check.sheet, *needles))
    if not rules:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}: missing {role} conditional rule "
            f"({', '.join(needles)})"
        )
        return
    actual = set().union(*(rule_colours(rule) for rule in rules))
    expected = {normalise_hex(COLORS[token]) for token in colour_tokens}
    missing = expected - actual
    if missing:
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}: {role} rule is missing design colours "
            f"{sorted(missing)}; found {sorted(actual)}"
        )
    if require_stop and not any(rule.stopIfTrue for rule in rules):
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}: {role} rule must stop lower-priority fills"
        )


def _item_attention_checks(check: SheetCheck, headers: list[object]) -> None:
    """Validate hierarchy-independent attention rules on Items."""
    ws = check.sheet
    header_letters = {
        name: get_column_letter(headers.index(name) + 1)
        for name in ("Type", "Parent", "Level", "Children", "Owner", "Delivery Health")
    }
    parent_rules = list(matching_rules(ws, "COUNTIF(dvItemIDs"))
    if not parent_rules:
        check.fails.append(f"{check.path.name}/Items: missing invalid Parent rule")
    parent_ref = f"${header_letters['Parent']}3"
    for rule in parent_rules:
        formula = " ".join(rule.formula or ())
        if f'{parent_ref}<>""' not in formula or f'{parent_ref}=""' in formula:
            check.fails.append(f"{check.path.name}/Items: blank Parent must remain unhighlighted")

    _item_required_field_rules(check, header_letters, "Owner")
    _item_required_field_rules(check, header_letters, "Delivery Health")


def _item_required_field_rules(
    check: SheetCheck,
    header_letters: dict[str, str],
    field: str,
) -> None:
    """Validate one active-work required-field attention rule."""
    rules = list(
        matching_rules(
            check.sheet,
            "dvStatusActive",
            f'${header_letters[field]}3=""',
        )
    )
    if not rules:
        check.fails.append(f"{check.path.name}/Items: missing active-work {field} rule")
    forbidden = tuple(f"${header_letters[name]}3" for name in ("Type", "Level", "Children"))
    for rule in rules:
        formula = " ".join(rule.formula or ())
        if any(reference in formula for reference in forbidden):
            check.fails.append(f"{check.path.name}/Items: {field} attention depends on hierarchy")
        if field == "Delivery Health":
            _delivery_health_attention_style(check, rule)


def _delivery_health_attention_style(check: SheetCheck, rule: Rule) -> None:
    """Require missing Delivery Health attention to use an amber border only."""
    dxf = rule.dxf
    if dxf is None or dxf.border is None:
        check.fails.append(f"{check.path.name}/Items: Delivery Health attention has no border")
    if dxf is not None and (dxf.font is not None or dxf.fill is not None):
        check.fails.append(f"{check.path.name}/Items: missing Delivery Health must be border-only")
    if normalise_hex(COLORS["rag_a_fg"]) not in rule_colours(rule):
        check.fails.append(
            f"{check.path.name}/Items: Delivery Health attention is missing the amber border"
        )


def _items_state_checks(check: SheetCheck) -> None:
    """Validate Items fills, prompts and semantic conditional formatting."""
    ws = check.sheet
    expect_fill(check, "A3", COLORS["example_bg"], "example data")
    blank_guards = list(matching_rules(ws, '$A3=""', "COUNTA($B3:$L3)=0"))
    if not any(rule.stopIfTrue is True and rule.dxf is None for rule in blank_guards):
        check.fails.append(
            f"{check.path.name}/Items: missing no-format stop-if-true guard for empty rows"
        )
    if style_color(ws["A3"].font.color) != normalise_hex(COLORS["example_fg"]):
        check.fails.append(f"{check.path.name}/Items!A3: example text does not use example_fg")
    headers = [cell.value for cell in ws[2]]
    _check_system_fill(check, headers, "ParentTitle", "calculated data")
    _check_system_fill(check, headers, "Created", "VBA/system data")
    expect_state_rule(
        check,
        "stale Latest Status",
        ("dvStatusActive", "cfgStaleDays"),
        ("rag_a_bg", "rag_a_fg"),
    )
    _item_attention_checks(check, headers)
    for rank, label in ((1, "green"), (2, "amber"), (3, "red")):
        if not list(matching_rules(ws, "dvDeliveryHealth", str(rank))):
            check.fails.append(
                f"{check.path.name}/Items: Delivery Health has no {label} semantic rule"
            )
    expect_state_rule(check, "blocked identity", ("=TRUE",), ("rag_r_bg",))


def _check_system_fill(
    check: SheetCheck,
    headers: list[object],
    header: str,
    role: str,
) -> None:
    """Require one calculated or system field to use the formula surface."""
    if header not in headers:
        return
    cell = check.sheet.cell(3, headers.index(header) + 1)
    actual = cell_rgb(cell.fill.fgColor)
    if actual != normalise_hex(COLORS["formula_bg"]):
        check.fails.append(
            f"{check.path.name}/{check.sheet.title}!{cell.coordinate}: {role} fill "
            f"is {actual or '<none>'}; expected formula_bg"
        )


def _raid_state_checks(check: SheetCheck) -> None:
    """Validate RAID fills, prompts and semantic conditional formatting."""
    ws = check.sheet
    expect_fill(check, "A3", COLORS["example_bg"], "example data")
    blank_guards = list(matching_rules(ws, '$A3=""', "COUNTA($B3:$I3,$K3:$L3)=0"))
    if not any(rule.stopIfTrue is True and rule.dxf is None for rule in blank_guards):
        check.fails.append(
            f"{check.path.name}/RAID: missing no-format stop-if-true guard for empty rows"
        )
    expect_state_rule(
        check,
        "overdue review",
        ("TODAY()", "dvRaidStatuses"),
        ("rag_r_bg", "rag_r_fg"),
    )
    severity_rules = list(matching_rules(ws, "dvSeverity"))
    actual = set().union(*(rule_colours(rule) for rule in severity_rules))
    expected = {
        normalise_hex(COLORS[token])
        for token in (
            "rag_g_bg",
            "rag_g_fg",
            "rag_a_bg",
            "rag_a_fg",
            "rag_r_bg",
            "rag_r_fg",
        )
    }
    if expected - actual:
        check.fails.append(
            f"{check.path.name}/RAID: severity rules do not cover every semantic pair"
        )
    headers = [cell.value for cell in ws[2]]
    for field in ("Prob", "Impact"):
        coordinate = ws.cell(3, headers.index(field) + 1).coordinate
        expect_input_prompt(
            check,
            coordinate,
            f"{field} rating",
            ("1-5", "Probability \u00d7 Impact", "1-25"),
        )
    _check_system_fill(check, headers, "Raised", "VBA/system data")


def _formula_text(ws: Worksheet, coordinate: str) -> str:
    """Return formula text from a scalar or openpyxl array-formula value.

    Returns:
        The formula text, or an empty string for a blank cell.

    """
    raw = ws[coordinate].value
    return str(getattr(raw, "text", raw) or "")


def _require_formula_tokens(
    check: SheetCheck,
    coordinate: str,
    role: str,
    tokens: tuple[str, ...],
) -> str:
    """Require formula tokens and return the normalized formula text.

    Returns:
        The normalized formula text.

    """
    formula = _formula_text(check.sheet, coordinate)
    check.fails.extend(
        f"{check.path.name}/{check.sheet.title}!{coordinate}: {role} is missing {token}"
        for token in tokens
        if token not in formula
    )
    return formula


def _overview_state_checks(check: SheetCheck) -> None:
    """Validate Overview formulas and urgency-band presentation."""
    ws = check.sheet
    scope_formula = _require_formula_tokens(
        check,
        "A3",
        "scope panel",
        (
            "cfgExecutiveStatusMaxLevel",
            "tblItems[Level]",
            "tblItems[Scope]",
            "tblItems[Parent]",
            "tblItems[A5]",
            "dvDeliveryHealth",
        ),
    )
    if "Calc!$N" in scope_formula:
        check.fails.append(
            f"{check.path.name}/Overview!A3: formula references the Calc N-column helper grid"
        )
    _require_formula_tokens(
        check,
        "F3",
        "Top RAID",
        ("tblRAID[Score]", "cfgAlertSevScore", "lstAlertRaid"),
    )
    _require_formula_tokens(
        check,
        "M3",
        "Coming up",
        ("tblItems[IsPoint]", "lstDecisionRaid", "tblRAID[NextReview]", "TODAY()"),
    )
    if ws["M2"].value != "Milestones / Decisions / Deadlines":
        check.fails.append(f"{check.path.name}/Overview!M2: Coming up header is {ws['M2'].value!r}")
    for token, colours in (
        ("cfgComingUrgentDays", ("brand_dark", "danger_strong")),
        ("cfgComingSoonDays", ("brand", "header_fg")),
        ("cfgComingNearDays", ("info_bg", "info_fg")),
        ("cfgComingHorizonDays", ("info_soft_bg", "info_soft_fg")),
    ):
        expect_state_rule(
            check,
            f"Coming Up {token}",
            (token, "$P3"),
            colours,
            require_stop=True,
        )
    check.fails.extend(
        f"{check.path.name}/Overview!{column}:{column}: numeric date helper "
        "must be visually hidden with ;;;"
        for column in ("E", "L", "P")
        if ws.column_dimensions[column].number_format != ";;;"
    )


def _plan_state_checks(check: SheetCheck) -> None:
    """Validate Plan glyphs, bars, point dates and ruler formatting."""
    ws = check.sheet
    grid_formula = _formula_text(ws, "F6")
    grid_logic = grid_formula.replace("_xlpm.", "")
    if "wk=INDEX" in grid_formula or "wk = INDEX" in grid_formula:
        check.fails.append(f"{check.path.name}/Plan!F6: bar glyphs are emitted only once")
    check.fails.extend(
        f"{check.path.name}/Plan!F6: bar formula is missing {glyph!r}"
        for glyph in ("\u2713", "\u25cf", "!", "\u00d7", "\u2014", "\u25c6")
        if glyph not in grid_formula
    )
    if '(pt=TRUE)*(es="")' in grid_logic:
        check.fails.append(
            f"{check.path.name}/Plan!F6: key-date rendering incorrectly depends on EffStart"
        )
    if 'IF(pt=TRUE,IF((du<>"")*(du>=wk)*(du<wk+7),"\u25c6"' not in grid_logic:
        check.fails.append(
            f"{check.path.name}/Plan!F6: due-only rows do not prioritize the Due-week diamond"
        )
    for category, background, foreground, label in (
        ("D", "bar_done_bg", "bar_done_fg", "done bar"),
        ("A", "bar_active_bg", "bar_active_fg", "active bar"),
        ("P", "bar_plan_bg", "bar_plan_fg", "planned bar"),
        ("O", "bar_over_bg", "bar_over_fg", "overdue bar"),
        ("C", "bar_cancel_bg", "bar_cancel_fg", "cancelled bar"),
    ):
        expect_state_rule(
            check,
            label,
            ('F6<>"\u25c6"', f'$BI6="{category}"'),
            (background, foreground),
        )
    expect_state_rule(check, "key-date point", ('F6="\u25c6"',), ("pt_next",))
    blank_guards = list(matching_rules(ws, '$A6=""', '$A6="\u2014 none \u2014"'))
    if not any(rule.stopIfTrue is True and rule.dxf is None for rule in blank_guards):
        check.fails.append(
            f"{check.path.name}/Plan: missing no-format stop-if-true guard for unused rows"
        )
    expect_state_rule(check, "neutral Today rule", ("TODAY()", "F$5+7"), ("today",))
    expect_border_only_rules(check, "Today ruler", ("TODAY()", "F$5+7"))
    expect_border_only_rules(check, "month ruler", ("MONTH(F$5)", "MONTH(F$5-7)"))


def _config_state_checks(check: SheetCheck) -> None:
    """Validate Config inputs, prompts and Boolean paste-safety rules."""
    ws = check.sheet
    expect_fill(check, "B4", COLORS["input_bg"], "editable setting")
    setting_rows = {
        str(ws.cell(row=row, column=1).value): row
        for row in range(4, ws.max_row + 1)
        if ws.cell(row=row, column=1).value not in {None, ""}
    }
    dynamic_coordinates = (
        ("B4", "numeric setting"),
        (f"B{setting_rows.get('ItemIDPrefix', 0)}", "ID-prefix setting"),
        (
            f"B{setting_rows.get('ExecutiveStatusMaxLevel', 0)}",
            "scope-panel level setting",
        ),
        ("K4", "type level"),
        ("Y4", "severity threshold"),
    )
    for coordinate, role in dynamic_coordinates:
        if coordinate == "B0":
            check.fails.append(f"{check.path.name}/Config: missing {role}")
        else:
            expect_stop_validation(check, coordinate, role)
    expect_input_prompt(
        check,
        "Y4",
        "severity threshold",
        ("Probability \u00d7 Impact", "1-25"),
    )
    if not list(matching_rules(ws, "ISLOGICAL")):
        check.fails.append(f"{check.path.name}/Config: Boolean role paste-safety rules missing")


def state_and_data_checks(path: Path, wb: Workbook, fails: list[str]) -> None:
    """Prove that data surfaces and behavioural rules use the design system."""
    checks = {
        "Items": _items_state_checks,
        "RAID": _raid_state_checks,
        "Overview": _overview_state_checks,
        "Plan": _plan_state_checks,
        "Config": _config_state_checks,
    }
    for sheet_name, run_checks in checks.items():
        if sheet_name in wb.sheetnames:
            run_checks(SheetCheck(path, wb[sheet_name], fails))


def visible_sheet_chrome_checks(path: Path, wb: Workbook, fails: list[str]) -> None:
    """Check workbook-wide visible-sheet accessibility and opening state."""
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible:
        fails.append(f"{path.name}: workbook has no visible sheets")
        return
    for ws in visible:
        if ws.sheet_view.showGridLines is not False:
            fails.append(f"{path.name}/{ws.title}: gridlines are not hidden")
        # Excel omits zoomScale at its 100% default, so None and 100 are the
        # same intended opening state.
        if ws.sheet_view.zoomScale not in {None, 100}:
            fails.append(
                f"{path.name}/{ws.title}: opening zoom is {ws.sheet_view.zoomScale}%; expected 100%"
            )
        value = ws["A1"].value
        if value is None or (isinstance(value, str) and not value.strip()):
            fails.append(
                f"{path.name}/{ws.title}: A1 is blank; visible sheets need a "
                "meaningful title/section label for navigation and screen readers"
            )


def _page_title_checks(path: Path, wb: Workbook, fails: list[str]) -> None:
    """Validate the shared page-title rail on operational sheets."""
    for sheet in ("Plan", "Items", "RAID", "Config"):
        if sheet not in wb.sheetnames:
            fails.append(f"{path.name}: missing design surface {sheet!r}")
            continue
        ws = wb[sheet]
        check = SheetCheck(path, ws, fails)
        expect_height(check, 1, ROWS["page_title"], "page-title")
        title = ws["A1"]
        if title.font.name != TYPOGRAPHY["display_font"]:
            fails.append(
                f"{path.name}/{sheet}!A1: title font is {title.font.name!r}; "
                f"expected {TYPOGRAPHY['display_font']!r}"
            )
        if title.font.sz is None or not math.isclose(
            float(title.font.sz), float(TYPOGRAPHY["page_title"]), abs_tol=0.1
        ):
            fails.append(
                f"{path.name}/{sheet}!A1: title size is {title.font.sz!r}; "
                f"expected {TYPOGRAPHY['page_title']} pt"
            )
        if not title.font.bold:
            fails.append(f"{path.name}/{sheet}!A1: page title is not bold")


def _column_alignment_contract(
    check: SheetCheck,
    specifications: tuple[tuple[str, str, str, str], ...],
) -> None:
    """Validate column-header presence and representative body alignment."""
    headers = {cell.value: cell.column for cell in check.sheet[2]}
    for header, horizontal, vertical, role in specifications:
        if header not in headers:
            check.fails.append(f"{check.path.name}/{check.sheet.title}: missing {header!r} column")
            continue
        coordinate = check.sheet.cell(3, headers[header]).coordinate
        expect_alignment(check, coordinate, horizontal, vertical, role)


def _items_component_checks(check: SheetCheck) -> None:
    """Validate Items title content, capacity guard and column alignment."""
    ws = check.sheet
    stray = [
        ws.cell(1, column).value
        for column in range(4, 14)
        if isinstance(ws.cell(1, column).value, str)
        and ws.cell(1, column).value.strip()
        and not ws.cell(1, column).value.startswith("=")
    ]
    if stray:
        check.fails.append(f"{check.path.name}/Items row 1 contains explanatory copy: {stray!r}")
    if f"ROWS(tblItems[ID])>{DATA_ROWS}" not in str(ws["D1"].value or ""):
        check.fails.append(f"{check.path.name}/Items!D1: capacity warning formula is missing")
    _column_alignment_contract(
        check,
        (
            ("ID", "left", "center", "compact text"),
            ("Start", "right", "center", "date"),
            ("Delivery Health", "left", "center", "delivery state"),
            ("Latest Status", "left", "top", "wrapped narrative"),
            ("IsBlocked", "center", "center", "calculated Boolean"),
            ("IsPoint", "center", "center", "calculated Boolean"),
        ),
    )


def _raid_component_checks(check: SheetCheck) -> None:
    """Validate representative RAID column alignment."""
    _column_alignment_contract(
        check,
        (
            ("RaidID", "left", "center", "compact text"),
            ("Detail", "left", "top", "wrapped narrative"),
            ("Prob", "right", "center", "number"),
            ("NextReview", "right", "center", "date"),
        ),
    )


def _overview_component_checks(check: SheetCheck) -> None:
    """Validate Overview panel geometry, fills and alignment."""
    expect_height(check, 1, ROWS["panel_title"], "panel-title")
    expect_height(check, 2, ROWS["panel_header"], "panel-header")
    expect_height(check, 3, ROWS["panel_body"], "panel-body")
    expect_fill(check, "A1", COLORS["header_bg"], "panel-title")
    expect_fill(check, "A2", COLORS["brand_tint"], "panel-header")
    for coordinate, horizontal, vertical, role in (
        ("A1", "left", "center", "panel title"),
        ("A2", "left", "center", "panel header"),
        ("A3", "left", "top", "panel text"),
    ):
        expect_alignment(check, coordinate, horizontal, vertical, role)
    for column in ("D", "J", "N", "T"):
        expect_column_alignment(check, column, "right", "top", "panel date")


def _plan_component_checks(check: SheetCheck) -> None:
    """Validate Plan rail, axes, controls and depth contract."""
    ws = check.sheet
    if '="As of "&TEXT(TODAY()' not in str(ws["D1"].value or ""):
        check.fails.append(
            f"{check.path.name}/Plan!D1: reporting-date rail is not a factual As of date"
        )
    expect_height(check, 4, ROWS["axis_month"], "month-axis")
    expect_height(check, 5, ROWS["axis_week"], "week-axis")
    expect_fill(check, "F4", COLORS["brand_tint"], "month-axis anchor")
    if style_color(ws["F4"].font.color) != normalise_hex(COLORS["header_bg"]):
        check.fails.append(f"{check.path.name}/Plan!F4: month-axis text is not header ink")
    expect_fill(check, "F5", COLORS["header_bg"], "week-axis anchor")
    if style_color(ws["F5"].font.color) != normalise_hex(COLORS["header_fg"]):
        check.fails.append(f"{check.path.name}/Plan!F5: week-axis text is not white")
    expect_fill(check, "F1", COLORS["bar_done_bg"], "top-rail legend")
    expect_fill(check, "A5", COLORS["header_bg"], "view-table header")
    for coordinate in ("B2", "C2", "C3", "E2"):
        expect_fill(check, coordinate, COLORS["input_bg"], "editable filter")
    for coordinate in ("B2", "B3", "C2", "C3", "E2", "E3"):
        expect_stop_validation(check, coordinate, "Plan input")
    for coordinate, horizontal, vertical, role in (
        ("A2", "left", "center", "filter label"),
        ("B2", "left", "center", "text filter"),
        ("B3", "left", "center", "numeric filter"),
        ("C2", "left", "center", "extra scope filter"),
        ("C3", "left", "center", "extra scope filter"),
        ("D2", "left", "center", "date filter label"),
        ("E2", "right", "center", "date filter"),
        ("F5", "center", "center", "timeline axis"),
    ):
        expect_alignment(check, coordinate, horizontal, vertical, role)
    if ws["B3"].value != MAX_HIERARCHY_LEVEL:
        check.fails.append(
            f"{check.path.name}/Plan!B3: default hierarchy depth is {ws['B3'].value!r}; "
            f"expected {MAX_HIERARCHY_LEVEL}"
        )
    depth_rules = validations_covering(ws, "B3")
    if not any(str(MAX_HIERARCHY_LEVEL) in str(rule.formula1) for rule in depth_rules):
        check.fails.append(
            f"{check.path.name}/Plan!B3: depth validation excludes level {MAX_HIERARCHY_LEVEL}"
        )


def _config_table_checks(check: SheetCheck) -> None:
    """Validate Config's parallel table-band layout."""
    expected_refs = {
        "tblStatuses": "E3:H9",
        "tblTypes": "J3:K17",
        "tblPriorities": "M3:M8",
        "tblTeams": "O3:O4",
        "tblRaidTypes": "Q3:S8",
        "tblRaidStatuses": "U3:V6",
        "tblSeverity": "X3:Y7",
        "tblDeliveryHealth": "AA3:AA7",
        "tblPeople": "AC3:AE4",
    }
    for table_name, expected_ref in expected_refs.items():
        table = check.sheet.tables.get(table_name)
        if table is None:
            check.fails.append(f"{check.path.name}/Config: missing {table_name}")
        elif table.ref != expected_ref:
            check.fails.append(
                f"{check.path.name}/Config: {table_name} is {table.ref}; "
                f"expected parallel band {expected_ref}"
            )


def _config_gutter_checks(check: SheetCheck) -> None:
    """Validate the narrow empty gutters between Config bands."""
    for gutter in ("D", "I", "L", "N", "P", "T", "W", "Z", "AB", "AF"):
        width = check.sheet.column_dimensions[gutter].width
        if width is None or not math.isclose(float(width), math.e, abs_tol=0.2):
            check.fails.append(
                f"{check.path.name}/Config: gutter {gutter} width is {width!r}; "
                "expected one thin 2-unit column"
            )
        if check.sheet[f"{gutter}3"].value not in {None, ""}:
            check.fails.append(f"{check.path.name}/Config!{gutter}3: gutter is not empty")


def _config_component_checks(check: SheetCheck) -> None:
    """Validate Config density, layout, alignment and hierarchy guards."""
    ws = check.sheet
    default_height = ws.sheet_format.defaultRowHeight
    if default_height is None or not math.isclose(
        float(default_height), float(ROWS["data_compact"]), abs_tol=0.1
    ):
        check.fails.append(
            f"{check.path.name}/Config: default data row height is {default_height!r}; "
            f"expected {ROWS['data_compact']} pt"
        )
    _config_table_checks(check)
    _config_gutter_checks(check)
    for coordinate, horizontal, vertical, role in (
        ("A3", "left", "center", "Config header"),
        ("A4", "left", "center", "Config label"),
        ("B4", "right", "center", "numeric setting"),
        ("F4", "center", "center", "native role checkbox"),
        ("J4", "left", "center", "taxonomy text"),
        ("K4", "right", "center", "taxonomy level"),
    ):
        expect_alignment(check, coordinate, horizontal, vertical, role)
    level_rules = validations_covering(ws, "K4")
    if not any(str(rule.formula2) == str(MAX_HIERARCHY_LEVEL) for rule in level_rules):
        check.fails.append(
            f"{check.path.name}/Config!K4: hierarchy maximum is not {MAX_HIERARCHY_LEVEL}"
        )
    health_gap_rules = list(matching_rules(ws, '$AA4=""', f"COUNTA($AA5:$AA${CONFIG_ROWS + 3})>0"))
    if not any(rule.stopIfTrue is True for rule in health_gap_rules):
        check.fails.append(
            f"{check.path.name}/Config: Delivery Health internal gaps are not guarded"
        )


def _operational_table_checks(check: SheetCheck, table_name: str) -> None:
    """Validate one editable table's recurring header and body geometry."""
    ws = check.sheet
    if table_name not in ws.tables:
        check.fails.append(f"{check.path.name}/{ws.title}: missing {table_name}")
        return
    min_col, min_row, _, _ = range_boundaries(ws.tables[table_name].ref)
    expect_height(check, min_row, ROWS["table_header"], "operational-table header")
    header = ws.cell(min_row, min_col)
    actual_fill = cell_rgb(header.fill.fgColor)
    actual_font = cell_rgb(header.font.color)
    wanted_fill = normalise_hex(COLORS["header_bg"])
    wanted_font = normalise_hex(COLORS["header_fg"])
    if header.fill.fill_type != "solid" or actual_fill != wanted_fill:
        check.fails.append(
            f"{check.path.name}/{ws.title}: {table_name} header fill is "
            f"{actual_fill or '<not explicit solid RGB>'}; expected {wanted_fill}"
        )
    if actual_font != wanted_font:
        check.fails.append(
            f"{check.path.name}/{ws.title}: {table_name} header text is "
            f"{actual_font or '<not explicit RGB>'}; expected {wanted_font}"
        )
    if header.font.name != TYPOGRAPHY["body_font"] or not header.font.bold:
        check.fails.append(
            f"{check.path.name}/{ws.title}: {table_name} header typography must be "
            f"bold {TYPOGRAPHY['body_font']}"
        )
    default_height = ws.sheet_format.defaultRowHeight
    if default_height is None or not math.isclose(
        float(default_height), float(ROWS["data_compact"]), abs_tol=0.1
    ):
        check.fails.append(
            f"{check.path.name}/{ws.title}: default data row height is "
            f"{default_height!r}; expected {ROWS['data_compact']} pt"
        )


def component_checks(path: Path, wb: Workbook, fails: list[str]) -> None:
    """Spot-check stable components without depending on calculated values."""
    _page_title_checks(path, wb, fails)
    surface_checks = {
        "Items": _items_component_checks,
        "RAID": _raid_component_checks,
        "Overview": _overview_component_checks,
        "Plan": _plan_component_checks,
        "Config": _config_component_checks,
    }
    for sheet_name, run_checks in surface_checks.items():
        if sheet_name in wb.sheetnames:
            run_checks(SheetCheck(path, wb[sheet_name], fails))
    for sheet_name, table_name in (("Items", "tblItems"), ("RAID", "tblRAID")):
        if sheet_name in wb.sheetnames:
            _operational_table_checks(SheetCheck(path, wb[sheet_name], fails), table_name)


def workbook_checks(path: Path, fails: list[str]) -> None:
    """Run package, workbook-state and component checks for one release file."""
    label = path.name
    if not path.exists():
        fails.append(f"{label}: file does not exist ({path})")
        return
    try:
        with zipfile.ZipFile(path) as package:
            bad_member = package.testzip()
            if bad_member:
                fails.append(f"{label}: corrupt ZIP member {bad_member}")
                return
            check = PackageCheck(path, package, fails)
            package_theme_checks(check)
            package_window_checks(check)
            package_calculation_checks(check)
            package_formula_checks(check)
            package_button_checks(check)
    except zipfile.BadZipFile as exc:
        fails.append(f"{label}: not a valid Excel package: {exc}")
        return

    try:
        wb = openpyxl.load_workbook(
            path,
            data_only=False,
            read_only=False,
            keep_vba=path.suffix.lower() == ".xlsm",
        )
    except (OSError, ValueError, KeyError, InvalidFileException, zipfile.BadZipFile) as exc:
        fails.append(f"{label}: openpyxl could not open workbook: {exc}")
        return
    try:
        visible_sheet_chrome_checks(path, wb, fails)
        component_checks(path, wb, fails)
        state_and_data_checks(path, wb, fails)
    finally:
        wb.close()


def default_paths() -> list[Path]:
    """Return the required release workbook paths.

    Returns:
        The macro-free and macro-enabled release paths.

    """
    return [DIST / "PM_Workbook.xlsx", DIST / "PM_Workbook.xlsm"]


def main(argv: list[str] | None = None) -> int:
    """Run design-system QA for explicit or default release paths.

    Returns:
        Zero when every check passes; otherwise one.

    """
    args = sys.argv[1:] if argv is None else argv
    paths = [Path(value).expanduser().resolve() for value in args] or default_paths()
    fails: list[str] = []
    token_checks(fails)

    for path in paths:
        workbook_checks(path, fails)

    if fails:
        sys.stdout.write("DESIGN QA FAIL\n")
        sys.stdout.write("".join(f" - {failure}\n" for failure in fails))
        return 1

    sys.stdout.write("DESIGN QA PASS\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
