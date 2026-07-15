"""Verify Overview ranking, capacity, scope depth and Plan coherence."""

import sys
from datetime import UTC, date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries

from ..pipeline import build_one
from ..spec import config
from ..spec.capacity import PLAN_ROWS
from .common import temporary_examples, temporary_workbook, workbook_error_cells
from .excel import recalculate

FLOAT_TOLERANCE = 1e-9
EXPECTED_SCOPE_SELECTORS = 4
OVERVIEW_INDENT = "\u203a "
HIERARCHY_TYPES = ("Project", "Release", "Phase", "Team", "Task", "Sub Task")


def _check(fails: list[str], label: str, got: object, want: object) -> None:
    matches = (
        abs(got - want) < FLOAT_TOLERANCE
        if isinstance(want, float) and isinstance(got, (int, float))
        else got == want
    )
    if not matches:
        fails.append(f"{label}: got {got!r}, want {want!r}")


def _items(today: date) -> list[dict]:
    return [
        {
            "ID": "I-1001",
            "Title": "Project Alpha",
            "Type": "Project",
            "Status": "In Progress",
            "Owner": "Alice",
            "Due": today + timedelta(days=60),
            "Delivery Health": "Off track",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1002",
            "Title": "Blocked predecessor",
            "Type": "Story",
            "Parent": "I-1001",
            "Status": "Backlog",
            "Owner": "Alice",
            "Due": today + timedelta(days=20),
            "Delivery Health": "Blocked",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1003",
            "Title": "Critical overdue work",
            "Type": "Story",
            "Parent": "I-1001",
            "Status": "In Progress",
            "Priority": "P0",
            "Owner": "Alice",
            "Due": today - timedelta(days=1),
            "BlockedBy": "I-1002",
            "Updated": today - timedelta(days=1),
            "InProgressSince": today - timedelta(days=10),
        },
        {
            "ID": "I-1004",
            "Title": "Second overdue item",
            "Type": "Story",
            "Parent": "I-1001",
            "Status": "In Progress",
            "Priority": "P1",
            "Owner": "Alice",
            "Due": today - timedelta(days=10),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1000",
            "Title": "Cancelled same-date release",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Cancelled",
            "Owner": "Alice",
            "Due": today + timedelta(days=10),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1005",
            "Title": "Steering approval",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Due": today + timedelta(days=10),
            "Delivery Health": "At risk",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1012",
            "Title": "Approval preparation",
            "Type": "Phase",
            "Parent": "I-1005",
            "Status": "In Progress",
            "Owner": "Alice",
            "Start": today + timedelta(days=2),
            # Deliberately later than the parent: neither Overview nor Plan may
            # adopt this descendant date for I-1005.
            "Due": today + timedelta(days=17),
            "Delivery Health": "Blocked",
            "Updated": today - timedelta(days=1),
            "InProgressSince": today - timedelta(days=1),
        },
        {
            "ID": "I-1007",
            "Title": "Pilot ready",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Owner": "Alice",
            "Due": today + timedelta(days=11),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1008",
            "Title": "Data migration",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Owner": "Alice",
            "Due": today + timedelta(days=12),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1009",
            "Title": "Training complete",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Owner": "Alice",
            "Due": today + timedelta(days=13),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1010",
            "Title": "Launch readiness",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Owner": "Alice",
            "Due": today + timedelta(days=14),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1011",
            "Title": "Go live",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Ready",
            "Owner": "Alice",
            "Due": today + timedelta(days=15),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-1006",
            "Title": "Delivered",
            "Type": "Task",
            "Parent": "I-1001",
            "Status": "Done",
            "Priority": "P2",
            "Owner": "Alice",
            "DoneDate": today - timedelta(days=3),
            "Updated": today - timedelta(days=3),
        },
        {
            "ID": "I-2001",
            "Title": "Project Beta",
            "Type": "Project",
            "Status": "In Progress",
            "Owner": "Bob",
            "Due": today + timedelta(days=90),
            "Delivery Health": "On track",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-2002",
            "Title": "Off-track Beta work",
            "Type": "Task",
            "Parent": "I-2001",
            "Status": "In Progress",
            "Priority": "P2",
            "Owner": "Bob",
            "Due": today + timedelta(days=20),
            "Delivery Health": "Off track",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-2003",
            "Title": "Undated release",
            "Type": "Release",
            "Parent": "I-2001",
            "Status": "Ready",
            "Owner": "Bob",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-2004",
            "Title": "Start-only Beta work",
            "Type": "Task",
            "Parent": "I-2001",
            "Status": "Ready",
            "Owner": "Bob",
            "Start": today + timedelta(days=5),
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-3001",
            "Title": "Product Gamma",
            "Type": "Product",
            "Status": "In Progress",
            "Owner": "Cara",
            "Start": today,
            "Due": today + timedelta(days=120),
            "Delivery Health": "On track",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-3002",
            "Title": "At-risk product work",
            "Type": "Task",
            "Parent": "I-3001",
            "Status": "In Progress",
            "Priority": "P3",
            "Owner": "Cara",
            "Start": today + timedelta(days=1),
            "Due": today + timedelta(days=30),
            "Delivery Health": "At risk",
            "Updated": today - timedelta(days=1),
        },
        {
            "ID": "I-9998",
            "Title": "Deleted historical release",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Deleted",
            "Priority": "P0",
            "Owner": "Alice",
            "Due": today + timedelta(days=1),
            "Delivery Health": "Blocked",
            "Updated": today,
        },
    ]


def _raid(today: date) -> list[dict]:
    return [
        {
            "RaidID": "R-001",
            "Type": "Decision",
            "Title": "Approve steering route",
            "RelatedID": "I-1001",
            "Owner": "Alice",
            "Status": "Open",
            "NextReview": today + timedelta(days=2),
            "Raised": today - timedelta(days=2),
            "Updated": today - timedelta(days=1),
        },
        {
            "RaidID": "R-002",
            "Type": "Risk",
            "Title": "Supplier dependency",
            "RelatedID": "I-1003",
            "Owner": "Alice",
            "Status": "Monitoring",
            "Prob": 3,
            "Impact": 4,
            "Response": "Escalate with supplier",
            "NextReview": today + timedelta(days=3),
            "Raised": today - timedelta(days=5),
            "Updated": today - timedelta(days=1),
        },
        {
            "RaidID": "R-003",
            "Type": "Risk",
            "Title": "Closed Beta risk",
            "RelatedID": "I-2001",
            "Owner": "Bob",
            "Status": "Closed",
            "Prob": 5,
            "Impact": 5,
            "Raised": today - timedelta(days=10),
            "Closed": today - timedelta(days=2),
            "Updated": today - timedelta(days=2),
        },
        {
            "RaidID": "R-004",
            "Type": "Risk",
            "Title": "Low scoring open risk",
            "RelatedID": "I-2001",
            "Owner": "Bob",
            "Status": "Open",
            "Prob": 2,
            "Impact": 2,
            "Response": "Monitor locally",
            "NextReview": today + timedelta(days=4),
            "Raised": today - timedelta(days=3),
            "Updated": today - timedelta(days=1),
        },
        {
            "RaidID": "R-005",
            "Type": "Risk",
            "Title": "Deleted historical risk",
            "RelatedID": "I-1001",
            "Owner": "Alice",
            "Status": "Deleted",
            "Prob": 5,
            "Impact": 5,
            "Raised": today - timedelta(days=5),
            "Closed": today,
            "Updated": today,
        },
    ]


def _direct_child_chain(today: date, child_level: int) -> list[dict]:
    """Return one hierarchy with an on-track parent and directly blocked child.

    Returns:
        A hierarchy fixture through the requested child level.

    """
    rows = []
    for level in range(1, child_level + 1):
        is_target = level == child_level - 1
        is_child = level == child_level
        rows.append({
            "ID": f"I-4{level:03d}",
            "Title": f"Level {level} item",
            "Type": HIERARCHY_TYPES[level - 1],
            "Parent": "" if level == 1 else f"I-4{level - 1:03d}",
            "Status": "In Progress" if is_target or is_child else "Done",
            "Owner": "Alice",
            "Start": today,
            "Due": today + timedelta(days=level),
            "Delivery Health": "Blocked" if is_child else "On track",
            "Updated": today,
        })
    return rows


def _check_deleted_table_row(
    workbook: openpyxl.Workbook,
    record: tuple[str, str, str, str],
    fails: list[str],
) -> None:
    table_name, sheet_name, key, record_id = record
    worksheet = workbook[sheet_name]
    min_col, header_row, max_col, max_row = range_boundaries(worksheet.tables[table_name].ref)
    headers = {
        worksheet.cell(header_row, column).value: column for column in range(min_col, max_col + 1)
    }
    matching_rows = [
        row
        for row in range(header_row + 1, max_row + 1)
        if worksheet.cell(row, headers[key]).value == record_id
    ]
    if not matching_rows:
        fails.append(f"Deleted history row {record_id} is not visible in {sheet_name}")
    elif worksheet.cell(matching_rows[0], headers["Status"]).value != "Deleted":
        fails.append(f"Deleted history row {record_id} lost its configured Status")


def _check_deleted_history(out: Path, fails: list[str]) -> None:
    workbook = openpyxl.load_workbook(out, data_only=False)
    try:
        _check_deleted_table_row(
            workbook,
            ("tblItems", "Items", "ID", "I-9998"),
            fails,
        )
        _check_deleted_table_row(
            workbook,
            ("tblRAID", "RAID", "RaidID", "R-005"),
            fails,
        )
        hardcoded = [
            f"{worksheet.title}!{cell.coordinate}"
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "f" and "deleted" in str(cell.value).casefold()
        ]
        if hardcoded:
            fails.append(f"Formulas hardcode the Deleted label: {hardcoded[:10]!r}")
    finally:
        workbook.close()


def _chain_label(level: int) -> str:
    """Return the rendered Overview label for one direct-child fixture row.

    Returns:
        The indented item label emitted by the Overview formula.

    """
    return (
        f"{OVERVIEW_INDENT * (level - 1)}I-4{level:03d} · "
        f"{HIERARCHY_TYPES[level - 1]} · Level {level} item"
    )


def _check_primary_workbook(out: Path, today: date, fails: list[str]) -> None:
    workbook = openpyxl.load_workbook(out, data_only=True)
    try:
        overview = workbook["Overview"]
        summary = {
            str(overview.cell(row, 1).value or ""): overview.cell(row, 2).value
            for row in range(3, 8)
        }
        expected_health = {
            f"{OVERVIEW_INDENT * 2}I-1012 · Phase · Approval preparation": "Blocked",
            f"{OVERVIEW_INDENT * 4}I-1002 · Story · Blocked predecessor": "Blocked",
            "I-1001 · Project · Project Alpha": "Off track",
            "I-2001 · Project · Project Beta": "On track",
            "I-3001 · Product · Product Gamma": "On track",
        }
        for label, health in expected_health.items():
            _check(fails, f"direct health for {label}", summary.get(label), health)
        _check(fails, "RAID type first", overview["F3"].value, "Risk")
        _check(fails, "RAID description", overview["G3"].value, "Supplier dependency")
        _check(fails, "RAID severity", overview["H3"].value, "High")
        _check(fails, "RAID latest status", overview["K3"].value, "Escalate with supplier")
        _check(
            fails, "Coming-up decision", overview["M3"].value, "Decision · Approve steering route"
        )
        _check(
            fails,
            "Coming-up date",
            overview["N3"].value,
            (today + timedelta(days=2)).strftime("%-d %b %Y"),
        )
        _check(fails, "Coming-up release", overview["M4"].value, "Release · Steering approval")
        _check(fails, "Coming-up fifth row", overview["M7"].value, "Release · Training complete")
        _check(fails, "Coming-up disclosure", overview["O1"].value, "Showing 5 of 7")
        _check(fails, "Recent completion", overview["Q3"].value, "Delivered")

        plan = workbook["Plan"]
        release_rows = [
            row for row in range(6, PLAN_ROWS + 6) if plan.cell(row=row, column=1).value == "I-1005"
        ]
        if not release_rows:
            fails.append("Plan does not show the due-only Release with a scheduled child")
        else:
            glyphs = [plan.cell(release_rows[0], column).value for column in range(6, 58)]
            if glyphs.count("◆") != 1:
                fails.append(
                    "Due-only item must show exactly one Plan key-date diamond from its own "
                    f"Due date: {glyphs!r}"
                )
            interval_glyphs = {"✓", "●", "!", "\u00d7", "—"}
            if any(glyph in interval_glyphs for glyph in glyphs):
                fails.append(
                    f"Due-only item gained a Plan interval glyph from descendant dates: {glyphs!r}"
                )

        visible = " ".join(
            str(overview.cell(row, column).value or "")
            for row in range(1, 9)
            for column in range(1, 22)
        )
        records_expected_off_panel = (
            "Closed Beta risk",
            "Low scoring open risk",
            "Undated release",
            "Cancelled same-date release",
            "Go live",
            "Deleted historical release",
            "Deleted historical risk",
        )
        fails.extend(
            f"panel filter included {record}"
            for record in records_expected_off_panel
            if record in visible
        )
        fails.extend(f"error {value}" for value in workbook_error_cells(workbook))
    finally:
        workbook.close()
    _check_deleted_history(out, fails)


def _config_setting_row(out: Path) -> int | None:
    formula_workbook = openpyxl.load_workbook(out, data_only=False)
    try:
        config = formula_workbook["Config"]
        scope_setting_row = next(
            (
                row
                for row in range(1, config.max_row + 1)
                if config.cell(row, 1).value == "ExecutiveStatusMaxLevel"
            ),
            None,
        )
    finally:
        formula_workbook.close()

    return scope_setting_row


def _check_level_two(
    out: Path,
    scope_setting_row: int | None,
    today: date,
    fails: list[str],
) -> None:
    if scope_setting_row is None:
        fails.append("Config ExecutiveStatusMaxLevel setting missing")
    else:
        recalculate(out, sheet="Config", cell=f"B{scope_setting_row}", value="2")
        level_workbook = openpyxl.load_workbook(out, data_only=True)
        try:
            level_overview = level_workbook["Overview"]
            labels = [str(level_overview.cell(row, 1).value or "") for row in range(3, 8)]
            release_rows = [
                row
                for row in range(3, 8)
                if "I-1005" in str(level_overview.cell(row, 1).value or "")
            ]
            blocked_child_rows = [
                row
                for row in range(3, 8)
                if "I-1012" in str(level_overview.cell(row, 1).value or "")
            ]
            if not release_rows:
                fails.append(
                    "Executive Status Summary at maximum level 2 does not show "
                    f"the Level-2 Release: {labels!r}"
                )
            else:
                release_row = release_rows[0]
                if level_overview.cell(release_row, 2).value != "At risk":
                    fails.append(
                        "Level-2 Release inherited a descendant's health: "
                        f"got {level_overview.cell(release_row, 2).value!r}, want 'At risk'"
                    )
                if level_overview.cell(release_row, 3).value != "Owner not set":
                    fails.append(
                        "Level-2 Release inherited its scope owner's name: "
                        f"got {level_overview.cell(release_row, 3).value!r}, "
                        "want 'Owner not set'"
                    )
                direct_due = (today + timedelta(days=10)).strftime("%-d %b %Y")
                if level_overview.cell(release_row, 4).value != direct_due:
                    fails.append(
                        "Level-2 Release inherited a descendant's Due date: "
                        f"got {level_overview.cell(release_row, 4).value!r}, "
                        f"want {direct_due!r}"
                    )
            if not blocked_child_rows:
                fails.append(
                    "Directly blocked Level-3 item did not bypass the level-2 display limit: "
                    f"{labels!r}"
                )
            elif level_overview.cell(blocked_child_rows[0], 2).value != "Blocked":
                fails.append(
                    "Directly blocked Level-3 item lost its own health: "
                    f"{level_overview.cell(blocked_child_rows[0], 2).value!r}"
                )
            if "Showing 5 of" not in str(level_overview["D1"].value or ""):
                fails.append(f"Scope-panel disclosure missing: {level_overview['D1'].value!r}")
            selectors = [level_workbook["Calc"].cell(row, 13).value for row in range(2, 20)]
            selectors = [value for value in selectors if value not in {None, ""}]
            if (
                len(selectors) != EXPECTED_SCOPE_SELECTORS
                or selectors[0] != "All"
                or any("Release" in str(value) for value in selectors)
            ):
                fails.append(f"Plan Scope selectors differ: {selectors!r}")
        finally:
            level_workbook.close()


def _check_plan_product_scope(out: Path, fails: list[str]) -> None:
    recalculate(out, sheet="Plan", cell="B2", value="I-3001 · Product Gamma")
    plan_workbook = openpyxl.load_workbook(out, data_only=True)
    try:
        plan = plan_workbook["Plan"]
        plan_ids = [plan.cell(row=row, column=1).value for row in range(6, PLAN_ROWS + 6)]
        plan_ids = [value for value in plan_ids if value not in {None, "", "— none —"}]
        fails.extend(
            f"Plan Product scope missing {expected}: {plan_ids!r}"
            for expected in ("I-3001", "I-3002")
            if expected not in plan_ids
        )
        unrelated = {"I-1001", "I-1002", "I-2001", "I-2002"} & set(plan_ids)
        if unrelated:
            fails.append(f"Plan Product scope includes unrelated IDs: {sorted(unrelated)}")
    finally:
        plan_workbook.close()


def _check_plan_missing_dates(out: Path, today: date, fails: list[str]) -> None:
    recalculate(out, sheet="Plan", cell="B2", value="I-2001 · Project Beta")
    beta_workbook = openpyxl.load_workbook(out, data_only=True)
    try:
        plan = beta_workbook["Plan"]
        rows = {
            str(plan.cell(row=row, column=1).value): row
            for row in range(6, PLAN_ROWS + 6)
            if plan.cell(row=row, column=1).value not in {None, "", "— none —"}
        }
        fails.extend(
            f"Plan hides incomplete-schedule item {item_id}: {sorted(rows)!r}"
            for item_id in ("I-2003", "I-2004")
            if item_id not in rows
        )

        undated_row = rows.get("I-2003")
        if undated_row is not None:
            _check(fails, "undated Plan Start", plan.cell(undated_row, 4).value, None)
            _check(fails, "undated Plan Due", plan.cell(undated_row, 5).value, None)
            glyphs = [plan.cell(undated_row, column).value for column in range(6, 58)]
            if any(glyph not in {None, ""} for glyph in glyphs):
                fails.append(f"undated Plan item gained a timeline glyph: {glyphs!r}")

        start_only_row = rows.get("I-2004")
        if start_only_row is not None:
            _check(
                fails,
                "start-only Plan Start",
                plan.cell(start_only_row, 4).value,
                datetime.combine(today + timedelta(days=5), datetime.min.time()),
            )
            _check(fails, "start-only Plan Due", plan.cell(start_only_row, 5).value, None)
            glyphs = [plan.cell(start_only_row, column).value for column in range(6, 58)]
            if any(glyph not in {None, ""} for glyph in glyphs):
                fails.append(f"start-only Plan item gained a timeline glyph: {glyphs!r}")

        counter = str(plan["F2"].value or "")
        if "not shown" in counter.lower() or not any(
            token in counter.lower() for token in ("schedule", "date")
        ):
            fails.append(f"Plan incomplete-schedule disclosure differs: {counter!r}")
    finally:
        beta_workbook.close()


def _run_overview_scenario(today: date, fails: list[str]) -> None:
    with temporary_examples() as examples, temporary_workbook("PM_overview_scenario") as out:
        examples.ITEMS_EXAMPLES = _items(today)
        examples.PEOPLE_EXAMPLES = [
            {"Person": "Alice", "Role": "PM", "Team": "Core"},
            {"Person": "Bob", "Role": "Lead", "Team": "Core"},
            {"Person": "Cara", "Role": "Product lead", "Team": "Core"},
        ]
        examples.RAID_EXAMPLES = _raid(today)
        build_one(out, with_vba=False)
        recalculate(out)

        _check_primary_workbook(out, today, fails)
        scope_setting_row = _config_setting_row(out)
        _check_level_two(out, scope_setting_row, today, fails)
        _check_plan_product_scope(out, fails)
        _check_plan_missing_dates(out, today, fails)


def _check_item_local_health(today: date, fails: list[str]) -> None:
    # At every supported child level, the blocked child appears as itself even
    # below the depth limit while the displayed parent keeps its own health.
    for child_level in range(2, 7):
        with (
            temporary_examples() as examples,
            temporary_workbook(f"PM_direct_child_level_{child_level}") as out,
        ):
            examples.ITEMS_EXAMPLES = _direct_child_chain(today, child_level)
            examples.PEOPLE_EXAMPLES = [
                {"Person": "Alice", "Role": "PM", "Team": "Core"},
            ]
            examples.RAID_EXAMPLES = []
            build_one(out, with_vba=False)
            executive_setting_row = 4 + next(
                index
                for index, setting in enumerate(config.SETTINGS)
                if setting[0] == "cfgExecutiveStatusMaxLevel"
            )
            recalculate(
                out,
                sheet="Config",
                cell=f"B{executive_setting_row}",
                value=child_level - 1,
            )
            child_workbook = openpyxl.load_workbook(out, data_only=True)
            try:
                overview = child_workbook["Overview"]
                summary = {
                    str(overview.cell(row, 1).value or ""): overview.cell(row, 2).value
                    for row in range(3, 8)
                }
                child_label = _chain_label(child_level)
                parent_level = child_level - 1
                parent_label = _chain_label(parent_level)
                if summary.get(child_label) != "Blocked":
                    fails.append(
                        f"directly blocked level {child_level} item missing or changed: "
                        f"got {summary.get(child_label)!r}, want 'Blocked'; rows={summary!r}"
                    )
                if summary.get(parent_label) != "On track":
                    fails.append(
                        f"level {parent_level} parent inherited child health: "
                        f"got {summary.get(parent_label)!r}, want 'On track'; rows={summary!r}"
                    )
                fails.extend(
                    f"direct-child level {child_level}: {value}"
                    for value in workbook_error_cells(child_workbook)
                )
            finally:
                child_workbook.close()


def main() -> None:
    """Run all Overview and Plan behavioral scenarios."""
    today = datetime.now(tz=UTC).astimezone().date()
    fails: list[str] = []
    _run_overview_scenario(today, fails)
    _check_item_local_health(today, fails)
    if fails:
        sys.stdout.write("OVERVIEW QA FAIL\n")
        sys.stdout.write("".join(f" - {failure}\n" for failure in fails[:40]))
        sys.exit(1)
    sys.stdout.write("OVERVIEW QA PASS\n")


if __name__ == "__main__":
    main()
