"""Regression contracts for the provider-neutral agent change-set format."""

from __future__ import annotations

import json
import math
import shutil
import tempfile
import unittest
from copy import deepcopy
from datetime import UTC, date, datetime
from pathlib import Path

import openpyxl
from jsonschema import Draft202012Validator
from openpyxl.utils.cell import range_boundaries

from .. import pipeline
from ..data.bridge import describe_workbook
from ..data.contract import (
    CHANGESET_SCHEMA,
    CONTRACT_NAME,
    CONTRACT_VERSION,
    ITEM_WRITABLE_FIELDS,
    ContractError,
    canonical_json,
    parse_changeset,
)
from ..data.inject import injected_source, validate_snapshot
from ..data.schema import schema_fingerprint
from ..spec.capacity import CONFIG_ROWS, DATA_ROWS
from .test_data_layer import _snapshot


def _change_set() -> dict[str, object]:
    """Return one complete valid contract document.

    Returns:
        A mutable change-set fixture.

    """
    return {
        "contract": "excel-pm-agent-change-set",
        "version": "1.0.0",
        "request_id": "3e5f7e30-7df5-4f16-8ce8-8725a02f3d51",
        "created_at": "2026-07-15T10:00:00Z",
        "target": {
            "workbook_sha256": "1" * 64,
            "workbook_schema_fingerprint": "2" * 64,
            "build_schema_fingerprint": "3" * 64,
        },
        "producer": {"name": "contract tests", "version": "1.0", "extensions": {}},
        "source": {
            "description": "pasted weekly status",
            "uri": "clipboard:weekly-status",
            "retrieved_at": "2026-07-15T09:55:00Z",
            "extensions": {},
        },
        "operations": [
            {
                "operation_id": "create-project",
                "op": "upsert",
                "entity": "item",
                "identity": {
                    "source": {
                        "namespace": "paste:weekly-status",
                        "record_id": "project-001",
                    }
                },
                "client_ref": "project",
                "set": {
                    "Type": "Project",
                    "Title": "Delivery programme",
                    "Status": "In Progress",
                    "Delivery Health": "On track",
                },
                "clear": [],
                "extensions": {},
            }
        ],
        "extensions": {},
    }


def _payload(document: dict[str, object]) -> bytes:
    """Encode one fixture as strict JSON bytes.

    Returns:
        UTF-8 JSON bytes.

    """
    return json.dumps(document, ensure_ascii=False, allow_nan=False).encode()


class SchemaContractTests(unittest.TestCase):
    """Pin the generated public JSON Schema and its supported shapes."""

    def test_schema_is_valid_draft_2020_12(self) -> None:
        """Publish a self-consistent Draft 2020-12 schema."""
        Draft202012Validator.check_schema(CHANGESET_SCHEMA)
        self.assertEqual(
            CHANGESET_SCHEMA["$schema"],
            "https://json-schema.org/draft/2020-12/schema",
        )
        self.assertEqual(CONTRACT_NAME, "excel-pm-agent-change-set")
        self.assertEqual(CONTRACT_VERSION, "1.0.0")

    def test_valid_item_and_raid_operations_parse(self) -> None:
        """Accept every public operation and structured reference shape."""
        document = _change_set()
        operations = document["operations"]
        self.assertIsInstance(operations, list)
        raid = {
            "operation_id": "create-risk",
            "op": "upsert",
            "entity": "raid",
            "identity": {"source": {"namespace": "paste:weekly-status", "record_id": "risk-001"}},
            "set": {
                "Type": "Risk",
                "Title": "Supplier delay",
                "Status": "Open",
                "RelatedID": {"client_ref": "project"},
                "Prob": 3,
                "Impact": 4,
            },
            "clear": ["NextReview"],
            "extensions": {},
        }
        operations.append(raid)
        operations.extend((
            {
                "operation_id": "delete-item",
                "op": "mark_deleted",
                "entity": "item",
                "identity": {"workbook_id": "I-1001"},
                "extensions": {},
            },
            {
                "operation_id": "delete-risk",
                "op": "mark_deleted",
                "entity": "raid",
                "identity": {
                    "source": {
                        "namespace": "paste:weekly-status",
                        "record_id": "risk-001",
                    }
                },
            },
        ))
        parsed = parse_changeset(_payload(document))
        self.assertEqual(parsed["operations"], operations)

    def test_external_identifiers_must_be_strings(self) -> None:
        """Reject numeric workbook, source, client and relationship identifiers."""
        mutations = (
            lambda operation: operation.__setitem__("identity", {"workbook_id": 3001}),
            lambda operation: operation["identity"]["source"].__setitem__("record_id", 42),
            lambda operation: operation.__setitem__("client_ref", 42),
            lambda operation: operation["set"].__setitem__("Parent", {"workbook_id": 3001}),
        )
        for mutate in mutations:
            with self.subTest(mutation=mutate):
                document = _change_set()
                operation = document["operations"][0]
                mutate(operation)
                with self.assertRaises(ContractError) as caught:
                    parse_changeset(_payload(document))
                self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")

    def test_unsupported_domain_operation_is_rejected(self) -> None:
        """Keep the public mutation boundary limited to upsert and mark_deleted."""
        document = _change_set()
        document["operations"][0]["op"] = "delete"
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")

    def test_closed_objects_reject_unknown_fields(self) -> None:
        """Reject mutation-like fields outside the documented core contract."""
        document = _change_set()
        document["surprise"] = True
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")
        self.assertEqual(caught.exception.diagnostics[0].pointer, "")

    def test_extensions_cannot_smuggle_writes_into_core_objects(self) -> None:
        """Treat extensions as inert metadata while still allowing nested JSON."""
        document = _change_set()
        document["extensions"] = {"vendor": {"set": {"Status": "Done"}}}
        parsed = parse_changeset(_payload(document))
        self.assertEqual(parsed["extensions"], document["extensions"])

    def test_null_is_not_a_clear_instruction(self) -> None:
        """Require explicit clear field names rather than JSON null."""
        document = _change_set()
        operations = document["operations"]
        self.assertIsInstance(operations, list)
        operations[0]["set"]["Due"] = None
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")

    def test_set_and_clear_cannot_name_the_same_field(self) -> None:
        """Reject contradictory preservation semantics within one upsert."""
        document = _change_set()
        operations = document["operations"]
        self.assertIsInstance(operations, list)
        operations[0]["clear"] = ["Title"]
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        diagnostic = caught.exception.diagnostics[0]
        self.assertEqual(
            set(diagnostic.as_dict()),
            {"code", "severity", "phase", "pointer", "operation_id", "message", "hint"},
        )
        self.assertEqual(diagnostic.code, "contract.field_conflict")
        self.assertEqual(diagnostic.severity, "error")
        self.assertEqual(diagnostic.phase, "schema")
        self.assertEqual(diagnostic.pointer, "/operations/0")
        self.assertEqual(diagnostic.operation_id, "create-project")
        self.assertTrue(diagnostic.message)
        self.assertTrue(diagnostic.hint)

    def test_formats_are_checked_explicitly(self) -> None:
        """Reject malformed UUID, timestamp and full-date values."""
        for pointer, mutate in (
            ("/request_id", lambda document: document.__setitem__("request_id", "not-a-uuid")),
            (
                "/created_at",
                lambda document: document.__setitem__("created_at", "2026-07-15 10:00"),
            ),
            (
                "/operations/0/set/Due",
                lambda document: document["operations"][0]["set"].__setitem__("Due", "15/07/2026"),
            ),
        ):
            with self.subTest(pointer=pointer):
                document = _change_set()
                mutate(document)
                with self.assertRaises(ContractError) as caught:
                    parse_changeset(_payload(document))
                self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")
                self.assertEqual(caught.exception.diagnostics[0].pointer, pointer)


class StrictJsonTests(unittest.TestCase):
    """Reject inputs that Python's permissive JSON decoder would accept."""

    def test_malformed_utf8_is_rejected(self) -> None:
        """Require one valid UTF-8 JSON document."""
        with self.assertRaises(ContractError) as caught:
            parse_changeset(b'{"title":"\xff"}')
        self.assertEqual(caught.exception.diagnostics[0].code, "json.invalid_utf8")

    def test_duplicate_keys_are_rejected(self) -> None:
        """Never silently select one value from duplicate JSON object keys."""
        with self.assertRaises(ContractError) as caught:
            parse_changeset(b'{"contract":"a","contract":"b"}')
        self.assertEqual(caught.exception.diagnostics[0].code, "json.duplicate_key")

    def test_non_finite_numbers_are_rejected(self) -> None:
        """Reject every non-I-JSON numeric representation."""
        for value in (b"NaN", b"Infinity", b"-Infinity", b"1e400"):
            with self.subTest(value=value):
                payload = b'{"extensions":{"number":' + value + b"}}"
                with self.assertRaises(ContractError) as caught:
                    parse_changeset(payload)
                self.assertEqual(caught.exception.diagnostics[0].code, "json.non_finite")

    def test_unpaired_surrogates_are_rejected(self) -> None:
        """Reject Unicode values that cannot be represented as valid XML text."""
        payload = _payload(_change_set()).replace(b"Delivery programme", b"\\ud800")
        with self.assertRaises(ContractError) as caught:
            parse_changeset(payload)
        self.assertEqual(caught.exception.diagnostics[0].code, "json.unpaired_surrogate")

    def test_illegal_xml_characters_are_rejected(self) -> None:
        """Reject control characters Excel's XML package cannot store."""
        document = _change_set()
        operations = document["operations"]
        self.assertIsInstance(operations, list)
        operations[0]["set"]["Title"] = "bad\u000bvalue"
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "json.illegal_xml")

    def test_excel_cell_limit_is_enforced(self) -> None:
        """Reject text that cannot fit in one Excel cell."""
        document = _change_set()
        operations = document["operations"]
        self.assertIsInstance(operations, list)
        operations[0]["set"]["Title"] = "x" * 32_768
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "contract.schema")

    def test_integers_outside_binary64_safe_range_are_rejected(self) -> None:
        """Keep extension metadata interoperable with I-JSON consumers."""
        document = _change_set()
        document["extensions"] = {"unsafe": 9_007_199_254_740_992}
        with self.assertRaises(ContractError) as caught:
            parse_changeset(_payload(document))
        self.assertEqual(caught.exception.diagnostics[0].code, "json.number_range")


class CanonicalJsonTests(unittest.TestCase):
    """Pin stable token input serialization independently of JSON formatting."""

    def test_object_key_order_and_whitespace_do_not_change_bytes(self) -> None:
        """Sort object keys and use compact UTF-8 JSON."""
        first = {"z": 1, "a": {"é": "literal", "date": date(2026, 7, 15)}}
        second = {"a": {"date": date(2026, 7, 15), "é": "literal"}, "z": 1}
        self.assertEqual(canonical_json(first), canonical_json(second))
        self.assertIn("é".encode(), canonical_json(first))
        self.assertNotIn(b" ", canonical_json(first))

    def test_canonical_json_rejects_non_finite_values(self) -> None:
        """Never hash a platform-specific non-finite JSON spelling."""
        for value in (math.nan, math.inf, -math.inf):
            with self.subTest(value=value), self.assertRaises(ValueError):
                canonical_json({"value": value})

    def test_parsing_does_not_mutate_the_input_shape(self) -> None:
        """Return values exactly as represented by the validated contract."""
        document = _change_set()
        before = deepcopy(document)
        self.assertEqual(parse_changeset(_payload(document)), before)


class DescribeContractTests(unittest.TestCase):
    """Pin the deterministic, provider-neutral workbook description."""

    directory: tempfile.TemporaryDirectory[str]
    workbook: Path
    now = datetime(2026, 7, 15, 11, 30, tzinfo=UTC)

    @classmethod
    def setUpClass(cls) -> None:
        """Build one current-structure workbook for read-only descriptions."""
        super().setUpClass()
        cls.directory = tempfile.TemporaryDirectory()
        cls.addClassCleanup(cls.directory.cleanup)
        cls.workbook = Path(cls.directory.name) / "describe.xlsx"
        snapshot = _snapshot()
        with injected_source(snapshot, validate_snapshot(snapshot)):
            pipeline.build_one(cls.workbook, with_vba=False)

    def test_describe_returns_schema_targets_capacities_and_choices(self) -> None:
        """Expose every value an agent needs to construct a bound change set."""
        result = describe_workbook(self.workbook, now=self.now)

        self.assertEqual(result["result"], "describe")
        self.assertEqual(result["contract"]["name"], CONTRACT_NAME)
        self.assertEqual(result["contract"]["version"], CONTRACT_VERSION)
        self.assertEqual(result["contract"]["schema"], CHANGESET_SCHEMA)
        self.assertRegex(result["target"]["workbook_sha256"], r"^[0-9a-f]{64}$")
        self.assertRegex(
            result["target"]["workbook_schema_fingerprint"],
            r"^[0-9a-f]{64}$",
        )
        self.assertEqual(result["target"]["build_schema_fingerprint"], schema_fingerprint())
        self.assertEqual(result["effective_date"], "2026-07-15")
        self.assertEqual(
            result["capacities"], {"items": DATA_ROWS, "raid": DATA_ROWS, "config": CONFIG_ROWS}
        )
        self.assertEqual(result["writable_fields"]["item"], list(ITEM_WRITABLE_FIELDS))
        self.assertIn(
            {
                "Status": "Deleted",
                "IsActive": False,
                "IsDone": True,
                "IsCancelled": True,
                "IsDeleted": True,
            },
            result["config"]["tables"]["tblStatuses"],
        )

    def test_describe_records_include_durable_identity_and_owned_columns(self) -> None:
        """Return ordered authored records with explicit blank source identity cells."""
        result = describe_workbook(self.workbook, now=self.now)
        items = result["records"]["items"]
        raid = result["records"]["raid"]
        self.assertTrue(items)
        self.assertTrue(raid)
        self.assertIn("ID", items[0])
        self.assertIn("Source", items[0])
        self.assertIn("Source ID", items[0])
        self.assertIn("RaidID", raid[0])
        self.assertIn("Source", raid[0])
        self.assertIn("Source ID", raid[0])

        item_columns = {entry["name"]: entry for entry in result["columns"]["item"]}
        self.assertEqual(item_columns["Title"]["ownership"], "input")
        self.assertTrue(item_columns["Title"]["writable"])
        self.assertEqual(item_columns["ID"]["ownership"], "vba")
        self.assertFalse(item_columns["ID"]["writable"])
        self.assertEqual(item_columns["Source"]["ownership"], "source_identity")
        self.assertFalse(item_columns["Source"]["writable"])
        self.assertEqual(item_columns["WbsKey"]["ownership"], "formula")

    def test_describe_is_byte_stable_for_the_same_effective_date(self) -> None:
        """Exclude volatile export timestamps and preserve deterministic ordering."""
        first = describe_workbook(self.workbook, now=self.now)
        second = describe_workbook(self.workbook, now=self.now)
        self.assertEqual(canonical_json(first), canonical_json(second))

    def test_observed_fingerprint_changes_without_changing_build_fingerprint(self) -> None:
        """Distinguish the live workbook shape from the current source schema."""
        legacy = Path(self.directory.name) / "legacy-shape.xlsx"
        shutil.copy2(self.workbook, legacy)
        workbook = openpyxl.load_workbook(legacy)
        worksheet = workbook["Items"]
        min_col, header_row, max_col, _max_row = range_boundaries(worksheet.tables["tblItems"].ref)
        header = next(
            worksheet.cell(header_row, column)
            for column in range(min_col, max_col + 1)
            if worksheet.cell(header_row, column).value == "Source"
        )
        header.value = "Legacy Source"
        workbook.save(legacy)

        current = describe_workbook(self.workbook, now=self.now)
        observed = describe_workbook(legacy, now=self.now)
        self.assertNotEqual(
            observed["target"]["workbook_schema_fingerprint"],
            current["target"]["workbook_schema_fingerprint"],
        )
        self.assertEqual(
            observed["target"]["build_schema_fingerprint"],
            current["target"]["build_schema_fingerprint"],
        )
        self.assertIn("Source", observed["schema_notes"]["added_columns"]["tblItems"])
        self.assertIn(
            "Legacy Source",
            observed["schema_notes"]["unknown_columns"]["tblItems"],
        )


if __name__ == "__main__":
    unittest.main()
