"""Prove that hierarchy rows never borrow operational values from one another.

The static pass scans every formula-bearing surface on all six sheets.  The
Excel scenario then uses deliberately conflicting values across a six-level
hierarchy so inherited dates, owners, health or status-derived presentation
cannot accidentally look correct.
"""

from __future__ import annotations

import argparse
import logging
import re
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from ..pipeline import SHEETS, build_one
from ..spec import config
from ..spec.capacity import PLAN_WEEKS
from ..spec.items import DIRECT_BLOCKED_HEALTH_FORMULA
from .common import temporary_examples, temporary_workbook, workbook_error_cells
from .excel import recalculate

if TYPE_CHECKING:
    from collections.abc import Iterator, Sequence

    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger(__name__)
OPERATIONAL_COLUMNS = (
    "Start",
    "Due",
    "Status",
    "Delivery Health",
    "Owner",
    "Priority",
    "Latest Status",
    "Created",
    "Updated",
    "InProgressSince",
    "DoneDate",
    "BlockedSince",
    "LatestUpdateOn",
)
RELATIONSHIP_COLUMNS = ("Parent", "A2", "A3", "A4", "A5", "Scope")
OBSOLETE_SCHEDULE_COLUMNS = ("EffStart", "EffDue")
FIRST_PLAN_ROW = 6


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--static-only",
        action="store_true",
        help="scan a raw build without opening desktop Excel",
    )
    return parser


def _formula_text(value: object) -> str:
    return str(getattr(value, "text", value) or "")


def _normalise(formula: object) -> str:
    text = _formula_text(formula)
    for prefix in ("_xlfn.", "_xlws.", "_xlpm."):
        text = text.replace(prefix, "")
    text = "".join(text.split()).upper()
    text = re.sub(r"\[\[#THISROW\],\[([^\]]+)\]\]", r"[@[\1]]", text)
    return re.sub(r"\[\[#THISROW\],([^\]]+)\]", r"[@\1]", text)


def _cell_formula_records(worksheet: Worksheet) -> Iterator[tuple[str, str]]:
    """Yield cell formulas with sheet coordinates.

    Yields:
        A formula context and normalized formula text.

    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" or isinstance(cell.value, ArrayFormula):
                yield f"{worksheet.title}!{cell.coordinate}", _normalise(cell.value)


def _table_formula_records(worksheet: Worksheet) -> Iterator[tuple[str, str]]:
    """Yield calculated table-column formulas.

    Yields:
        A formula context and normalized formula text.

    """
    for table in worksheet.tables.values():
        for column in table.tableColumns:
            formula = column.calculatedColumnFormula
            if formula is not None:
                yield (
                    f"{worksheet.title}!{table.name}[{column.name}]",
                    _normalise(getattr(formula, "text", formula)),
                )


def _rule_formula_records(worksheet: Worksheet) -> Iterator[tuple[str, str]]:
    """Yield conditional-format and data-validation formulas.

    Yields:
        A formula context and normalized formula text.

    """
    for target, rules in worksheet.conditional_formatting._cf_rules.items():
        for rule_index, rule in enumerate(rules, start=1):
            for formula in rule.formula or ():
                yield f"{worksheet.title}!CF:{target}:{rule_index}", _normalise(formula)
    for validation_index, validation in enumerate(
        worksheet.data_validations.dataValidation,
        start=1,
    ):
        for field in ("formula1", "formula2"):
            formula = getattr(validation, field)
            if formula:
                yield (
                    f"{worksheet.title}!DV:{validation_index}:{field}",
                    _normalise(formula),
                )


def _formula_records(workbook: Workbook) -> Iterator[tuple[str, str]]:
    """Yield every workbook formula-bearing surface.

    Yields:
        A formula context and normalized formula text.

    """
    for worksheet in workbook.worksheets:
        yield from _cell_formula_records(worksheet)
        yield from _table_formula_records(worksheet)
        yield from _rule_formula_records(worksheet)
    for name in workbook.defined_names:
        yield f"Name:{name}", _normalise(workbook.defined_names[name].value)


def _cell_formula(workbook: Workbook, sheet: str, reference: str) -> str:
    return _normalise(workbook[sheet][reference].value)


def _table_formula_map(workbook: Workbook, sheet: str, table: str) -> dict[str, str]:
    return {
        column.name: _normalise(getattr(column.calculatedColumnFormula, "text", ""))
        for column in workbook[sheet].tables[table].tableColumns
        if column.calculatedColumnFormula is not None
    }


def _require_tokens(
    failures: list[str],
    label: str,
    formula: str,
    tokens: Sequence[str],
) -> None:
    failures.extend(
        f"{label} does not use direct {token}"
        for token in tokens
        if _normalise(token) not in formula
    )


def _forbidden_adoption_tokens() -> tuple[str, ...]:
    tokens: list[str] = []
    for relationship in RELATIONSHIP_COLUMNS:
        for operational in OPERATIONAL_COLUMNS:
            tokens.extend((
                f"XLOOKUP(tblItems[{relationship}],tblItems[ID],tblItems[{operational}]",
                f"fnItemLookup([@{relationship}],tblItems[{operational}])",
                f"INDEX(tblItems[{operational}],MATCH([@{relationship}],tblItems[ID]",
                f"FILTER(tblItems[{operational}],tblItems[{relationship}]",
                f"FILTER(tblItems[{operational}],(tblItems[{relationship}]",
            ))
            tokens.extend(
                f"{aggregate}(tblItems[{operational}],tblItems[{relationship}]"
                for aggregate in ("MINIFS", "MAXIFS", "SUMIFS", "AVERAGEIFS")
            )
    return tuple(_normalise(token) for token in tokens)


def _policy_failures(
    workbook: Workbook,
    records: Sequence[tuple[str, str]],
) -> list[str]:
    failures: list[str] = []
    if workbook.sheetnames != list(SHEETS):
        failures.append(
            f"formula-policy scan covers {workbook.sheetnames!r}, want all {list(SHEETS)!r}"
        )
    items_table = workbook["Items"].tables["tblItems"]
    headers = [column.name for column in items_table.tableColumns]
    failures.extend(
        f"tblItems still contains obsolete calculated column {column}"
        for column in OBSOLETE_SCHEDULE_COLUMNS
        if column in headers
    )
    obsolete_tokens = tuple(_normalise(f"tblItems[{name}]") for name in OBSOLETE_SCHEDULE_COLUMNS)
    for token in obsolete_tokens:
        hits = [context for context, formula in records if token in formula]
        if hits:
            failures.append(f"obsolete schedule field {token} remains in {', '.join(hits[:6])}")

    adoption_tokens = _forbidden_adoption_tokens()
    adoption_hits = [
        context
        for context, formula in records
        if any(token in formula for token in adoption_tokens)
    ]
    if adoption_hits:
        failures.append(
            "operational values still travel through hierarchy relationships in "
            + ", ".join(dict.fromkeys(adoption_hits[:8]))
        )
    return failures


def _view_formula_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []

    overview = _cell_formula(workbook, "Overview", "A3")
    _require_tokens(
        failures,
        "Overview!A3",
        overview,
        (
            "FILTER(tblItems[Owner],pred",
            "FILTER(tblItems[Due],pred",
            '"Owner not set"',
            '"Due date not set"',
        ),
    )
    owner_fallback = _normalise("XLOOKUP(tblItems[Scope],tblItems[ID],tblItems[Owner]")
    if owner_fallback in overview:
        failures.append("Overview!A3 still substitutes the Level-1 scope owner")

    direct_blocked = _normalise(DIRECT_BLOCKED_HEALTH_FORMULA)
    for sheet, reference in (("Overview", "A3"), ("Calc", "AK2")):
        if direct_blocked not in _cell_formula(workbook, sheet, reference):
            failures.append(
                f"{sheet}!{reference} does not share the final-nonblank Blocked identity"
            )
    if direct_blocked not in _table_formula_map(workbook, "Items", "tblItems").get("IsBlocked", ""):
        failures.append(
            "Items!tblItems[IsBlocked] does not share the final-nonblank Blocked identity"
        )

    plan_ids = _cell_formula(workbook, "Plan", "A6")
    if any(
        _normalise(f"tblItems[{column}]") in plan_ids
        for column in (*OPERATIONAL_COLUMNS, *OBSOLETE_SCHEDULE_COLUMNS, "IsPoint")
    ):
        failures.append("Plan!A6 still hides or admits rows according to an operational value")
    _require_tokens(
        failures,
        "Plan!D6",
        _cell_formula(workbook, "Plan", "D6"),
        ("tblItems[Start]",),
    )
    _require_tokens(
        failures,
        "Plan!E6",
        _cell_formula(workbook, "Plan", "E6"),
        ("tblItems[Due]",),
    )
    _require_tokens(
        failures,
        "Plan!BI6",
        _cell_formula(workbook, "Plan", "BI6"),
        ("tblItems[Status]", "tblItems[Due]"),
    )
    _require_tokens(
        failures,
        "Plan!F5",
        _cell_formula(workbook, "Plan", "F5"),
        ("tblItems[Start]", "tblItems[Due]"),
    )
    plan_grid = _cell_formula(workbook, "Plan", "F6")
    _require_tokens(
        failures,
        "Plan!F6",
        plan_grid,
        ("tblItems[Start]", "tblItems[Due]", "tblItems[IsPoint]", '"◆"'),
    )
    expected_grid_branches = (
        'IF((ids="")+(ids="— none —"),"",',
        'IF(pt=TRUE,IF((du<>"")*(du>=wk)*(du<wk+7),"◆",""),',
        'IF((es<>"")*(du<>"")*(wk+6>=es)*(wk<=du),glyph,"")',
    )
    failures.extend(
        f"Plan!F6 does not preserve direct-date branch {_normalise(branch)}"
        for branch in expected_grid_branches
        if _normalise(branch) not in plan_grid
    )

    sort_date = _normalise(workbook.defined_names["fnSortDate"].value)
    _require_tokens(
        failures,
        "fnSortDate",
        sort_date,
        ("tblItems[Start]", "tblItems[Due]"),
    )
    wbs_key = _normalise(workbook.defined_names["fnWbsKey"].value)
    entry_order_tie = _normalise('TEXT(IFNA(XMATCH(id,tblItems[ID]),0),"00000")')
    if entry_order_tie not in wbs_key:
        failures.append("fnWbsKey does not break equal-date sibling ties by table entry order")
    return failures


def _items_due_cue_failures(
    workbook: Workbook,
    records: Sequence[tuple[str, str]],
) -> list[str]:
    failures: list[str] = []

    item_headers = {cell.value: index for index, cell in enumerate(workbook["Items"][2], start=1)}
    due_letter = get_column_letter(item_headers["Due"])
    status_letter = get_column_letter(item_headers["Status"])
    due_attention_formula = _normalise(
        f'AND(COUNTIFS(dvStatus,${status_letter}3,dvStatusActive,TRUE)>0,${due_letter}3="")'
    )
    item_cf = [formula for context, formula in records if context.startswith("Items!CF:")]
    if due_attention_formula not in item_cf:
        failures.append("Items active missing-Due cue is not the exact direct Status/Due condition")
    return failures


def _relationship_derivation_failures(workbook: Workbook) -> list[str]:
    failures: list[str] = []
    item_formulas = _table_formula_map(workbook, "Items", "tblItems")
    expected_items = {
        "ParentTitle": '=IF([@Parent]="","",fnItemLookup([@Parent],tblItems[Title]))',
        "ParentLevel": (
            '=IF([@Parent]="",0,IFNA(INDEX(tblItems[Level],MATCH([@Parent],tblItems[ID],0)),0))'
        ),
        "A2": '=IF([@Parent]="","",fnItemLookup([@Parent],tblItems[Parent]))',
        "A3": '=IF([@A2]="","",fnItemLookup([@A2],tblItems[Parent]))',
        "A4": '=IF([@A3]="","",fnItemLookup([@A3],tblItems[Parent]))',
        "A5": '=IF([@A4]="","",fnItemLookup([@A4],tblItems[Parent]))',
        "Level": '=IF([@ID]="",0,fnTypeLevel([@Type]))',
        "Scope": "=fnAncestorAtLevel([@ID],1,0)",
        "Children": '=IF([@ID]="",0,COUNTIF(tblItems[Parent],[@ID]))',
        "WaitingOn": "=fnDepOpen([@BlockedBy])",
        "BlockedRefsValid": "=fnRefsValid([@BlockedBy],[@ID])",
        "IsBlocked": (
            '=OR(AND([@[Delivery Health]]<>"",[@[Delivery Health]]='
            f'{DIRECT_BLOCKED_HEALTH_FORMULA}),[@WaitingOn]<>"")'
        ),
        "WbsKey": '=IF([@ID]="",REPT("Z",50),fnWbsKey([@ID],0))',
    }
    for column, expected in expected_items.items():
        got = item_formulas.get(column, "")
        if got != _normalise(expected).removeprefix("="):
            failures.append(f"Items!tblItems[{column}] relationship formula changed: got {got!r}")

    raid_scope = _table_formula_map(workbook, "RAID", "tblRAID").get("Scope", "")
    expected_raid_scope = _normalise(
        '=IF([@RelatedID]="","",IFNA(INDEX(tblItems[Scope],MATCH([@RelatedID],tblItems[ID],0)),""))'
    ).removeprefix("=")
    if raid_scope != expected_raid_scope:
        failures.append(f"RAID!tblRAID[Scope] relationship formula changed: got {raid_scope!r}")
    return failures


def _static_formula_failures(workbook: Workbook) -> list[str]:
    records = list(_formula_records(workbook))
    return [
        *_policy_failures(workbook, records),
        *_view_formula_failures(workbook),
        *_items_due_cue_failures(workbook, records),
        *_relationship_derivation_failures(workbook),
    ]


def _static_failures() -> list[str]:
    with TemporaryDirectory(prefix="PM_direct_item_policy.") as directory:
        out = Path(directory) / "PM_direct_item_policy.xlsx"
        build_one(out, with_vba=False)
        workbook = openpyxl.load_workbook(out, data_only=False)
        try:
            return _static_formula_failures(workbook)
        finally:
            workbook.close()


def _scenario_items(today: date) -> list[dict[str, object]]:
    common = {
        "Created": today - timedelta(days=100),
        "Updated": today - timedelta(days=1),
        "LatestUpdateOn": today - timedelta(days=1),
    }
    return [
        {
            **common,
            "ID": "I-1001",
            "Title": "Direct product",
            "Type": "Product",
            "Status": "In Progress",
            "Delivery Health": "On track",
            "Priority": "P4",
            "Owner": "Root owner",
            "Latest Status": "Root owns this narrative",
            "InProgressSince": today - timedelta(days=60),
        },
        {
            **common,
            "ID": "I-1002",
            "Title": "Blank release schedule",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Review",
            "Delivery Health": "At risk",
            "Priority": "P3",
            "Latest Status": "Release owns this narrative",
            "InProgressSince": today - timedelta(days=50),
        },
        {
            **common,
            "ID": "I-1003",
            "Title": "Direct key date",
            "Type": "Phase",
            "Parent": "I-1002",
            "Status": "Ready",
            "Due": today - timedelta(days=21),
            "Delivery Health": "Off track",
            "Priority": "P2",
            "Owner": "Phase owner",
            "Latest Status": "Phase owns this narrative",
        },
        {
            **common,
            "ID": "I-1004",
            "Title": "Direct start only",
            "Type": "Feature",
            "Parent": "I-1003",
            "Status": "In Progress",
            "Start": today - timedelta(days=35),
            "Delivery Health": "On track",
            "Priority": "P1",
            "Owner": "Feature owner",
            "Latest Status": "Feature owns this narrative",
            "InProgressSince": today - timedelta(days=35),
        },
        {
            **common,
            "ID": "I-1005",
            "Title": "Direct interval",
            "Type": "Task",
            "Parent": "I-1004",
            "Status": "In Progress",
            "Start": today - timedelta(days=28),
            "Due": today - timedelta(days=14),
            "Delivery Health": "Off track",
            "Priority": "P0",
            "Owner": "Task owner",
            "Latest Status": "Task owns this narrative",
            "InProgressSince": today - timedelta(days=28),
        },
        {
            **common,
            "ID": "I-1006",
            "Title": "Direct blocked sub task",
            "Type": "Sub Task",
            "Parent": "I-1005",
            "Status": "Ready",
            "Start": today - timedelta(days=12),
            "Due": today - timedelta(days=7),
            "Delivery Health": "Blocked",
            "Priority": "P2",
            "Owner": "Sub task owner",
            "Latest Status": "Sub task owns this narrative",
            "BlockedSince": today - timedelta(days=4),
        },
        {
            **common,
            "ID": "I-1007",
            "Title": "Fully undated sibling",
            "Type": "Release",
            "Parent": "I-1001",
            "Status": "Backlog",
            "Delivery Health": "On track",
            "Priority": "P3",
            "Owner": "Sibling owner",
            "Latest Status": "Sibling owns this narrative",
        },
    ]


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def _comparable(value: object) -> object:
    converted = _as_date(value)
    return converted if converted is not None else value


def _overview_rows(workbook: Workbook) -> dict[str, tuple[object, ...]]:
    overview = workbook["Overview"]
    rows: dict[str, tuple[object, ...]] = {}
    for row in range(3, 8):
        label = str(overview.cell(row, 1).value or "")
        for item_id in ("I-1001", "I-1002", "I-1006", "I-1007"):
            if item_id in label:
                rows[item_id] = tuple(overview.cell(row, column).value for column in range(1, 5))
    return rows


def _plan_axis_failures(plan: Worksheet, today: date) -> list[str]:
    axis = [
        _as_date(plan.cell(5, column).value)
        for column in range(6, 6 + PLAN_WEEKS)
        if _as_date(plan.cell(5, column).value) is not None
    ]
    earliest = today - timedelta(days=35)
    latest = today - timedelta(days=7)
    if not axis or axis[0] > earliest or axis[-1] + timedelta(days=6) < latest:
        return [f"Plan automatic window does not cover direct dates: {axis!r}"]
    return []


def _plan_scenario_failures(
    workbook: Workbook,
    fixture: Sequence[dict[str, object]],
    today: date,
) -> list[str]:
    failures: list[str] = []
    plan = workbook["Plan"]
    ordered_ids = [
        str(plan.cell(row, 1).value)
        for row in range(FIRST_PLAN_ROW, FIRST_PLAN_ROW + len(fixture) + 2)
        if plan.cell(row, 1).value
    ]
    expected_order = [str(item["ID"]) for item in fixture]
    if ordered_ids != expected_order:
        failures.append(
            f"Plan WBS order: got {ordered_ids!r}, want entry-order ties {expected_order!r}"
        )
    plan_rows = {
        str(plan.cell(row, 1).value): row
        for row in range(FIRST_PLAN_ROW, FIRST_PLAN_ROW + len(fixture) + 2)
        if plan.cell(row, 1).value
    }
    missing = sorted({str(item["ID"]) for item in fixture} - set(plan_rows))
    if missing:
        failures.append(f"Plan omitted in-scope items with partial/blank dates: {missing}")

    expected_dates = {
        "I-1001": (None, None),
        "I-1002": (None, None),
        "I-1003": (None, today - timedelta(days=21)),
        "I-1004": (today - timedelta(days=35), None),
        "I-1005": (today - timedelta(days=28), today - timedelta(days=14)),
        "I-1006": (today - timedelta(days=12), today - timedelta(days=7)),
        "I-1007": (None, None),
    }
    expected_categories = {
        "I-1001": "A",
        "I-1002": "A",
        "I-1003": "O",
        "I-1004": "A",
        "I-1005": "O",
        "I-1006": "O",
        "I-1007": "P",
    }
    for item_id, row in plan_rows.items():
        if item_id not in expected_dates:
            continue
        got_dates = (_as_date(plan.cell(row, 4).value), _as_date(plan.cell(row, 5).value))
        if got_dates != expected_dates[item_id]:
            failures.append(
                f"Plan {item_id} dates: got {got_dates!r}, want {expected_dates[item_id]!r}"
            )
        category = plan.cell(row, 61).value
        if category != expected_categories[item_id]:
            failures.append(
                f"Plan {item_id} category: got {category!r}, want {expected_categories[item_id]!r}"
            )
        timeline = [plan.cell(row, column).value for column in range(6, 6 + PLAN_WEEKS)]
        marks = [value for value in timeline if value not in {None, ""}]
        if item_id == "I-1003" and marks != ["◆"]:
            failures.append(f"Plan direct due-only key date marks: got {marks!r}, want ['◆']")
        if item_id in {"I-1001", "I-1002", "I-1004", "I-1007"} and marks:
            failures.append(
                f"Plan {item_id} drew a timeline mark without both direct dates: {marks!r}"
            )
        if item_id in {"I-1005", "I-1006"} and (not marks or set(marks) != {"!"}):
            failures.append(
                f"Plan {item_id} direct interval marks: got {marks!r}, want overdue interval"
            )

    failures.extend(_plan_axis_failures(plan, today))
    return failures


def _overview_scenario_failures(workbook: Workbook, today: date) -> list[str]:
    failures: list[str] = []
    blocked_due = today - timedelta(days=7)
    expected = {
        "I-1001": ("On track", "Root owner", "Due date not set"),
        "I-1002": ("At risk", "Owner not set", "Due date not set"),
        "I-1006": (
            "Blocked",
            "Sub task owner",
            f"{blocked_due.day} {blocked_due:%b %Y}",
        ),
        "I-1007": ("On track", "Sibling owner", "Due date not set"),
    }
    rows = _overview_rows(workbook)
    for item_id, wanted in expected.items():
        row = rows.get(item_id)
        if row is None:
            failures.append(f"Overview omitted direct item {item_id}")
        elif row[1:4] != wanted:
            failures.append(f"Overview {item_id} direct values: got {row[1:4]!r}, want {wanted!r}")
    return failures


def _items_scenario_failures(
    workbook: Workbook,
    fixture: Sequence[dict[str, object]],
) -> list[str]:
    failures: list[str] = []
    items = workbook["Items"]
    columns = {cell.value: index for index, cell in enumerate(items[2], start=1)}
    rows = {str(items.cell(row, columns["ID"]).value): row for row in range(3, 3 + len(fixture))}
    for expected in fixture:
        item_id = str(expected["ID"])
        row = rows[item_id]
        for column in OPERATIONAL_COLUMNS:
            got = _comparable(items.cell(row, columns[column]).value)
            wanted = _comparable(expected.get(column))
            if got != wanted:
                failures.append(
                    f"Items {item_id} {column}: got {got!r}, want direct value {wanted!r}"
                )
    return failures


def _ancestor_surface_snapshot(workbook: Workbook) -> dict[str, tuple[object, ...]]:
    ancestor_ids = ("I-1001", "I-1002")
    snapshot: dict[str, tuple[object, ...]] = {}

    items = workbook["Items"]
    item_columns = {cell.value: index for index, cell in enumerate(items[2], start=1)}
    item_rows = {
        str(items.cell(row, item_columns["ID"]).value): row
        for row in range(3, 10)
        if items.cell(row, item_columns["ID"]).value
    }
    for item_id in ancestor_ids:
        row = item_rows[item_id]
        snapshot[f"Items:{item_id}"] = tuple(
            _comparable(items.cell(row, item_columns[column]).value)
            for column in OPERATIONAL_COLUMNS
        )

    plan = workbook["Plan"]
    plan_rows = {
        str(plan.cell(row, 1).value): row
        for row in range(FIRST_PLAN_ROW, FIRST_PLAN_ROW + 9)
        if plan.cell(row, 1).value
    }
    for item_id in ancestor_ids:
        row = plan_rows.get(item_id)
        snapshot[f"Plan:{item_id}"] = (
            ()
            if row is None
            else tuple(_comparable(plan.cell(row, column).value) for column in (3, 4, 5, 61))
        )

    overview_rows = _overview_rows(workbook)
    for item_id in ancestor_ids:
        snapshot[f"Overview:{item_id}"] = tuple(
            _comparable(value) for value in overview_rows.get(item_id, ())
        )
    return snapshot


def _item_cell_reference(workbook: Workbook, item_id: str, column: str) -> str:
    items = workbook["Items"]
    columns = {cell.value: index for index, cell in enumerate(items[2], start=1)}
    id_column = columns["ID"]
    row = next(
        row
        for row in range(3, items.max_row + 1)
        if str(items.cell(row, id_column).value or "") == item_id
    )
    return f"{get_column_letter(columns[column])}{row}"


def _scenario_failures() -> list[str]:
    today = datetime.now(tz=UTC).date()
    fixture = _scenario_items(today)
    with temporary_examples() as examples, temporary_workbook("PM_direct_item_values") as out:
        examples.ITEMS_EXAMPLES = fixture
        examples.PEOPLE_EXAMPLES = [
            {"Person": owner, "Role": "Owner", "Team": "Core"}
            for owner in (
                "Root owner",
                "Phase owner",
                "Feature owner",
                "Task owner",
                "Sub task owner",
                "Sibling owner",
                "Changed descendant owner",
            )
        ]
        examples.RAID_EXAMPLES = []
        build_one(out, with_vba=False)
        setting_row = next(
            4 + index
            for index, setting in enumerate(config.SETTINGS)
            if setting[0] == "cfgExecutiveStatusMaxLevel"
        )
        recalculate(out, sheet="Config", cell=f"B{setting_row}", value="2")

        workbook = openpyxl.load_workbook(out, data_only=True)
        try:
            failures = [
                *(f"calculated error {error}" for error in workbook_error_cells(workbook)),
                *_items_scenario_failures(workbook, fixture),
                *_plan_scenario_failures(workbook, fixture, today),
                *_overview_scenario_failures(workbook, today),
            ]
            ancestor_before = _ancestor_surface_snapshot(workbook)
            changed_owner_cell = _item_cell_reference(workbook, "I-1006", "Owner")
        finally:
            workbook.close()

        recalculate(
            out,
            sheet="Items",
            cell=changed_owner_cell,
            value="Changed descendant owner",
        )
        changed = openpyxl.load_workbook(out, data_only=True)
        try:
            ancestor_after = _ancestor_surface_snapshot(changed)
            if ancestor_after != ancestor_before:
                failures.append(
                    "changing a descendant Owner changed an ancestor reporting surface: "
                    f"before={ancestor_before!r}, after={ancestor_after!r}"
                )
            plan = changed["Plan"]
            changed_row = next(
                (
                    row
                    for row in range(FIRST_PLAN_ROW, FIRST_PLAN_ROW + len(fixture) + 2)
                    if plan.cell(row, 1).value == "I-1006"
                ),
                None,
            )
            if changed_row is None:
                failures.append("descendant mutation row I-1006 disappeared from Plan")
            elif plan.cell(changed_row, 3).value != "Changed descendant owner":
                failures.append("descendant mutation did not reach its own direct Plan Owner")
        finally:
            changed.close()
        return failures


def main(argv: Sequence[str] | None = None) -> int:
    """Run the static policy and optional desktop-Excel regression scenario.

    Returns:
        A process exit status: zero for a clean workbook, one for failures.

    """
    args = _parser().parse_args(argv)
    failures = _static_failures()
    if not args.static_only and not failures:
        failures.extend(_scenario_failures())
    if failures:
        LOGGER.error("DIRECT-ITEM VALUE QA: %s FAILURE(S)", len(failures))
        for failure in failures:
            LOGGER.error("  FAIL %s", failure)
        return 1
    LOGGER.info("DIRECT-ITEM VALUE QA: ALL PASS")
    return 0


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    raise SystemExit(main())
