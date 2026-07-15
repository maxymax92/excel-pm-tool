"""Build the deterministic, populated demonstration workbook.

The scenario replaces example and Config data within this Python process,
builds one macro-enabled workbook, restores module state and runs package-level
structural checks. Release artifacts are left untouched.

Run from the repository root:

    .venv/bin/python -m build.scenarios.ship_demo

Output:

    dist/PM_Workbook_Ship_Demo.xlsm

The fixed 13 Jul 2026 anchor makes the scenario reproducible.  The data is
purpose-built to exercise the hierarchy, Overview panels, Plan states,
validation surfaces and conditional-format states without seeding invalid
structural data.
"""

from __future__ import annotations

import copy
import logging
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, timedelta
from enum import Enum
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.utils.cell import range_boundaries

from .. import pipeline
from ..data.inject import swapped_module_state
from ..paths import DIST
from ..spec import config, examples
from ..spec import items as item_spec

if TYPE_CHECKING:
    from collections.abc import Iterator
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

AS_OF = date(2026, 7, 13)
OUTPUT = DIST / "PM_Workbook_Ship_Demo.xlsm"
LOGGER = logging.getLogger(__name__)
EXPECTED_ITEM_COUNT = 30
EXPECTED_RAID_COUNT = 9
MIN_KEY_DATE_LEVEL = 2
EXPECTED_HIERARCHY_LEVELS = 6
EXPECTED_RECENT_COUNT = 6


class _ShipDemoProblem(Enum):
    QA = "ship-demo QA: {}"
    RESTORE = "demonstration build failed and source restoration also failed: {}: {}"


class ShipDemoError(RuntimeError):
    """Report invalid demonstration data or failed source restoration."""

    def __init__(self, problem: _ShipDemoProblem, *details: object) -> None:
        """Create a stable demonstration diagnostic."""
        super().__init__(problem.value.format(*details))


def d(offset: int) -> date:
    """Return a deterministic scenario date at the requested offset.

    Returns:
        The anchor date plus ``offset`` days.

    """
    return AS_OF + timedelta(days=offset)


@dataclass(frozen=True, kw_only=True, slots=True)
class ItemSeed:
    """Typed source values for one populated demonstration item."""

    item_id: str
    title: str
    item_type: str
    status: str
    priority: str
    owner: str
    latest: str
    parent: str = ""
    start: int | None = None
    due: int | None = None
    delivery_health: str = ""
    blocked_by: str = ""
    blocked_since: int | None = None
    updated: int = -1
    latest_on: int | None = -1
    done: int | None = None


def _item(seed: ItemSeed) -> dict[str, object]:
    row = {
        "ID": seed.item_id,
        "Title": seed.title,
        "Type": seed.item_type,
        "Status": seed.status,
        "Priority": seed.priority,
        "Owner": seed.owner,
        "Latest Status": seed.latest,
        "Created": d(-90),
        "Updated": d(seed.updated),
    }
    if seed.parent:
        row["Parent"] = seed.parent
    if seed.start is not None:
        row["Start"] = d(seed.start)
    if seed.due is not None:
        row["Due"] = d(seed.due)
    if seed.delivery_health:
        row["Delivery Health"] = seed.delivery_health
    if seed.blocked_by:
        row["BlockedBy"] = seed.blocked_by
    if seed.blocked_since is not None:
        row["BlockedSince"] = d(seed.blocked_since)
    if seed.latest_on is not None:
        row["LatestUpdateOn"] = d(seed.latest_on)
    if seed.status in {"In Progress", "Review"} and seed.start is not None:
        row["InProgressSince"] = d(seed.start)
    if seed.done is not None:
        row["DoneDate"] = d(seed.done)
        row["Updated"] = d(seed.done)
        row["LatestUpdateOn"] = d(seed.done)
    return row


ITEMS = [
    # Aurora: full six-level branch plus every schedule/state edge.
    _item(
        ItemSeed(
            item_id="I-1001",
            title="Aurora pricing pilot",
            item_type="Project",
            status="In Progress",
            priority="P1",
            owner="Maya Chen",
            latest="Pilot build is moving; supplier security is the main risk.",
            start=-28,
            due=60,
            delivery_health="At risk",
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1002",
            title="Pilot rollout",
            item_type="Initiative",
            status="In Progress",
            priority="P1",
            owner="Priya Shah",
            latest="Pilot governance is set; launch controls are being closed.",
            parent="I-1001",
            start=-21,
            due=45,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1003",
            title="Build and validate",
            item_type="Phase",
            status="In Progress",
            priority="P1",
            owner="Priya Shah",
            latest="Build is in validation with no new scope change.",
            parent="I-1002",
            start=-21,
            due=30,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1004",
            title="Delivery squad",
            item_type="Team",
            status="In Progress",
            priority="P2",
            owner="Priya Shah",
            latest="Squad capacity is stable for the delivery window.",
            parent="I-1003",
            start=-20,
            due=28,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1005",
            title="Approval workflow",
            item_type="Feature",
            status="Review",
            priority="P1",
            owner="Elliot Hughes",
            latest="Workflow is in review with Legal and Operations.",
            parent="I-1003",
            start=-14,
            due=6,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1006",
            title="Stakeholder approval rules",
            item_type="Story",
            status="Review",
            priority="P1",
            owner="Elliot Hughes",
            latest="Approval rules are ready for final stakeholder review.",
            parent="I-1005",
            start=-10,
            due=3,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1007",
            title="Duplicate approval event",
            item_type="Bug",
            status="In Progress",
            priority="P0",
            owner="Luca Martin",
            latest="Duplicate event remains blocked by webhook credentials.",
            parent="I-1006",
            start=-12,
            due=-1,
            blocked_by="I-1008",
            latest_on=-9,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1008",
            title="Supplier webhook credentials",
            item_type="Sub Task",
            status="Backlog",
            priority="P0",
            owner="Luca Martin",
            latest="Supplier security approval is pending.",
            parent="I-1006",
            start=1,
            due=5,
            delivery_health="Blocked",
            blocked_since=-4,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1009",
            title="Pilot go/no-go",
            item_type="Release",
            status="Ready",
            priority="P0",
            owner="Maya Chen",
            latest="Steering pack is ready for the go/no-go decision.",
            parent="I-1001",
            due=6,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1010",
            title="Data migration rehearsal",
            item_type="Phase",
            status="Ready",
            priority="P1",
            owner="Nia Okafor",
            latest="Rehearsal data set is ready for validation.",
            parent="I-1002",
            due=18,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1011",
            title="Wave 1 launch",
            item_type="Release",
            status="Ready",
            priority="P0",
            owner="Maya Chen",
            latest="Wave 1 remains scheduled after pilot acceptance.",
            parent="I-1001",
            due=75,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1012",
            title="UAT regression pack",
            item_type="Test Case",
            status="Done",
            priority="P2",
            owner="Theo Grant",
            latest="Regression pack passed with no release blockers.",
            parent="I-1005",
            start=-12,
            due=-3,
            done=-3,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1013",
            title="Decision pack approved",
            item_type="Deliverable",
            status="Done",
            priority="P2",
            owner="Priya Shah",
            latest="Decision pack was approved and distributed.",
            parent="I-1003",
            start=-20,
            due=-8,
            done=-8,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1014",
            title="Contingency route",
            item_type="Release",
            status="Cancelled",
            priority="P4",
            owner="Maya Chen",
            latest="Primary-route approval closed the contingency route.",
            parent="I-1001",
            start=-15,
            due=14,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-1015",
            title="Support handover",
            item_type="Task",
            status="In Progress",
            priority="P3",
            owner="",
            latest="Owner and due date are being confirmed.",
            parent="I-1002",
            start=-2,
        )
    ),
    # Atlas exercises a healthy Product scope, planned work, narrative aging and points.
    _item(
        ItemSeed(
            item_id="I-2001",
            title="Atlas client portal",
            item_type="Product",
            status="In Progress",
            priority="P2",
            owner="Elliot Hughes",
            latest="Onboarding release remains on plan.",
            start=-35,
            due=60,
            delivery_health="On track",
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2002",
            title="Onboarding release",
            item_type="Release",
            status="In Progress",
            priority="P1",
            owner="Elliot Hughes",
            latest="Release scope is stable and delivery is tracking.",
            parent="I-2001",
            start=-15,
            due=45,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2003",
            title="Accessibility improvements",
            item_type="Epic",
            status="In Progress",
            priority="P1",
            owner="Elliot Hughes",
            latest="Accessibility work is progressing through review.",
            parent="I-2002",
            start=-10,
            due=20,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2004",
            title="Keyboard navigation",
            item_type="Story",
            status="Done",
            priority="P1",
            owner="Luca Martin",
            latest="Keyboard navigation passed acceptance testing.",
            parent="I-2003",
            start=-10,
            due=-1,
            done=-1,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2005",
            title="Screen-reader labels",
            item_type="Task",
            status="Review",
            priority="P1",
            owner="Luca Martin",
            latest="Labels are in review; final audit evidence is pending.",
            parent="I-2003",
            start=-7,
            due=4,
            latest_on=-8,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2006",
            title="Contrast regression",
            item_type="Bug",
            status="Done",
            priority="P1",
            owner="Luca Martin",
            latest="Contrast checks passed verification.",
            parent="I-2005",
            start=-5,
            due=-2,
            done=-2,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2007",
            title="Partner data contract",
            item_type="Deliverable",
            status="Backlog",
            priority="P2",
            owner="Nia Okafor",
            latest="Contract drafting starts after onboarding sign-off.",
            parent="I-2002",
            start=10,
            due=40,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2008",
            title="Data quality checks",
            item_type="Test Case",
            status="Backlog",
            priority="P4",
            owner="Theo Grant",
            latest="Test design is queued behind the data contract.",
            parent="I-2007",
            start=15,
            due=35,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2009",
            title="Content sign-off",
            item_type="Story",
            status="Ready",
            priority="P2",
            owner="Elliot Hughes",
            latest="Content is ready for business sign-off.",
            parent="I-2003",
            due=12,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-2010",
            title="Design approval",
            item_type="Deliverable",
            status="Done",
            priority="P2",
            owner="Elliot Hughes",
            latest="Design approval was recorded by the sponsor.",
            parent="I-2002",
            due=-5,
            done=-5,
        )
    ),
    # Beacon exercises off-track delivery health and narrative aging with valid structure.
    _item(
        ItemSeed(
            item_id="I-3001",
            title="Beacon data foundation",
            item_type="Project",
            status="In Progress",
            priority="P0",
            owner="Nia Okafor",
            latest="Recovery plan is awaiting sponsor confirmation.",
            start=-50,
            due=30,
            delivery_health="Off track",
            updated=-10,
            latest_on=-10,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-3002",
            title="Migration readiness",
            item_type="Initiative",
            status="In Progress",
            priority="P0",
            owner="Nia Okafor",
            latest="Readiness remains constrained by reconciliation defects.",
            parent="I-3001",
            start=-40,
            due=20,
            updated=-10,
            latest_on=-10,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-3003",
            title="Reconciliation engine",
            item_type="Feature",
            status="In Progress",
            priority="P0",
            owner="Nia Okafor",
            latest="Engine output still requires manual reconciliation.",
            parent="I-3002",
            start=-30,
            due=10,
            updated=-10,
            latest_on=-10,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-3004",
            title="Reconcile customer balances",
            item_type="Task",
            status="In Progress",
            priority="P0",
            owner="Nia Okafor",
            latest="Balance exceptions remain above the agreed tolerance.",
            parent="I-3003",
            start=-15,
            due=-4,
            updated=-10,
            latest_on=-10,
        )
    ),
    _item(
        ItemSeed(
            item_id="I-3005",
            title="Migration smoke test",
            item_type="Test Case",
            status="Done",
            priority="P2",
            owner="Theo Grant",
            latest="Smoke test completed with follow-up defects logged.",
            parent="I-3003",
            start=-20,
            due=-12,
            done=-12,
        )
    ),
]


PEOPLE = [
    {"Person": "Maya Chen", "Role": "Programme Lead", "Team": "Transformation"},
    {"Person": "Elliot Hughes", "Role": "Product Manager", "Team": "Product"},
    {"Person": "Priya Shah", "Role": "Delivery Lead", "Team": "Delivery"},
    {"Person": "Luca Martin", "Role": "Engineering Lead", "Team": "Engineering"},
    {"Person": "Nia Okafor", "Role": "Data Lead", "Team": "Data"},
    {"Person": "Theo Grant", "Role": "QA Lead", "Team": "Quality"},
]

TEAMS = ["Transformation", "Product", "Delivery", "Engineering", "Data", "Quality"]


RAID = [
    {
        "RaidID": "R-001",
        "Type": "Issue",
        "Title": "Production-like environment unstable",
        "Detail": "Intermittent resets are disrupting end-to-end approval testing.",
        "RelatedID": "I-1007",
        "Owner": "Luca Martin",
        "Status": "Open",
        "Prob": 5,
        "Impact": 4,
        "Response": "Stabilise the environment and run a daily verification check.",
        "NextReview": d(1),
        "Raised": d(-5),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-002",
        "Type": "Risk",
        "Title": "Supplier approval may miss pilot",
        "Detail": "Security approval for the webhook credentials is behind plan.",
        "RelatedID": "I-1008",
        "Owner": "Priya Shah",
        "Status": "Monitoring",
        "Prob": 4,
        "Impact": 4,
        "Response": "Escalate approval and retain a controlled contingency credential path.",
        "NextReview": d(-1),
        "Raised": d(-20),
        "Updated": d(-2),
    },
    {
        "RaidID": "R-003",
        "Type": "Dependency",
        "Title": "Identity platform capacity window",
        "Detail": "The accessibility release depends on an identity-platform change window.",
        "RelatedID": "I-2005",
        "Owner": "Luca Martin",
        "Status": "Open",
        "Prob": 3,
        "Impact": 4,
        "Response": "Reserve the change window and confirm rollback ownership.",
        "NextReview": d(5),
        "Raised": d(-8),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-004",
        "Type": "Risk",
        "Title": "Cutover reconciliation underestimated",
        "Detail": "Manual exception handling may exceed the cutover support capacity.",
        "RelatedID": "I-3004",
        "Owner": "Nia Okafor",
        "Status": "Open",
        "Prob": 3,
        "Impact": 3,
        "Response": "Reduce exception volume and add a second reconciliation shift.",
        "NextReview": d(8),
        "Raised": d(-7),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-005",
        "Type": "Assumption",
        "Title": "Pilot volumes remain below 500 cases",
        "Detail": "The operating model assumes a controlled first-wave volume.",
        "RelatedID": "I-1001",
        "Owner": "Maya Chen",
        "Status": "Monitoring",
        "Prob": 2,
        "Impact": 3,
        "Response": "Validate weekly against intake and activate throttling if required.",
        "NextReview": d(14),
        "Raised": d(-12),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-006",
        "Type": "Decision",
        "Title": "Approve pilot scope and controls",
        "Detail": "Steering approval is required before the pilot can be released.",
        "RelatedID": "I-1001",
        "Owner": "Maya Chen",
        "Status": "Open",
        "Response": "Approve the proposed scope, controls and named exception owners.",
        "NextReview": d(2),
        "Raised": d(-7),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-007",
        "Type": "Decision",
        "Title": "Approve market rollout",
        "Detail": "The sponsor must approve expansion after pilot evidence is reviewed.",
        "RelatedID": "I-2001",
        "Owner": "Elliot Hughes",
        "Status": "Open",
        "Response": "Review adoption, quality and support evidence before approval.",
        "NextReview": d(45),
        "Raised": d(-2),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-008",
        "Type": "Risk",
        "Title": "Minor training attendance variance",
        "Detail": "A small group may miss the first live training session.",
        "RelatedID": "I-2007",
        "Owner": "Theo Grant",
        "Status": "Open",
        "Prob": 1,
        "Impact": 2,
        "Response": "Track attendance and publish the recorded session.",
        "NextReview": d(20),
        "Raised": d(-3),
        "Updated": d(-1),
    },
    {
        "RaidID": "R-009",
        "Type": "Issue",
        "Title": "Export delimiter issue",
        "Detail": "A delimiter issue affected one extract during rehearsal.",
        "RelatedID": "I-3005",
        "Owner": "Nia Okafor",
        "Status": "Closed",
        "Prob": 5,
        "Impact": 5,
        "Response": "Patch deployed; corrected extract reconciled successfully.",
        "NextReview": d(-6),
        "Raised": d(-25),
        "Closed": d(-5),
        "Updated": d(-5),
    },
]


SETTING_OVERRIDES = {
    "cfgDueSoonDays": 5,
    "cfgBlockedRedDays": 3,
    "cfgStaleDays": 7,
    "cfgReportDays": 14,
    "cfgExecutiveStatusMaxLevel": 1,
    "cfgKeyDateMaxLevel": 4,
    "cfgComingUrgentDays": 3,
    "cfgComingSoonDays": 7,
    "cfgComingNearDays": 30,
    "cfgComingHorizonDays": 60,
    "cfgAlertSevScore": 9,
    "cfgNextItemID": 3006,
    "cfgNextRaidID": 10,
}


def _require(condition: object, message: str) -> None:
    if not condition:
        raise ShipDemoError(_ShipDemoProblem.QA, message)


def _patched_settings() -> list[tuple[str, object, str]]:
    seen: set[str] = set()
    rows: list[tuple[str, object, str]] = []
    for name, default_value, description in config.SETTINGS:
        setting_value = default_value
        if name in SETTING_OVERRIDES:
            setting_value = SETTING_OVERRIDES[name]
            seen.add(name)
        rows.append((name, setting_value, description))
    missing = set(SETTING_OVERRIDES) - seen
    _require(not missing, f"settings are unavailable: {sorted(missing)}")
    return rows


@contextmanager
def demo_source() -> Iterator[None]:
    """Patch imported data/config objects only for the duration of one build.

    Yields:
        Nothing; the demonstration state lives for the context body.

    """
    replacements: dict[tuple[object, str], object] = {
        (examples, "ITEMS_EXAMPLES"): copy.deepcopy(ITEMS),
        (examples, "PEOPLE_EXAMPLES"): copy.deepcopy(PEOPLE),
        (examples, "RAID_EXAMPLES"): copy.deepcopy(RAID),
        (config, "TEAMS"): list(TEAMS),
        (config, "SETTINGS"): _patched_settings(),
    }

    def _restore_error(cleanup_error: BaseException) -> Exception:
        return ShipDemoError(
            _ShipDemoProblem.RESTORE,
            type(cleanup_error).__name__,
            cleanup_error,
        )

    with swapped_module_state(replacements, _restore_error):
        yield


def _source_qa() -> None:
    item_ids = [row["ID"] for row in ITEMS]
    raid_ids = [row["RaidID"] for row in RAID]
    _require(
        len(ITEMS) == EXPECTED_ITEM_COUNT,
        f"expected {EXPECTED_ITEM_COUNT} Items rows, got {len(ITEMS)}",
    )
    _require(
        len(RAID) == EXPECTED_RAID_COUNT,
        f"expected {EXPECTED_RAID_COUNT} RAID rows, got {len(RAID)}",
    )
    _require(len(item_ids) == len(set(item_ids)), "duplicate Item IDs")
    _require(len(raid_ids) == len(set(raid_ids)), "duplicate RAID IDs")
    _require(
        all("EXAMPLE" not in str(row).upper() for row in ITEMS + RAID),
        "example marker leaked into the populated copy",
    )

    levels = dict(config.TYPES)
    by_id = {row["ID"]: row for row in ITEMS}
    for row in ITEMS:
        _require(row["Type"] in levels, f"{row['ID']} uses unknown type {row['Type']!r}")
        parent = row.get("Parent", "")
        if parent:
            _require(parent in by_id, f"{row['ID']} has unknown parent {parent}")
            _require(
                levels[by_id[parent]["Type"]] < levels[row["Type"]],
                f"{row['ID']} parent is not at a higher hierarchy level",
            )
        else:
            _require(levels[row["Type"]] == 1, f"non-Level-1 row {row['ID']} has no parent")

    _require(
        {levels[row["Type"]] for row in ITEMS} == set(range(1, EXPECTED_HIERARCHY_LEVELS + 1)),
        "all six hierarchy levels are not represented",
    )
    _require(
        {row["Type"] for row in ITEMS} == set(levels),
        "all configured item types are not represented",
    )
    _require(
        {row["Status"] for row in ITEMS} == {status[0] for status in config.STATUSES},
        "all configured item statuses are not represented",
    )
    _require(
        {row["Priority"] for row in ITEMS} == set(config.PRIORITIES),
        "all configured priorities are not represented",
    )
    _require(
        {row.get("Delivery Health") for row in ITEMS if row.get("Delivery Health")}
        == set(config.DELIVERY_HEALTH),
        "all delivery-health states are not represented",
    )
    _require(any(row.get("BlockedBy") for row in ITEMS), "dependency blocker is not represented")

    raid_type_roles = {name: (alert, decision) for name, alert, decision in config.RAID_TYPES}
    closed_statuses = {name for name, closed in config.RAID_STATUSES if closed}
    _require(
        {row["Type"] for row in RAID} == set(raid_type_roles),
        "all configured RAID types are not represented",
    )
    _require(
        {row["Status"] for row in RAID} == {status[0] for status in config.RAID_STATUSES},
        "all configured RAID statuses are not represented",
    )

    top_raid = [
        row["RaidID"]
        for row in RAID
        if row["Status"] not in closed_statuses
        and raid_type_roles[row["Type"]][0]
        and row.get("Prob")
        and row.get("Impact")
        and row["Prob"] * row["Impact"] >= SETTING_OVERRIDES["cfgAlertSevScore"]
    ]
    _require(
        top_raid == ["R-001", "R-002", "R-003", "R-004"],
        f"unexpected Top RAID source set: {top_raid}",
    )

    done_statuses = {name for name, _active, done, _cancelled in config.STATUSES if done}
    cancelled = {name for name, _active, _done, is_cancelled in config.STATUSES if is_cancelled}
    item_events = [
        (row["Due"], row["ID"])
        for row in ITEMS
        if row.get("Due")
        and not row.get("Start")
        and MIN_KEY_DATE_LEVEL <= levels[row["Type"]] <= SETTING_OVERRIDES["cfgKeyDateMaxLevel"]
        and row["Due"] >= AS_OF
        and row["Status"] not in done_statuses
        and row["Status"] not in cancelled
    ]
    decision_events = [
        (row["NextReview"], row["RaidID"])
        for row in RAID
        if row["Status"] not in closed_statuses
        and raid_type_roles[row["Type"]][1]
        and row.get("NextReview")
        and row["NextReview"] >= AS_OF
    ]
    coming = [key for _date, key in sorted(item_events + decision_events)]
    _require(
        coming == ["R-006", "I-1009", "I-1010", "R-007", "I-1011"],
        f"unexpected Coming Up source order: {coming}",
    )

    recent = [
        row["ID"]
        for row in ITEMS
        if row.get("DoneDate") and row["DoneDate"] >= d(-14) and row["Status"] not in cancelled
    ]
    _require(
        len(recent) == EXPECTED_RECENT_COUNT,
        f"expected {EXPECTED_RECENT_COUNT} recent completions, got {recent}",
    )
    _require(
        sum(1 for row in ITEMS if row.get("Start") and not row.get("Due")) == 1,
        "Plan missing-date disclosure should have exactly one source row",
    )


def _table_values(
    worksheet: Worksheet,
    table_name: str,
) -> tuple[list[str], list[list[object]]]:
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [str(worksheet.cell(min_row, column).value) for column in range(min_col, max_col + 1)]
    rows = [
        [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        for row in range(min_row + 1, max_row + 1)
    ]
    return headers, rows


def _formula_column_qa(
    worksheet: Worksheet,
    table_name: str,
    specs: list[item_spec.ColumnSpec],
) -> None:
    headers, rows = _table_values(worksheet, table_name)
    for spec in specs:
        if spec["kind"] != "F":
            continue
        col = headers.index(spec["name"])
        formulas = [row[col] for row in rows]
        _require(
            all(isinstance(value, str) and value.startswith("=") for value in formulas),
            f"{table_name}[{spec['name']}] contains a non-scalar formula cell",
        )
        _require(len(set(formulas)) == 1, f"{table_name}[{spec['name']}] has inconsistent formulas")


def _package_qa(path: Path) -> None:
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    _require(
        wb.sheetnames == pipeline.SHEETS,
        f"sheet order is {wb.sheetnames}, expected {pipeline.SHEETS}",
    )

    item_headers, item_rows = _table_values(wb["Items"], "tblItems")
    raid_headers, raid_rows = _table_values(wb["RAID"], "tblRAID")
    _require(
        len(item_rows) == len(ITEMS), f"tblItems has {len(item_rows)} rows, expected {len(ITEMS)}"
    )
    _require(
        len(raid_rows) == len(RAID), f"tblRAID has {len(raid_rows)} rows, expected {len(RAID)}"
    )
    _require(
        [row[item_headers.index("ID")] for row in item_rows] == [row["ID"] for row in ITEMS],
        "Items order/data changed in package",
    )
    _require(
        [row[raid_headers.index("RaidID")] for row in raid_rows] == [row["RaidID"] for row in RAID],
        "RAID order/data changed in package",
    )

    _formula_column_qa(wb["Items"], "tblItems", item_spec.ITEMS_COLUMNS)
    _formula_column_qa(wb["RAID"], "tblRAID", item_spec.RAID_COLUMNS)

    config = wb["Config"]
    labels = {
        config.cell(row, 1).value: config.cell(row, 2).value
        for row in range(1, config.max_row + 1)
        if config.cell(row, 1).value
    }
    for name, expected in SETTING_OVERRIDES.items():
        label = name.removeprefix("cfg")
        _require(
            labels.get(label) == expected,
            f"Config {label} is {labels.get(label)!r}, expected {expected!r}",
        )
    people_headers, people_rows = _table_values(config, "tblPeople")
    person_col = people_headers.index("Person")
    _require(
        [row[person_col] for row in people_rows] == [row["Person"] for row in PEOPLE],
        "Config people were not seeded",
    )

    with zipfile.ZipFile(path) as package:
        names = set(package.namelist())
        _require("xl/vbaProject.bin" in names, "macro binary is not embedded")
        drawing_xml = "\n".join(
            package.read(name).decode("utf-8")
            for name in sorted(names)
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        )
        _require(
            'macro="[0]!ExportMarkdown"' in drawing_xml, "ExportMarkdown action is not attached"
        )
        _require('macro="[0]!OrganiseItems"' in drawing_xml, "OrganiseItems action is not attached")
        table_xml = "\n".join(
            package.read(name).decode("utf-8")
            for name in sorted(names)
            if name.startswith("xl/tables/table") and name.endswith(".xml")
        )
        expected_formula_columns = sum(
            spec["kind"] == "F" for spec in item_spec.ITEMS_COLUMNS
        ) + sum(spec["kind"] == "F" for spec in item_spec.RAID_COLUMNS)
        _require(
            table_xml.count("<calculatedColumnFormula") == expected_formula_columns,
            "calculated-column metadata count changed",
        )


def main() -> None:
    """Build and verify the deterministic populated demonstration workbook."""
    _source_qa()
    pipeline.require_current_vba()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with demo_source():
        pipeline.build_one(OUTPUT, with_vba=True)
    _package_qa(OUTPUT)
    LOGGER.info(
        "SHIP DEMO: PASS (%s items, %s RAID, %s people)",
        len(ITEMS),
        len(RAID),
        len(PEOPLE),
    )
    LOGGER.info("%s", OUTPUT)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    main()
